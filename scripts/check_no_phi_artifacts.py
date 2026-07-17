#!/usr/bin/env python3
"""
Block runtime artifacts and PHI-bearing data files from entering the repository.

The existing privacy gate (``git_hook_privacy_checks.py``) inspects *Python source*
for PII in log lines and docs. It has no concept of committed **data files**, which
is how three ``.pytest-tmp-*/dicom_viewer_config_test_signals.json`` files -- each
holding a real ``recent_files`` list with local drive paths and an identifiable
filename -- were committed in May 2026 and survived in history. A .gitignore rule
does not help: it does not untrack what is already in the index.

This checker closes that gap with two independent rules:

  1. **Path denylist.** Runtime artifacts that capture user state (pytest temp dirs,
     the app's own config JSON, generated reports) may never be tracked, wherever
     they appear.
  2. **Content scan.** Data files are searched for PHI/PII indicators: home-directory
     absolute paths, the app config's ``recent_files``/``last_path`` keys, and
     populated DICOM patient tags.

Usage:
    python scripts/check_no_phi_artifacts.py            # scan all tracked files (CI)
    python scripts/check_no_phi_artifacts.py --staged   # scan staged files (pre-commit)

Exit code: 0 clean, 1 if anything is flagged.
"""

from __future__ import annotations

import argparse
import bz2
import getpass
import gzip
import hashlib
import io
import ipaddress
import json
import lzma
import os
import re
import socket
import subprocess
import sys
import tarfile
import zipfile
from contextlib import ExitStack
from pathlib import Path
from typing import Any, BinaryIO, Protocol

# --- Rule 1: paths that must never be tracked -------------------------------
# Matched with Path.match / prefix semantics against the repo-relative posix path.
FORBIDDEN_PATH_PATTERNS: list[tuple[str, str]] = [
    (r"^\.pytest-tmp", "pytest temp dir: captures live app config (recent_files)"),
    (r"^pytest-of-", "pytest temp dir"),
    (r"(^|/)\.pytest_cache/", "pytest cache"),
    (
        r"(^|/)dicom_viewer_config.*\.json$",
        "app config: contains recent_files / local paths",
    ),
    (r"^pyright-report\.txt$", "generated type-check report (absolute local paths)"),
    (r"^backups/", "local source backups: not for version control"),
    (r"^data/(?!\.gitkeep$)", "local/imported data directory"),
    (r"^decoder-spike-artifacts/", "local decoder corpus and reports"),
    (r"^logs/", "runtime logs may contain local or clinical context"),
    (r"^resources/screenshots-ignored/", "local screenshots may contain PHI/PII"),
    (r"^sample-DICOM-gitignored/", "local DICOM study directory"),
    (r"^test-DICOM-data/", "local DICOM/QC study directory"),
    (r"^\.sonar-local/", "local analysis state and coverage data"),
    (r"^\.phi-tools/", "isolated local PHI scanner environment"),
    (r"^tmp/", "local temporary workspace"),
    (r"(^|/)\.DS_Store$", "macOS metadata"),
    (r"(^|/)\.cache/", "cache directory"),
]
FORBIDDEN_ARTIFACT_SUFFIXES = {".cache", ".err", ".log", ".out", ".pkl", ".trace"}

# These exact ignore rules are repository privacy controls, not convenience
# entries.  The staged-index check below prevents a partial-staging trick from
# removing a rule while leaving it present only in the working tree.
REQUIRED_GITIGNORE_RULES = frozenset(
    {
        "*.db",
        "*.dcm",
        "*.dicom",
        "*.ima",
        "*.sqlite",
        "*.sqlite3",
        ".pytest-tmp-*/",
        ".pytest-tmp/",
        ".phi-tools/",
        ".sonar-local/",
        "backups/",
        "data/*",
        "decoder-spike-artifacts/",
        "logs/",
        "resources/screenshots-ignored/",
        "sample-DICOM-gitignored/",
        "test-DICOM-data/",
        "tmp/",
    }
)

# --- Rule 2: PHI/PII indicators inside data files ---------------------------
DATA_SUFFIXES = {
    ".json",
    ".csv",
    ".txt",
    ".log",
    ".yaml",
    ".yml",
    ".ini",
    ".cfg",
    ".xml",
    ".html",
    ".htm",
    ".ipynb",
    ".md",
    ".rst",
    ".svg",
    ".tex",
    ".tsv",
    ".ps",
    ".eps",
}
SPREADSHEET_SUFFIXES = {".xlsx", ".xlsm"}
DICOM_SUFFIXES = {".dcm", ".dicom", ".ima"}
NOTEBOOK_SUFFIXES = {".ipynb"}
PDF_SUFFIXES = {".pdf"}
POSTSCRIPT_SUFFIXES = {".eps", ".ps"}
# Archives and document packages can contain DICOM, private exports, and images.
# They are inspected in memory only and still require a hash-bound human review.
ARCHIVE_SUFFIXES = {
    ".7z",
    ".bz2",
    ".gz",
    ".rar",
    ".tar",
    ".xz",
    ".zip",
    ".zst",
}
OFFICE_DOCUMENT_SUFFIXES = {
    ".docm",
    ".docx",
    ".key",
    ".numbers",
    ".odp",
    ".ods",
    ".odt",
    ".pages",
    ".potm",
    ".potx",
    ".ppsm",
    ".ppsx",
    ".pptm",
    ".pptx",
}
CONTAINER_SUFFIXES = ARCHIVE_SUFFIXES | OFFICE_DOCUMENT_SUFFIXES | SPREADSHEET_SUFFIXES
SUPPORTED_SINGLE_FILE_COMPRESSION_SUFFIXES = {".bz2", ".gz", ".xz"}
UNSUPPORTED_ARCHIVE_SUFFIXES = {".7z", ".rar", ".zst"}
MAX_CONTAINER_DEPTH = 3
MAX_CONTAINER_MEMBERS = 500
MAX_CONTAINER_MEMBER_BYTES = 32 * 1024 * 1024
MAX_CONTAINER_TOTAL_BYTES = 96 * 1024 * 1024
MAX_TEXT_BYTES = 16 * 1024 * 1024
# SVG is text-readable and is also content-scanned above.  The other formats are
# opaque containers, so they need an explicit hash review before admission.
IMAGE_SUFFIXES = {
    ".avif",
    ".bmp",
    ".gif",
    ".heic",
    ".ico",
    ".icns",
    ".jpeg",
    ".jpg",
    ".jp2",
    ".jxl",
    ".png",
    ".svg",
    ".tif",
    ".tiff",
    ".webp",
}
APPROVED_MEDIA_MANIFEST = "security/approved-media-sha256.json"
APPROVED_TEXT_EXCEPTIONS_MANIFEST = "security/approved-phi-text-exceptions.json"
SHA256_PATTERN = re.compile(r"^[0-9a-f]{64}$")

# A file without a suffix can hide a binary export or clinical data.  Existing
# reviewed assets live in the hash manifest; every new one blocks until reviewed.
EXTENSIONLESS_EXEMPT = {
    ".cursorignore",
    ".cursorindexingignore",
    ".gitattributes",
    ".gitignore",
}

# These DICOM attributes must never hold a real value in a tracked fixture,
# including when nested in a sequence.  ``Dataset.iterall()`` visits sequence
# items recursively.
DICOM_IDENTIFIER_KEYWORDS = {
    "AccessionNumber",
    "InstitutionAddress",
    "InstitutionName",
    "IssuerOfPatientID",
    "OperatorsName",
    "OtherPatientIDs",
    "OtherPatientNames",
    "PatientAddress",
    "PatientBirthDate",
    "PatientBirthTime",
    "PatientComments",
    "PatientID",
    "PatientName",
    "PatientSex",
    "PatientTelephoneNumbers",
    "PerformingPhysicianName",
    "ReferringPhysicianName",
    "StationName",
    "StudyID",
}

CONTENT_RULES: list[tuple[re.Pattern[str], str]] = [
    (
        re.compile(
            r'"(recent_files|last_path|last_export_path|last_pylinac_output_path)"'
        ),
        "app config key that records user file paths",
    ),
    (
        re.compile(r"[A-Za-z]:\\+Users\\+|[A-Za-z]:\\+To\b", re.I),
        "Windows absolute user path",
    ),
    (
        re.compile(r"/(Users|home)/(?!runner\b|user\b)[A-Za-z0-9._-]+/"),
        "POSIX absolute home path",
    ),
    (re.compile(r"\\\\[^\\\s]+\\[^\\\s]+"), "UNC path"),
    (
        re.compile(r"\b[a-z][a-z0-9+.-]*://[^\s/@:]+:[^\s/@]+@", re.I),
        "authenticated URL",
    ),
    (
        re.compile(r'"Patient(Name|ID|BirthDate)"\s*:\s*"(?!\s*")[^"]+'),
        "populated DICOM patient tag",
    ),
    (re.compile(r"\b(?:dicom|pacs)://[^\s/]+", re.I), "DICOM/PACS endpoint"),
    (
        re.compile(r"\b(?:called|calling)[ _-]?ae(?:title)?\s*[:=]\s*\S+", re.I),
        "DICOM AE endpoint",
    ),
    (
        re.compile(r"\b[a-z0-9][a-z0-9.-]*\.(?:corp|internal|lan|local)\b", re.I),
        "internal hostname",
    ),
]
DOCUMENTATION_NETWORKS = (
    ipaddress.ip_network("192.0.2.0/24"),
    ipaddress.ip_network("198.51.100.0/24"),
    ipaddress.ip_network("203.0.113.0/24"),
    ipaddress.ip_network("2001:db8::/32"),
)
IP_ADDRESS_TOKEN = re.compile(
    r"(?<![0-9A-Fa-f:.])(?:[0-9]{1,3}(?:\.[0-9]{1,3}){3}|[0-9A-Fa-f]{0,4}:[0-9A-Fa-f:.]+)(?![0-9A-Fa-f:.])"
)
NETWORK_ADDRESS_CONTEXT = re.compile(
    r"(?:\b(?:ip|ipv4|ipv6|host|hostname|endpoint|server|address|listen|bind|"
    r"connect|url|proxy|pacs|dicom|ssh|tcp|udp)\b|https?://)",
    re.I,
)
SAFE_INTERNAL_HOSTNAMES = frozenset({"host.docker.internal"})
GENERIC_IDENTITIES = frozenset({"localhost", "runner", "user", "username"})

SENSITIVE_FILENAME_PATTERNS: tuple[re.Pattern[str], ...] = (
    re.compile(
        r"(?:^|[._-])(?:patientid|mrn|accession|account)[._-][A-Za-z0-9^.-]{4,}",
        re.I,
    ),
    re.compile(r"(?:^|[._-])patientname[._-][A-Za-z]+(?:\^[A-Za-z]+|[-_][A-Za-z]+)", re.I),
    re.compile(r"(?:^|[._-])\d+(?:\.\d+){3,}(?:[._-]|$)"),
)


def local_identities() -> frozenset[str]:
    """Return local account/host tokens without persisting or reporting them."""
    values = {
        getpass.getuser(),
        os.environ.get("USER", ""),
        os.environ.get("USERNAME", ""),
        socket.gethostname(),
        socket.getfqdn(),
    }
    return frozenset(
        value.lower()
        for value in values
        if len(value) >= 4 and value.lower() not in GENERIC_IDENTITIES
    )


def _contains_local_identity(text: str, identities: frozenset[str]) -> bool:
    return any(
        re.search(
            rf"(?<![A-Za-z0-9_.-]){re.escape(identity)}(?![A-Za-z0-9_.-])",
            text,
            re.I,
        )
        for identity in identities
    )


def _address_requires_redaction(address: ipaddress.IPv4Address | ipaddress.IPv6Address) -> bool:
    """Block public/private addresses; allow loopback and documentation ranges."""
    if address.is_loopback or any(address in network for network in DOCUMENTATION_NETWORKS):
        return False
    return address.is_global or address.is_private


def address_requires_redaction(
    address: ipaddress.IPv4Address | ipaddress.IPv6Address,
) -> bool:
    """Public network policy adapter for redacted repository preflight."""

    return _address_requires_redaction(address)


def _path_reasons(
    path: str, identities: frozenset[str] | None = None
) -> list[str]:
    """Classify a repository/archive path without returning matched components."""
    reasons: list[str] = []
    normalized = path.replace("\\", "/")
    path_parts = [part for part in normalized.split("/") if part]
    if normalized.startswith(("/", "//")) or re.match(r"^[A-Za-z]:/", normalized):
        reasons.append("absolute path name")
    if any(part == ".." for part in path_parts):
        reasons.append("path traversal name")
    if any(pattern.search(part) for part in path_parts for pattern in SENSITIVE_FILENAME_PATTERNS):
        reasons.append("sensitive-looking filename")
    if _contains_local_identity(normalized, local_identities() if identities is None else identities):
        reasons.append("local identity in filename")
    for match in IP_ADDRESS_TOKEN.finditer(normalized):
        try:
            address = ipaddress.ip_address(match.group())
        except ValueError:
            continue
        if _address_requires_redaction(address):
            reasons.append("network address in filename")
    return list(dict.fromkeys(reasons))


def path_reasons(
    path: str,
    identities: frozenset[str] | None = None,
) -> list[str]:
    """Public category-only path policy for repository preflight."""

    return _path_reasons(path, identities)


def _safe_display_path(path: str, identities: frozenset[str] | None = None) -> str:
    """Avoid echoing a filename when the filename is itself the finding."""
    return "[redacted repository path]" if _path_reasons(path, identities) else path


def _run(args: list[str], root: Path) -> str:
    return subprocess.run(
        args, cwd=root, capture_output=True, text=True, check=True
    ).stdout


def tracked_files(root: Path) -> list[str]:
    return [p for p in _run(["git", "ls-files"], root).splitlines() if p]


def staged_files(root: Path) -> list[str]:
    out = _run(["git", "diff", "--cached", "--name-only", "--diff-filter=ACM"], root)
    return [p for p in out.splitlines() if p]


def check_gitignore_policy(root: Path, *, staged: bool) -> list[str]:
    """Require privacy-critical ignore rules in the worktree or staged blob."""
    try:
        if staged:
            raw = subprocess.run(
                ["git", "show", ":.gitignore"],
                cwd=root,
                capture_output=True,
                check=True,
            ).stdout
            text = raw.decode("utf-8", errors="surrogateescape")
        else:
            text = (root / ".gitignore").read_text(encoding="utf-8")
    except (OSError, subprocess.CalledProcessError):
        return [".gitignore: privacy policy file is missing or unreadable"]

    active_rules = {
        line.strip()
        for line in text.splitlines()
        if line.strip() and not line.lstrip().startswith("#")
    }
    missing = sorted(REQUIRED_GITIGNORE_RULES - active_rules)
    return [
        ".gitignore: required privacy ignore rule is missing "
        f"({rule})"
        for rule in missing
    ]


def check_symlinks(paths: list[str], root: Path) -> list[str]:
    """Reject indexed symlinks whose targets are absolute or leave the repository."""
    wanted = set(paths)
    problems: list[str] = []
    try:
        entries = subprocess.run(
            ["git", "ls-files", "--stage", "-z"],
            cwd=root,
            capture_output=True,
            check=True,
        ).stdout.split(b"\0")
    except subprocess.CalledProcessError:
        return ["[repository index]: symlinks could not be inspected"]

    root_resolved = root.resolve()
    for raw_entry in entries:
        if not raw_entry or b"\t" not in raw_entry:
            continue
        metadata, raw_path = raw_entry.split(b"\t", 1)
        if not metadata.startswith(b"120000 "):
            continue
        path = raw_path.decode("utf-8", errors="surrogateescape")
        if path not in wanted:
            continue
        display_path = _safe_display_path(path)
        try:
            target = subprocess.run(
                ["git", "show", f":{path}"],
                cwd=root,
                capture_output=True,
                check=True,
            ).stdout.decode("utf-8", errors="surrogateescape")
        except subprocess.CalledProcessError:
            problems.append(f"{display_path}: symlink target could not be inspected")
            continue
        normalized_target = target.replace("\\", "/")
        absolute = normalized_target.startswith(("/", "//")) or bool(
            re.match(r"^[A-Za-z]:/", normalized_target)
        )
        resolved = (root / Path(path).parent / target).resolve(strict=False)
        if absolute:
            problems.append(f"{display_path}: absolute symlink target")
        elif not resolved.is_relative_to(root_resolved):
            problems.append(f"{display_path}: repository-escaping symlink target")
    return problems


def check_paths(paths: list[str]) -> list[str]:
    """Rule 1: forbidden paths, regardless of content."""
    problems: list[str] = []
    identities = local_identities()
    for path in paths:
        display_path = _safe_display_path(path, identities)
        for reason in _path_reasons(path, identities):
            problems.append(f"{display_path}: forbidden path ({reason})")
        if path == ".gitmodules" or path.endswith("/.gitmodules"):
            problems.append(f"{display_path}: repository remote configuration")
        if Path(path).suffix.lower() in FORBIDDEN_ARTIFACT_SUFFIXES:
            problems.append(f"{display_path}: forbidden runtime artifact")
            continue
        for pattern, why in FORBIDDEN_PATH_PATTERNS:
            if re.search(pattern, path):
                problems.append(f"{display_path}: forbidden artifact ({why})")
                break
    return problems


def _content_reasons(
    text: str, identities: frozenset[str] | None = None
) -> list[str]:
    """Return PHI/PII rule categories found in one text value, never the value."""
    reasons: list[str] = []
    for pattern, why in CONTENT_RULES:
        match = pattern.search(text)
        if match is None:
            continue
        if (
            why == "internal hostname"
            and match.group().lower() in SAFE_INTERNAL_HOSTNAMES
        ):
            continue
        reasons.append(why)
    local_terms = local_identities() if identities is None else identities
    if _contains_local_identity(text, local_terms):
        reasons.append("local account or hostname")
    for match in IP_ADDRESS_TOKEN.finditer(text):
        try:
            address = ipaddress.ip_address(match.group())
        except ValueError:
            continue
        context = text[max(0, match.start() - 100) : match.end() + 100]
        if not NETWORK_ADDRESS_CONTEXT.search(context):
            continue
        if _address_requires_redaction(address):
            reasons.append(
                "private-network address"
                if address.is_private
                else "public-network address"
            )
    return list(dict.fromkeys(reasons))


def check_contents(paths: list[str], root: Path) -> list[str]:
    """Rule 2: PHI/PII indicators inside data files."""
    problems = []
    approved = _approved_text_exceptions(root)
    for path in paths:
        if Path(path).suffix.lower() not in DATA_SUFFIXES:
            continue
        full = root / path
        if full.is_symlink() or not full.is_file():
            continue
        try:
            if full.stat().st_size > MAX_TEXT_BYTES:
                problems.append(f"{path}: text exceeds the inspection limit")
                continue
            text = full.read_text(encoding="utf-8", errors="ignore")
        except OSError as exc:
            problems.append(
                f"{path}: text could not be inspected ({type(exc).__name__})"
            )
            continue
        for lineno, line in enumerate(text.splitlines(), start=1):
            for reason in _content_reasons(line):
                if not _is_approved_text_exception(path, reason, root, approved):
                    problems.append(f"{path}:{lineno}: possible PHI/PII ({reason})")
    return problems


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _approved_text_exceptions(root: Path) -> dict[str, object]:
    """Load reviewed, hash-bound synthetic-text exceptions by repository path."""
    try:
        payload = json.loads(
            (root / APPROVED_TEXT_EXCEPTIONS_MANIFEST).read_text(encoding="utf-8")
        )
    except (OSError, json.JSONDecodeError):
        return {}
    if not isinstance(payload, dict):
        return {}
    files = payload.get("files", {})
    return files if isinstance(files, dict) else {}


def _is_approved_text_exception(
    path: str, reason: str, root: Path, approved: dict[str, object]
) -> bool:
    """Allow only a reviewed rule in an unchanged synthetic text fixture."""
    entry = approved.get(path)
    if not isinstance(entry, dict):
        return False
    expected_hash = entry.get("sha256")
    allowed_rules = entry.get("allowed_rules")
    return (
        isinstance(expected_hash, str)
        and isinstance(allowed_rules, list)
        and reason in allowed_rules
        and (root / path).is_file()
        and _sha256(root / path) == expected_hash
    )


def _approved_media(root: Path) -> dict[str, str]:
    manifest = root / APPROVED_MEDIA_MANIFEST
    try:
        payload = json.loads(manifest.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    files = payload.get("files", {})
    return files if isinstance(files, dict) else {}


def check_approval_manifests(root: Path) -> list[str]:
    """Fail closed when a reviewed-asset manifest is malformed or names risky paths."""
    problems: list[str] = []
    manifest_paths = (APPROVED_MEDIA_MANIFEST, APPROVED_TEXT_EXCEPTIONS_MANIFEST)
    if not any((root / relpath).exists() for relpath in manifest_paths):
        return []
    for relpath in manifest_paths:
        if not (root / relpath).exists():
            continue
        try:
            payload = json.loads((root / relpath).read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            problems.append(
                f"{relpath}: approval manifest could not be inspected "
                f"({type(exc).__name__})"
            )
            continue
        if not isinstance(payload, dict) or not isinstance(payload.get("files", {}), dict):
            problems.append(f"{relpath}: approval manifest has invalid structure")
            continue
        entries = payload.get("files", {})
        assert isinstance(entries, dict)
        for approved_path, entry in entries.items():
            if not isinstance(approved_path, str) or _path_reasons(approved_path):
                problems.append(f"{relpath}: approval manifest contains an unsafe path")
                continue
            expected_hash = entry if isinstance(entry, str) else entry.get("sha256") if isinstance(entry, dict) else None
            if not isinstance(expected_hash, str) or not SHA256_PATTERN.fullmatch(expected_hash):
                problems.append(f"{relpath}: approval manifest contains an invalid digest")
        if relpath == APPROVED_MEDIA_MANIFEST:
            trees = payload.get("image_trees", {})
            if not isinstance(trees, dict):
                problems.append(f"{relpath}: approval manifest has invalid image-tree structure")
            else:
                for directory, expected_hash in trees.items():
                    if (
                        not isinstance(directory, str)
                        or _path_reasons(directory)
                        or not isinstance(expected_hash, str)
                        or not SHA256_PATTERN.fullmatch(expected_hash)
                    ):
                        problems.append(f"{relpath}: approval manifest contains an invalid image tree")
    return list(dict.fromkeys(problems))


def _is_approved_media(path: str, root: Path, approved: dict[str, str]) -> bool:
    expected = approved.get(path)
    return (
        isinstance(expected, str)
        and (root / path).is_file()
        and _sha256(root / path) == expected
    )


def _image_tree_sha256(root: Path, directory: str) -> str:
    """Hash the paths and bytes of all tracked reviewable images in a directory."""
    prefix = f"{directory.rstrip('/')}/"
    digest = hashlib.sha256()
    for path in tracked_files(root):
        if (
            not path.startswith(prefix)
            or Path(path).suffix.lower() not in IMAGE_SUFFIXES
        ):
            continue
        full = root / path
        if not full.is_file():
            continue
        digest.update(path.encode("utf-8"))
        digest.update(b"\0")
        digest.update(bytes.fromhex(_sha256(full)))
    return digest.hexdigest()


def _approved_image_trees(manifest: dict[str, object]) -> dict[str, str]:
    """Return approved image-directory hashes from the review manifest."""
    trees = manifest.get("image_trees", {})
    if not isinstance(trees, dict):
        return {}
    return {
        directory.rstrip("/"): expected
        for directory, expected in trees.items()
        if isinstance(directory, str) and isinstance(expected, str)
    }


def _is_approved_image_tree(
    path: str, root: Path, approved_trees: dict[str, str], tree_digests: dict[str, str]
) -> bool:
    """Allow an image only when its entire reviewed directory is unchanged."""
    for directory, expected in approved_trees.items():
        if path.startswith(f"{directory}/"):
            actual = tree_digests.setdefault(
                directory, _image_tree_sha256(root, directory)
            )
            return actual == expected
    return False


def _check_spreadsheet(path: str, root: Path) -> list[str]:
    try:
        with (root / path).open("rb") as stream:
            return _check_spreadsheet_stream(path, stream, root)
    except OSError as exc:
        return [f"{path}: spreadsheet could not be inspected ({type(exc).__name__})"]


def _check_spreadsheet_stream(path: str, stream: BinaryIO, root: Path) -> list[str]:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(stream, read_only=True, data_only=False)
    except Exception as exc:
        return [f"{path}: spreadsheet could not be inspected ({type(exc).__name__})"]
    problems: list[str] = []
    approved = _approved_text_exceptions(root)
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for value in row:
                    if not isinstance(value, str):
                        continue
                    for reason in _content_reasons(value):
                        if not _is_approved_text_exception(
                            path, reason, root, approved
                        ):
                            problems.append(
                                f"{path}: worksheet {sheet.title!r}: possible PHI/PII ({reason})"
                            )
    finally:
        workbook.close()
    return problems


def _check_dicom(path: str, root: Path) -> list[str]:
    try:
        import pydicom

        dataset = pydicom.dcmread(root / path, stop_before_pixels=True, force=False)
    except Exception as exc:
        return [f"{path}: DICOM could not be inspected ({type(exc).__name__})"]
    return _check_dicom_dataset(path, dataset)


def _check_dicom_bytes(path: str, payload: bytes) -> list[str]:
    try:
        import pydicom

        dataset = pydicom.dcmread(
            io.BytesIO(payload), stop_before_pixels=True, force=False
        )
    except Exception as exc:
        return [f"{path}: archived DICOM could not be inspected ({type(exc).__name__})"]
    return _check_dicom_dataset(path, dataset)


def _check_dicom_dataset(path: str, dataset: Any) -> list[str]:
    problems: list[str] = []
    for element in dataset.iterall():
        value = str(element.value).strip()
        if element.tag.is_private and value:
            problems.append(f"{path}: populated private DICOM tag {element.tag}")
        if element.keyword not in DICOM_IDENTIFIER_KEYWORDS:
            continue
        synthetic_fixture = path.startswith("tests/fixtures/dicom_rdsr/") and value in {
            "Synthetic^RDSR",
            "SYN-RDSR-001",
        }
        if value and not synthetic_fixture:
            problems.append(
                f"{path}: populated nested DICOM identifier {element.keyword}"
            )
    return problems


def _has_dicom_preamble(path: Path) -> bool:
    """Recognize standard DICOM files even when their filename hides the type."""
    try:
        with path.open("rb") as stream:
            return stream.read(132)[128:132] == b"DICM"
    except OSError:
        return False


def _has_dicom_preamble_bytes(payload: bytes) -> bool:
    return len(payload) >= 132 and payload[128:132] == b"DICM"


def _check_notebook(path: str, root: Path) -> list[str]:
    """Require notebook outputs to be stripped before a reviewed notebook is admitted."""
    try:
        payload = json.loads((root / path).read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return [f"{path}: notebook could not be inspected"]
    cells = payload.get("cells", []) if isinstance(payload, dict) else []
    if not isinstance(cells, list):
        return [f"{path}: notebook has invalid cell structure"]
    if any(isinstance(cell, dict) and cell.get("outputs") for cell in cells):
        return [f"{path}: notebook outputs must be stripped before review"]
    return []


def _check_pdf(path: str, root: Path) -> list[str]:
    """Extract text from every PDF page; encrypted or unreadable files fail closed."""
    try:
        with (root / path).open("rb") as stream:
            return _check_pdf_stream(path, stream, root)
    except Exception as exc:
        return [f"{path}: PDF could not be inspected ({type(exc).__name__})"]


def _check_pdf_stream(path: str, stream: BinaryIO, root: Path) -> list[str]:
    try:
        from pypdf import PdfReader

        reader = PdfReader(stream)
        if reader.is_encrypted:
            return [f"{path}: encrypted PDF cannot be inspected"]
    except Exception as exc:
        return [f"{path}: PDF could not be inspected ({type(exc).__name__})"]

    problems: list[str] = []
    approved = _approved_text_exceptions(root)
    try:
        for page_number, page in enumerate(reader.pages, start=1):
            for reason in _content_reasons(page.extract_text() or ""):
                if not _is_approved_text_exception(path, reason, root, approved):
                    problems.append(
                        f"{path}: page {page_number}: possible PHI/PII ({reason})"
                    )
    except Exception as exc:
        return [f"{path}: PDF text extraction failed ({type(exc).__name__})"]
    return problems


class ReadableBinary(Protocol):
    """The minimal binary stream interface shared by archive readers."""

    def read(self, size: int = -1, /) -> bytes: ...


def _read_limited(stream: ReadableBinary, limit: int) -> bytes | None:
    """Read one untrusted member without permitting unbounded expansion."""
    payload = stream.read(limit + 1)
    return None if len(payload) > limit else payload


def _zip_members(payload: bytes) -> tuple[list[tuple[str, bytes]], str | None]:
    try:
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            members = [info for info in archive.infolist() if not info.is_dir()]
            if len(members) > MAX_CONTAINER_MEMBERS:
                return [], "too many archive members"
            if any(info.flag_bits & 0x1 for info in members):
                return [], "encrypted archive member cannot be inspected"
            if any((info.external_attr >> 16) & 0o170000 == 0o120000 for info in members):
                return [], "archive contains a link member"
            if sum(info.file_size for info in members) > MAX_CONTAINER_TOTAL_BYTES:
                return [], "archive expands beyond the inspection limit"
            result: list[tuple[str, bytes]] = []
            for info in members:
                if info.file_size > MAX_CONTAINER_MEMBER_BYTES:
                    return [], "archive member exceeds the inspection limit"
                with archive.open(info) as stream:
                    member = _read_limited(stream, MAX_CONTAINER_MEMBER_BYTES)
                if member is None:
                    return [], "archive member exceeds the inspection limit"
                result.append((info.filename, member))
            return result, None
    except (OSError, zipfile.BadZipFile, RuntimeError):
        return [], "archive could not be inspected"


def _tar_members(payload: bytes) -> tuple[list[tuple[str, bytes]], str | None]:
    try:
        with tarfile.open(fileobj=io.BytesIO(payload), mode="r:") as archive:
            all_members = archive.getmembers()
            if any(member.issym() or member.islnk() for member in all_members):
                return [], "archive contains a link member"
            members = [member for member in all_members if member.isfile()]
            if len(members) > MAX_CONTAINER_MEMBERS:
                return [], "too many archive members"
            if sum(member.size for member in members) > MAX_CONTAINER_TOTAL_BYTES:
                return [], "archive expands beyond the inspection limit"
            result: list[tuple[str, bytes]] = []
            for member_info in members:
                if member_info.size > MAX_CONTAINER_MEMBER_BYTES:
                    return [], "archive member exceeds the inspection limit"
                stream = archive.extractfile(member_info)
                if stream is None:
                    return [], "archive member could not be inspected"
                with stream:
                    member = _read_limited(stream, MAX_CONTAINER_MEMBER_BYTES)
                if member is None:
                    return [], "archive member exceeds the inspection limit"
                result.append((member_info.name, member))
            return result, None
    except (OSError, tarfile.TarError):
        return [], "archive could not be inspected"


def _compressed_member(
    path: str, payload: bytes
) -> tuple[list[tuple[str, bytes]], str | None]:
    suffix = Path(path).suffix.lower()
    try:
        with ExitStack() as stack:
            if suffix == ".gz":
                stream = stack.enter_context(
                    gzip.GzipFile(fileobj=io.BytesIO(payload), mode="rb")
                )
            elif suffix == ".bz2":
                stream = stack.enter_context(
                    bz2.BZ2File(io.BytesIO(payload), mode="rb")
                )
            elif suffix == ".xz":
                stream = stack.enter_context(
                    lzma.LZMAFile(io.BytesIO(payload), mode="rb")
                )
            else:
                return [], "unsupported archive type"
            member = _read_limited(stream, MAX_CONTAINER_TOTAL_BYTES)
    except (OSError, EOFError, lzma.LZMAError):
        return [], "archive could not be inspected"
    if member is None:
        return [], "archive expands beyond the inspection limit"
    return [(path.removesuffix(suffix), member)], None


def _container_members(
    path: str, payload: bytes
) -> tuple[list[tuple[str, bytes]], str | None]:
    """Return bounded in-memory members; unsupported containers fail closed."""
    suffix = Path(path).suffix.lower()
    if suffix in UNSUPPORTED_ARCHIVE_SUFFIXES:
        return [], "unsupported archive type"
    if zipfile.is_zipfile(io.BytesIO(payload)):
        return _zip_members(payload)
    if suffix in SUPPORTED_SINGLE_FILE_COMPRESSION_SUFFIXES:
        return _compressed_member(path, payload)
    if suffix == ".tar" or payload[257:262] == b"ustar":
        return _tar_members(payload)
    return [], "archive could not be inspected"


def _check_archived_text(path: str, payload: bytes, root: Path) -> list[str]:
    approved = _approved_text_exceptions(root)
    text = payload.decode("utf-8", errors="ignore")
    return [
        f"{path}: archived text: possible PHI/PII ({reason})"
        for reason in _content_reasons(text)
        if not _is_approved_text_exception(path, reason, root, approved)
    ]


def _looks_like_container(path: str, payload: bytes) -> bool:
    suffix = Path(path).suffix.lower()
    return (
        suffix in CONTAINER_SUFFIXES
        or zipfile.is_zipfile(io.BytesIO(payload))
        or payload[257:262] == b"ustar"
    )


def _check_container(
    path: str,
    root: Path,
    payload: bytes | None = None,
    depth: int = 0,
    container_name: str | None = None,
) -> list[str]:
    """Inspect nested archive/document contents without extracting them to disk."""
    if depth >= MAX_CONTAINER_DEPTH:
        return [f"{path}: archive nesting exceeds the inspection limit"]
    try:
        data = payload if payload is not None else (root / path).read_bytes()
    except OSError as exc:
        return [f"{path}: archive could not be inspected ({type(exc).__name__})"]
    if len(data) > MAX_CONTAINER_TOTAL_BYTES:
        return [f"{path}: archive exceeds the compressed inspection limit"]
    members, error = _container_members(container_name or path, data)
    if error:
        return [f"{path}: {error}"]

    problems: list[str] = []
    for member_name, member in members:
        member_reasons = _path_reasons(member_name)
        for reason in member_reasons:
            problems.append(f"{path}: archive contains a forbidden member name ({reason})")
        if Path(member_name).name == ".gitmodules":
            problems.append(f"{path}: archive contains repository remote configuration")
        suffix = Path(member_name).suffix.lower()
        if suffix in FORBIDDEN_ARTIFACT_SUFFIXES:
            problems.append(f"{path}: archive contains a forbidden runtime artifact")
        if suffix in DATA_SUFFIXES:
            problems.extend(_check_archived_text(path, member, root))
        is_dicom = suffix in DICOM_SUFFIXES or _has_dicom_preamble_bytes(member)
        if is_dicom:
            problems.extend(_check_dicom_bytes(path, member))
        elif suffix in SPREADSHEET_SUFFIXES:
            problems.extend(_check_spreadsheet_stream(path, io.BytesIO(member), root))
        elif suffix in PDF_SUFFIXES:
            problems.extend(_check_pdf_stream(path, io.BytesIO(member), root))
        if _looks_like_container(member_name, member):
            problems.extend(
                _check_container(path, root, member, depth + 1, member_name)
            )
    return problems


def check_reviewable_files(paths: list[str], root: Path) -> list[str]:
    """Fail closed for clinical/binary assets that require a human PHI review."""
    problems: list[str] = []
    manifest = root / APPROVED_MEDIA_MANIFEST
    try:
        manifest_payload = json.loads(manifest.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        manifest_payload = {}
    if not isinstance(manifest_payload, dict):
        manifest_payload = {}
    approved = _approved_media(root)
    approved_trees = _approved_image_trees(manifest_payload)
    tree_digests: dict[str, str] = {}
    for path in paths:
        if (root / path).is_symlink():
            continue
        suffix = Path(path).suffix.lower()
        is_extensionless = not suffix and path not in EXTENSIONLESS_EXEMPT
        has_dicom_preamble = _has_dicom_preamble(root / path)
        is_dicom = suffix in DICOM_SUFFIXES or has_dicom_preamble
        is_container = suffix in CONTAINER_SUFFIXES
        requires_review = (
            suffix in IMAGE_SUFFIXES
            or suffix in NOTEBOOK_SUFFIXES
            or suffix in PDF_SUFFIXES
            or suffix in POSTSCRIPT_SUFFIXES
            or is_container
            or is_extensionless
        )
        if is_dicom:
            problems.extend(_check_dicom(path, root))
            requires_review = True
        elif suffix in SPREADSHEET_SUFFIXES:
            problems.extend(_check_spreadsheet(path, root))
        elif suffix in NOTEBOOK_SUFFIXES:
            problems.extend(_check_notebook(path, root))
        elif suffix in PDF_SUFFIXES:
            problems.extend(_check_pdf(path, root))
        if is_container:
            problems.extend(_check_container(path, root))
        image_tree_approved = suffix in IMAGE_SUFFIXES and _is_approved_image_tree(
            path, root, approved_trees, tree_digests
        )
        if (
            requires_review
            and not image_tree_approved
            and not _is_approved_media(path, root, approved)
        ):
            asset_kind = (
                "DICOM"
                if is_dicom
                else "notebook"
                if suffix in NOTEBOOK_SUFFIXES
                else "PDF/PostScript"
                if suffix in PDF_SUFFIXES or suffix in POSTSCRIPT_SUFFIXES
                else "archive/document package"
                if is_container
                else "image/extensionless"
            )
            problems.append(
                f"{path}: unapproved {asset_kind} asset; "
                f"perform PHI/burned-in-text review and add its SHA-256 to {APPROVED_MEDIA_MANIFEST}"
            )
    return problems


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--staged", action="store_true", help="scan staged files only (pre-commit hook)"
    )
    parser.add_argument("--root", default=None, help="repository root")
    args = parser.parse_args()

    root = (
        Path(args.root)
        if args.root
        else Path(_run(["git", "rev-parse", "--show-toplevel"], Path.cwd()).strip())
    )

    paths = staged_files(root) if args.staged else tracked_files(root)
    scope = "staged" if args.staged else "tracked"

    problems = (
        check_gitignore_policy(root, staged=args.staged)
        + check_approval_manifests(root)
        + check_paths(paths)
        + check_symlinks(paths, root)
        + check_contents(paths, root)
        + check_reviewable_files(paths, root)
    )

    if problems:
        print(
            f"BLOCKED: {len(problems)} PHI/artifact issue(s) in {scope} files:\n",
            file=sys.stderr,
        )
        for p in problems:
            print(f"  {p}", file=sys.stderr)
        print(
            "\nThese files must not be committed. Remove them from the index:\n"
            "  git rm --cached <path>\n"
            "and confirm the path is covered by .gitignore.\n"
            "If this is a false positive, add an exemption in "
            "scripts/check_no_phi_artifacts.py.",
            file=sys.stderr,
        )
        return 1

    print(f"OK: no PHI artifacts ({len(paths)} {scope} file(s) scanned)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
