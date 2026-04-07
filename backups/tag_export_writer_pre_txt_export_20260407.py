"""
DICOM Tag Export Writer

Provides filename generation and file writing (Excel, CSV) for DICOM tag exports.
This module is pure logic with no Qt dependency.
"""

import csv
from pathlib import Path
from typing import Dict, List

from pydicom.dataset import Dataset

from core.dicom_parser import DICOMParser
from core.tag_export_catalog import missing_tag_export_display_fields


def generate_default_filename(
    studies: Dict[str, Dict[str, List[Dataset]]],
    selected_series: Dict[str, Dict[str, List[int]]],
) -> str:
    """Generate a default filename for tag export (Excel format)."""
    first_study_uid = list(selected_series.keys())[0]
    first_series_uid = list(selected_series[first_study_uid].keys())[0]
    first_instance_idx = selected_series[first_study_uid][first_series_uid][0]
    first_dataset = studies[first_study_uid][first_series_uid][first_instance_idx]

    modality = getattr(first_dataset, 'Modality', 'Unknown')
    accession = getattr(first_dataset, 'AccessionNumber', 'Unknown')
    return f"{modality} DICOM Tag Export {accession}.xlsx"


def write_excel_file(
    file_path: str,
    variation_analysis: Dict[str, Dict[str, List[str]]],
    studies: Dict[str, Dict[str, List[Dataset]]],
    selected_series: Dict[str, Dict[str, List[int]]],
    selected_tags: List[str],
    include_private: bool,
    include_missing_selected_tags: bool = True,
) -> None:
    """Write selected tags to an Excel file with per-instance export for varying tags."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        raise ImportError(
            "openpyxl library is required for Excel export. "
            "Install it with: pip install openpyxl"
        )

    wb = Workbook()
    default_ws = wb.active
    if default_ws is not None:
        wb.remove(default_ws)  # Remove default sheet

    # Create one sheet per study
    for study_uid, series_dict in selected_series.items():
        first_series_uid = list(series_dict.keys())[0]
        first_instance_idx = series_dict[first_series_uid][0]
        first_dataset = studies[study_uid][first_series_uid][first_instance_idx]
        study_desc = getattr(first_dataset, 'StudyDescription', 'Study')
        # Sanitize sheet name (max 31 chars, no special chars)
        sheet_name = study_desc[:31].replace('/', '-').replace('\\', '-').replace(':', '-')

        ws = wb.create_sheet(title=sheet_name)

        # Write header
        ws['A1'] = 'Instance'
        ws['B1'] = 'Tag Number'
        ws['C1'] = 'Name'
        ws['D1'] = 'Value'
        for cell in ('A1', 'B1', 'C1', 'D1'):
            ws[cell].font = Font(bold=True)

        row = 2

        # Export each selected series
        for series_uid, instance_indices in series_dict.items():
            datasets = studies[study_uid][series_uid]
            if not datasets:
                continue

            first_ds = datasets[0]

            # Write series header
            series_num = getattr(first_ds, 'SeriesNumber', '')
            series_desc = getattr(first_ds, 'SeriesDescription', 'Unknown')
            ws[f'A{row}'] = f"Series {series_num}: {series_desc}"
            ws[f'A{row}'].font = Font(bold=True, italic=True)
            ws.merge_cells(f'A{row}:D{row}')
            row += 1

            analysis = variation_analysis.get(
                series_uid, {'varying_tags': [], 'constant_tags': selected_tags}
            )
            varying_tags = analysis['varying_tags']
            constant_tags = analysis['constant_tags']

            # Export constant tags (once per series, using first instance)
            if constant_tags and instance_indices:
                first_instance_idx = instance_indices[0]
                if first_instance_idx < len(datasets):
                    dataset = datasets[first_instance_idx]
                    parser = DICOMParser(dataset)
                    all_tags = parser.get_all_tags(include_private=include_private)

                    for tag_str in constant_tags:
                        if tag_str in all_tags:
                            tag_data = all_tags[tag_str]
                            ws[f'A{row}'] = 'All'  # Indicates all instances
                            ws[f'B{row}'] = tag_data.get('tag', tag_str)
                            ws[f'C{row}'] = tag_data.get('name', '')

                            value = tag_data.get('value', '')
                            if isinstance(value, list):
                                value_str = ', '.join(str(v) for v in value)
                            else:
                                value_str = str(value)
                            ws[f'D{row}'] = value_str

                            row += 1
                        elif include_missing_selected_tags:
                            tag_num, tag_name = missing_tag_export_display_fields(tag_str)
                            ws[f'A{row}'] = 'All'
                            ws[f'B{row}'] = tag_num or tag_str
                            ws[f'C{row}'] = tag_name
                            ws[f'D{row}'] = ''
                            row += 1

            # Export varying tags (per instance)
            if varying_tags:
                for instance_idx in instance_indices:
                    if instance_idx >= len(datasets):
                        continue

                    dataset = datasets[instance_idx]
                    parser = DICOMParser(dataset)
                    all_tags = parser.get_all_tags(include_private=include_private)

                    instance_num = getattr(dataset, 'InstanceNumber', None)
                    instance_id = (
                        f"Instance {instance_num}"
                        if instance_num is not None
                        else f"Instance {instance_idx + 1}"
                    )

                    for tag_str in varying_tags:
                        if tag_str in all_tags:
                            tag_data = all_tags[tag_str]
                            ws[f'A{row}'] = instance_id
                            ws[f'B{row}'] = tag_data.get('tag', tag_str)
                            ws[f'C{row}'] = tag_data.get('name', '')

                            value = tag_data.get('value', '')
                            if isinstance(value, list):
                                value_str = ', '.join(str(v) for v in value)
                            else:
                                value_str = str(value)
                            ws[f'D{row}'] = value_str

                            row += 1
                        elif include_missing_selected_tags:
                            tag_num, tag_name = missing_tag_export_display_fields(tag_str)
                            ws[f'A{row}'] = instance_id
                            ws[f'B{row}'] = tag_num or tag_str
                            ws[f'C{row}'] = tag_name
                            ws[f'D{row}'] = ''
                            row += 1

            # Add blank row between series
            row += 1

    # Adjust column widths
    for sheet in wb.worksheets:
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 40
        sheet.column_dimensions['D'].width = 60

    wb.save(file_path)


def write_csv_files(
    base_file_path: str,
    variation_analysis: Dict[str, Dict[str, List[str]]],
    studies: Dict[str, Dict[str, List[Dataset]]],
    selected_series: Dict[str, Dict[str, List[int]]],
    selected_tags: List[str],
    include_private: bool,
    include_missing_selected_tags: bool = True,
) -> List[Path]:
    """
    Write selected tags to CSV files (one per study) with per-instance export for varying tags.

    Returns:
        List of created file paths.
    """
    base_path = Path(base_file_path)
    base_name = base_path.stem
    base_dir = base_path.parent
    exported_files: List[Path] = []

    # Create one CSV file per study
    for study_uid, series_dict in selected_series.items():
        first_series_uid = list(series_dict.keys())[0]
        first_instance_idx = series_dict[first_series_uid][0]
        first_dataset = studies[study_uid][first_series_uid][first_instance_idx]
        study_desc = getattr(first_dataset, 'StudyDescription', 'Study')
        safe_study_desc = study_desc.replace('/', '-').replace('\\', '-').replace(':', '-')[:50]

        if len(selected_series) > 1:
            csv_filename = f"{base_name}_{safe_study_desc}.csv"
        else:
            csv_filename = f"{base_name}.csv"

        csv_path = base_dir / csv_filename

        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)

            # Write header
            writer.writerow(['Instance', 'Tag Number', 'Name', 'Value'])

            # Export each selected series
            for series_uid, instance_indices in series_dict.items():
                datasets = studies[study_uid][series_uid]
                if not datasets:
                    continue

                first_ds = datasets[0]

                # Write series header
                series_num = getattr(first_ds, 'SeriesNumber', '')
                series_desc = getattr(first_ds, 'SeriesDescription', 'Unknown')
                writer.writerow([f"Series {series_num}: {series_desc}", '', '', ''])

                analysis = variation_analysis.get(
                    series_uid, {'varying_tags': [], 'constant_tags': selected_tags}
                )
                varying_tags = analysis['varying_tags']
                constant_tags = analysis['constant_tags']

                # Export constant tags (once per series, using first instance)
                if constant_tags and instance_indices:
                    first_instance_idx = instance_indices[0]
                    if first_instance_idx < len(datasets):
                        dataset = datasets[first_instance_idx]
                        parser = DICOMParser(dataset)
                        all_tags = parser.get_all_tags(include_private=include_private)

                        for tag_str in constant_tags:
                            if tag_str in all_tags:
                                tag_data = all_tags[tag_str]
                                tag_num = tag_data.get('tag', tag_str)
                                tag_name = tag_data.get('name', '')

                                value = tag_data.get('value', '')
                                if isinstance(value, list):
                                    value_str = ', '.join(str(v) for v in value)
                                else:
                                    value_str = str(value)

                                writer.writerow(['All', tag_num, tag_name, value_str])
                            elif include_missing_selected_tags:
                                tag_num, tag_name = missing_tag_export_display_fields(tag_str)
                                writer.writerow(['All', tag_num or tag_str, tag_name, ''])

                # Export varying tags (per instance)
                if varying_tags:
                    for instance_idx in instance_indices:
                        if instance_idx >= len(datasets):
                            continue

                        dataset = datasets[instance_idx]
                        parser = DICOMParser(dataset)
                        all_tags = parser.get_all_tags(include_private=include_private)

                        instance_num = getattr(dataset, 'InstanceNumber', None)
                        instance_id = (
                            f"Instance {instance_num}"
                            if instance_num is not None
                            else f"Instance {instance_idx + 1}"
                        )

                        for tag_str in varying_tags:
                            if tag_str in all_tags:
                                tag_data = all_tags[tag_str]
                                tag_num = tag_data.get('tag', tag_str)
                                tag_name = tag_data.get('name', '')

                                value = tag_data.get('value', '')
                                if isinstance(value, list):
                                    value_str = ', '.join(str(v) for v in value)
                                else:
                                    value_str = str(value)

                                writer.writerow([instance_id, tag_num, tag_name, value_str])
                            elif include_missing_selected_tags:
                                tag_num, tag_name = missing_tag_export_display_fields(tag_str)
                                writer.writerow([instance_id, tag_num or tag_str, tag_name, ''])

                # Add blank row between series
                writer.writerow([])

        exported_files.append(csv_path)

    return exported_files
