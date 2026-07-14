"""
Smoke tests for tag export writer backends.

Covers default filename generation and minimal CSV/XLSX/TXT output creation.
"""

import csv
import os
import sys
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from pydicom.dataset import Dataset
from pydicom.tag import Tag

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", "src"))

from core.tag_export_writer import (
    generate_default_filename,
    write_csv_files,
    write_excel_file,
    write_txt_files,
)


class TestTagExportWriterSmoke(unittest.TestCase):
    def _build_sample_data(self):
        ds1 = Dataset()
        ds1.StudyDescription = "Head CT"
        ds1.SeriesDescription = "Axial"
        ds1.SeriesNumber = 1
        ds1.InstanceNumber = 1
        ds1.PatientName = "Demo^Patient"
        ds1.Modality = "CT"
        ds1.AccessionNumber = "ACC123"

        ds2 = Dataset()
        ds2.StudyDescription = "Head CT"
        ds2.SeriesDescription = "Axial"
        ds2.SeriesNumber = 1
        ds2.InstanceNumber = 2
        ds2.PatientName = "Demo^Patient"
        ds2.Modality = "CT"
        ds2.AccessionNumber = "ACC123"

        studies = {"study1": {"series1": [ds1, ds2]}}
        selected_series = {"study1": {"series1": [0, 1]}}

        patient_name_tag = str(Tag(0x0010, 0x0010))
        instance_number_tag = str(Tag(0x0020, 0x0013))
        selected_tags = [patient_name_tag, instance_number_tag]

        variation_analysis = {
            "series1": {
                "varying_tags": [instance_number_tag],
                "constant_tags": [patient_name_tag],
            }
        }

        return studies, selected_series, selected_tags, variation_analysis

    def test_generate_default_filename(self):
        studies, selected_series, _, _ = self._build_sample_data()
        filename = generate_default_filename(studies, selected_series)
        self.assertEqual(filename, "CT DICOM Tag Export ACC123.xlsx")

    def test_write_csv_files_smoke(self):
        studies, selected_series, selected_tags, variation_analysis = self._build_sample_data()

        with TemporaryDirectory() as tmp_dir:
            out_base = str(Path(tmp_dir) / "tag_export.csv")
            exported = write_csv_files(
                out_base,
                variation_analysis,
                studies,
                selected_series,
                selected_tags,
                include_private=False,
            )

            self.assertEqual(len(exported), 1)
            csv_path = exported[0]
            self.assertTrue(csv_path.exists())

            with open(csv_path, newline="", encoding="utf-8") as f:
                rows = list(csv.reader(f))

            self.assertGreaterEqual(len(rows), 4)
            self.assertEqual(rows[0], ["Instance", "Tag Number", "Name", "Value"])
            self.assertTrue(any(r and r[0] == "All" for r in rows))
            self.assertTrue(any(r and r[0].startswith("Instance ") for r in rows))

    def test_write_excel_file_smoke(self):
        studies, selected_series, selected_tags, variation_analysis = self._build_sample_data()

        with TemporaryDirectory() as tmp_dir:
            xlsx_path = str(Path(tmp_dir) / "tag_export.xlsx")
            write_excel_file(
                xlsx_path,
                variation_analysis,
                studies,
                selected_series,
                selected_tags,
                include_private=False,
            )

            self.assertTrue(Path(xlsx_path).exists())

            from openpyxl import load_workbook

            wb = load_workbook(xlsx_path)
            self.assertGreaterEqual(len(wb.sheetnames), 1)
            ws = wb[wb.sheetnames[0]]
            self.assertEqual(ws["A1"].value, "Instance")
            self.assertEqual(ws["B1"].value, "Tag Number")
            self.assertEqual(ws["C1"].value, "Name")
            self.assertEqual(ws["D1"].value, "Value")

    def test_write_txt_files_smoke(self):
        studies, selected_series, selected_tags, variation_analysis = self._build_sample_data()

        with TemporaryDirectory() as tmp_dir:
            out_base = str(Path(tmp_dir) / "tag_export.txt")
            exported = write_txt_files(
                out_base,
                variation_analysis,
                studies,
                selected_series,
                selected_tags,
                include_private=False,
            )

            self.assertEqual(len(exported), 1)
            txt_path = exported[0]
            self.assertTrue(txt_path.exists())

            with open(txt_path, newline="", encoding="utf-8") as f:
                rows = list(csv.reader(f, delimiter="\t"))

            self.assertGreaterEqual(len(rows), 4)
            self.assertEqual(rows[0], ["Instance", "Tag Number", "Name", "Value"])
            self.assertTrue(any(r and r[0] == "All" for r in rows))
            self.assertTrue(any(r and r[0].startswith("Instance ") for r in rows))


if __name__ == "__main__":
    unittest.main()
