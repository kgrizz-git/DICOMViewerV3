"""
Tests for P2-I1: per-module image capture on QARequest/QAResult and the
analyzer.save_images() call in the ACR CT and MRI runners.

No live analyze(), no DICOM fixtures, no PHI. The pylinac analyzer is fully
mocked; we only assert the runner wires up the request fields, calls
save_images(directory=...), populates analyzed_module_images, and keeps the
paths out of JSON/CSV output.
"""

from __future__ import annotations

import builtins as _builtins
import io
import json
from pathlib import Path
from typing import Any
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from qa.analysis_types import QARequest, QAResult
from qa.qa_export import build_metrics_csv, build_single_run_document
from qa.qa_result_flatten import build_metric_rows, build_tabular_run
from qa.qa_xlsx_export import build_qa_workbook

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_request(**overrides: Any) -> QARequest:
    """A minimal QARequest with sensible defaults for the runners."""
    base: dict[str, Any] = {
        "analysis_type": "acr_ct",
        "dicom_paths": ["/fake/a.dcm"],
        "folder_path": None,
        "study_uid": "1.2.3",
        "series_uid": "1.2.3.4",
        "modality": "CT",
    }
    base.update(overrides)
    return QARequest(**base)


def _stub_analyzer(module_keys: tuple[str, ...] = ("hu", "uniformity")) -> Any:
    """A fake analyzer whose save_images() drops one PNG per module key."""
    analyzer = MagicMock()

    def _save_images(directory: str | None = None, **_kw: Any) -> list[Path]:
        assert directory is not None
        out = Path(directory)
        out.mkdir(parents=True, exist_ok=True)
        paths = []
        for key in module_keys:
            p = out / f"{key}.png"
            p.write_bytes(b"\x89PNG")
            paths.append(p)
        return paths

    analyzer.save_images.side_effect = _save_images
    # results_data(as_dict=True) → minimal raw_pylinac
    analyzer.results_data.return_value = {"num_images": 1}
    return analyzer


class _ChainPatcher:
    """Patch pylinac + ACR class imports in both runner modules.

    The runners import pylinac lazily inside the function body, so we intercept
    ``builtins.__import__`` to hand back a fake pylinac module and classes.
    """

    def __init__(self, analyzer: Any) -> None:
        self.analyzer = analyzer
        self._patches: list[Any] = []
        self._real_import = _builtins.__import__
        self._fake_pylinac = MagicMock()
        self._fake_pylinac.__version__ = "3.43.2"
        self._fake_pylinac.ACRCT = lambda *a, **k: (_ for _ in ()).throw(
            AssertionError("stock CT not used by default")
        )
        self._fake_pylinac.ACRMRILarge = lambda *a, **k: (_ for _ in ()).throw(
            AssertionError("stock MRI not used by default")
        )

    def _fake_import(self, name: str, *args: Any, **kwargs: Any) -> Any:
        if name == "pylinac":
            return self._fake_pylinac
        if name == "pylinac.acr":
            fake_acr = MagicMock()
            fake_acr.ACRCT = self._fake_pylinac.ACRCT
            fake_acr.ACRMRILarge = self._fake_pylinac.ACRMRILarge
            return fake_acr
        return self._real_import(name, *args, **kwargs)

    def __enter__(self) -> _ChainPatcher:
        # The runners import pylinac lazily inside the function body, and the
        # viewer subclasses from qa.pylinac_extent_subclasses. Patch the
        # subclass module-level names and intercept builtins.__import__ for the
        # top-level pylinac / from-pylinac imports.
        self._patches = [
            patch(
                "qa.pylinac_extent_subclasses.ACRCTForViewer",
                lambda *a, **k: self.analyzer,
            ),
            patch(
                "qa.pylinac_extent_subclasses.ACRMRILargeForViewer",
                lambda *a, **k: self.analyzer,
            ),
        ]
        started: list[Any] = []
        try:
            for p in self._patches:
                p.start()
                started.append(p)
            self._orig_builtin_import = _builtins.__import__
            _builtins.__import__ = self._fake_import
        except Exception:
            _builtins.__import__ = getattr(self, "_orig_builtin_import", _builtins.__import__)
            for p in reversed(started):
                p.stop()
            raise
        return self

    def __exit__(self, *exc: Any) -> None:
        _builtins.__import__ = self._orig_builtin_import
        for p in self._patches:
            p.stop()


# ---------------------------------------------------------------------------
# QARequest / QAResult field defaults
# ---------------------------------------------------------------------------


def test_qa_request_defaults() -> None:
    req = QARequest(analysis_type="acr_ct")
    assert req.embed_module_images_in_xlsx is True
    assert req.module_images_out_dir is None
    assert req.analyzed_image_out_path is None


def test_qa_result_default_module_images_empty() -> None:
    res = QAResult(success=True, analysis_type="acr_ct")
    assert res.analyzed_module_images == {}
    assert res.analyzed_image_path is None


# ---------------------------------------------------------------------------
# CT runner: embed on + dir set
# ---------------------------------------------------------------------------


def test_ct_embed_on_calls_save_images_populates_dict_no_composite(tmp_path: Path) -> None:
    analyzer = _stub_analyzer(("hu", "uniformity", "mtf"))
    out_dir = tmp_path / "ct_modules"

    with _ChainPatcher(analyzer):
        from qa.pylinac_acr_ct import run_acr_ct_analysis

        req = _make_request(
            embed_module_images_in_xlsx=True,
            module_images_out_dir=str(out_dir),
            analyzed_image_out_path=str(tmp_path / "composite.png"),
        )
        result = run_acr_ct_analysis(req)

    assert result.success is True
    analyzer.save_images.assert_called_once_with(directory=str(out_dir))
    # Composite save must NOT be called when embed is on + dir set
    analyzer.save_analyzed_image.assert_not_called()
    assert result.analyzed_image_path is None
    assert set(result.analyzed_module_images.keys()) == {"hu", "uniformity", "mtf"}
    for _label, path in result.analyzed_module_images.items():
        assert Path(path).is_absolute()
        assert Path(path).exists()
        assert Path(path).suffix == ".png"


def test_ct_ignores_leftover_pngs_in_out_dir(tmp_path: Path) -> None:
    """save_images return paths are used; leftover files in the dir are ignored."""
    analyzer = _stub_analyzer(("hu",))
    out_dir = tmp_path / "ct_modules"
    out_dir.mkdir()
    leftover = out_dir / "stale.png"
    leftover.write_bytes(b"\x89PNG")

    with _ChainPatcher(analyzer):
        from qa.pylinac_acr_ct import run_acr_ct_analysis

        req = _make_request(
            embed_module_images_in_xlsx=True,
            module_images_out_dir=str(out_dir),
        )
        result = run_acr_ct_analysis(req)

    assert set(result.analyzed_module_images.keys()) == {"hu"}
    assert "stale" not in result.analyzed_module_images
    assert str(leftover) not in result.analyzed_module_images.values()


def test_ct_embed_off_skips_composite_and_save_images(tmp_path: Path) -> None:
    """P2-X4: embed off → no composite save, no module images, no embeddable path."""
    analyzer = _stub_analyzer(("hu",))
    out_dir = tmp_path / "ct_modules"
    composite_path = tmp_path / "composite.png"

    def _save_composite(path: str) -> None:
        Path(path).write_bytes(b"\x89PNG")

    analyzer.save_analyzed_image.side_effect = _save_composite

    with _ChainPatcher(analyzer):
        from qa.pylinac_acr_ct import run_acr_ct_analysis

        req = _make_request(
            embed_module_images_in_xlsx=False,
            module_images_out_dir=str(out_dir),
            analyzed_image_out_path=str(composite_path),
        )
        result = run_acr_ct_analysis(req)

    assert result.success is True
    analyzer.save_images.assert_not_called()
    analyzer.save_analyzed_image.assert_not_called()
    assert result.analyzed_image_path is None
    assert result.analyzed_module_images == {}


def test_ct_embed_on_but_no_dir_falls_back_to_composite(tmp_path: Path) -> None:
    analyzer = _stub_analyzer(("hu",))
    composite_path = tmp_path / "composite.png"

    def _save_composite(path: str) -> None:
        Path(path).write_bytes(b"\x89PNG")

    analyzer.save_analyzed_image.side_effect = _save_composite

    with _ChainPatcher(analyzer):
        from qa.pylinac_acr_ct import run_acr_ct_analysis

        req = _make_request(
            embed_module_images_in_xlsx=True,
            module_images_out_dir=None,  # no dir → fall back
            analyzed_image_out_path=str(composite_path),
        )
        result = run_acr_ct_analysis(req)

    assert result.success is True
    analyzer.save_images.assert_not_called()
    analyzer.save_analyzed_image.assert_called_once_with(str(composite_path))
    assert result.analyzed_image_path == str(composite_path)
    assert result.analyzed_module_images == {}


def test_ct_save_images_failure_does_not_fail_run(tmp_path: Path) -> None:
    analyzer = _stub_analyzer(("hu",))
    out_dir = tmp_path / "ct_modules"
    analyzer.save_images.side_effect = RuntimeError("matplotlib boom")

    with _ChainPatcher(analyzer):
        from qa.pylinac_acr_ct import run_acr_ct_analysis

        req = _make_request(
            embed_module_images_in_xlsx=True,
            module_images_out_dir=str(out_dir),
        )
        result = run_acr_ct_analysis(req)

    assert result.success is True
    assert result.analyzed_module_images == {}


# ---------------------------------------------------------------------------
# MRI runner: embed on + dir set
# ---------------------------------------------------------------------------


def test_mri_embed_on_calls_save_images_populates_dict(tmp_path: Path) -> None:
    analyzer = _stub_analyzer(("geometric", "slice 1", "rMTF"))
    out_dir = tmp_path / "mri_modules"

    with _ChainPatcher(analyzer):
        from qa.pylinac_acr_mri import run_acr_mri_large_analysis

        req = _make_request(
            analysis_type="acr_mri_large",
            embed_module_images_in_xlsx=True,
            module_images_out_dir=str(out_dir),
        )
        result = run_acr_mri_large_analysis(req)

    assert result.success is True
    analyzer.save_images.assert_called_once_with(directory=str(out_dir))
    assert set(result.analyzed_module_images.keys()) == {"geometric", "slice 1", "rMTF"}
    for path in result.analyzed_module_images.values():
        assert Path(path).is_absolute()
        assert Path(path).exists()


def test_mri_embed_off_does_not_call_save_images(tmp_path: Path) -> None:
    analyzer = _stub_analyzer(("geometric",))
    out_dir = tmp_path / "mri_modules"

    with _ChainPatcher(analyzer):
        from qa.pylinac_acr_mri import run_acr_mri_large_analysis

        req = _make_request(
            analysis_type="acr_mri_large",
            embed_module_images_in_xlsx=False,
            module_images_out_dir=str(out_dir),
        )
        result = run_acr_mri_large_analysis(req)

    assert result.success is True
    analyzer.save_images.assert_not_called()
    assert result.analyzed_module_images == {}


def test_mri_save_images_failure_does_not_fail_run(tmp_path: Path) -> None:
    analyzer = _stub_analyzer(("geometric",))
    out_dir = tmp_path / "mri_modules"
    analyzer.save_images.side_effect = RuntimeError("matplotlib boom")

    with _ChainPatcher(analyzer):
        from qa.pylinac_acr_mri import run_acr_mri_large_analysis

        req = _make_request(
            analysis_type="acr_mri_large",
            embed_module_images_in_xlsx=True,
            module_images_out_dir=str(out_dir),
        )
        result = run_acr_mri_large_analysis(req)

    assert result.success is True
    assert result.analyzed_module_images == {}


# ---------------------------------------------------------------------------
# Serialization: module paths must not leak into JSON / CSV
# ---------------------------------------------------------------------------


def test_module_images_not_in_flatten_output() -> None:
    result = QAResult(
        success=True,
        analysis_type="acr_ct",
        raw_pylinac={"num_images": 1},
        metrics={"num_images": 1},
        analyzed_module_images={"hu": "/tmp/ct/hu.png", "mtf": "/tmp/ct/mtf.png"},
        analyzed_image_path="/tmp/ct/composite.png",
    )
    rows = dict(build_metric_rows(result))
    assert "analyzed_module_images" not in rows
    assert "analyzed_image_path" not in rows
    tab = build_tabular_run(result)
    assert "analyzed_module_images" not in tab
    assert "analyzed_image_path" not in tab


def test_module_images_not_in_json_document(tmp_path: Path) -> None:
    """The single-run JSON document must not carry module-image paths."""
    analyzer = _stub_analyzer(("hu", "mtf"))
    out_dir = tmp_path / "ct_modules"

    with _ChainPatcher(analyzer):
        from qa.pylinac_acr_ct import run_acr_ct_analysis

        req = _make_request(
            embed_module_images_in_xlsx=True,
            module_images_out_dir=str(out_dir),
        )
        result = run_acr_ct_analysis(req)

    doc = build_single_run_document(result, app_version="test")
    blob = json.dumps(doc)
    csv_text = build_metrics_csv(result)
    assert "analyzed_module_images" not in blob
    assert "analyzed_image_path" not in blob
    for path in result.analyzed_module_images.values():
        assert path not in blob
        assert path not in csv_text
    assert "analyzed_module_images" not in csv_text


# ---------------------------------------------------------------------------
# P2-X4 end-to-end: toggle-off → runner produces no embeddable image →
# build_qa_workbook skips the Images sheet (composite fallback must not run).
# ---------------------------------------------------------------------------


def _xlsx_sheet_names(wb: openpyxl.Workbook) -> list[str]:
    return wb.sheetnames


def _save_and_reload(wb: openpyxl.Workbook) -> openpyxl.Workbook:
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return openpyxl.load_workbook(buffer)


def test_p2x4_ct_embed_off_runner_result_skips_images_sheet(tmp_path: Path) -> None:
    """P2-X4: embed off → composite not saved → workbook has no Images sheet."""
    analyzer = _stub_analyzer(("hu",))
    out_dir = tmp_path / "ct_modules"
    composite_path = tmp_path / "composite.png"
    analyzer.save_analyzed_image.side_effect = lambda p: Path(p).write_bytes(b"\x89PNG")

    with _ChainPatcher(analyzer):
        from qa.pylinac_acr_ct import run_acr_ct_analysis

        req = _make_request(
            embed_module_images_in_xlsx=False,
            module_images_out_dir=str(out_dir),
            analyzed_image_out_path=str(composite_path),
        )
        result = run_acr_ct_analysis(req)

    # Runner must not leave any embeddable image when embed is off.
    assert result.analyzed_image_path is None
    assert result.analyzed_module_images == {}

    wb = build_qa_workbook([result], labels=["Run 1"])
    reloaded = _save_and_reload(wb)
    assert "Images" not in _xlsx_sheet_names(reloaded)
    summary_values = [
        c.value for row in reloaded["Summary"].iter_rows() for c in row
    ]
    assert any(
        isinstance(v, str) and "Images sheet skipped" in v for v in summary_values
    )


def test_p2x4_ct_embed_off_paths_denylisted_from_csv_json(tmp_path: Path) -> None:
    """P2-X4: even if a composite path existed, embed-off result has none to leak."""
    result = QAResult(
        success=True,
        analysis_type="acr_ct",
        metrics={"input_count": 5, "vanilla_pylinac": False},
        study_uid="1.2.3",
        series_uid="1.2.3.4",
        modality="CT",
        num_images=5,
        # Simulate the post-fix state: embed off → no paths at all.
        analyzed_image_path=None,
        analyzed_module_images={},
    )
    csv_text = build_metrics_csv(result)
    doc = build_single_run_document(result, app_version="test")
    blob = json.dumps(doc)
    assert "analyzed_image_path" not in blob
    assert "analyzed_module_images" not in blob
    assert "analyzed_image_path" not in csv_text
    assert "analyzed_module_images" not in csv_text


def test_p2x4_ct_embed_on_regression_modules_embed(tmp_path: Path) -> None:
    """P2-X4 regression: embed on + modules → Images sheet still embeds them."""
    pytest.importorskip("PIL")

    analyzer = _stub_analyzer(("hu", "mtf"))
    out_dir = tmp_path / "ct_modules"

    with _ChainPatcher(analyzer):
        from qa.pylinac_acr_ct import run_acr_ct_analysis

        req = _make_request(
            embed_module_images_in_xlsx=True,
            module_images_out_dir=str(out_dir),
        )
        result = run_acr_ct_analysis(req)

    assert result.analyzed_module_images
    assert result.analyzed_image_path is None  # composite skipped when embed on + dir set

    wb = build_qa_workbook([result], labels=["Run 1"])
    reloaded = _save_and_reload(wb)
    assert "Images" in _xlsx_sheet_names(reloaded)


def test_p2x4_ct_embed_on_no_dir_falls_back_to_composite(tmp_path: Path) -> None:
    """P2-X4 regression: embed on but no dir → composite fallback still works."""
    PILImage = pytest.importorskip("PIL.Image")

    analyzer = _stub_analyzer(("hu",))
    composite_path = tmp_path / "composite.png"
    PILImage.new("RGB", (200, 200), color="red").save(composite_path)
    analyzer.save_analyzed_image.side_effect = lambda p: None  # file already exists

    with _ChainPatcher(analyzer):
        from qa.pylinac_acr_ct import run_acr_ct_analysis

        req = _make_request(
            embed_module_images_in_xlsx=True,
            module_images_out_dir=None,  # no dir → composite fallback
            analyzed_image_out_path=str(composite_path),
        )
        result = run_acr_ct_analysis(req)

    assert result.analyzed_image_path == str(composite_path)

    wb = build_qa_workbook([result], labels=["Run 1"])
    reloaded = _save_and_reload(wb)
    assert "Images" in _xlsx_sheet_names(reloaded)
