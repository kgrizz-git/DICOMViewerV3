"""
Unit tests for Feature 3 — XLSX export (``qa.qa_xlsx_export``).

Covers ``build_qa_workbook`` via an in-memory openpyxl round-trip (single and
multi-row), the Pillow-missing degradation path (Images sheet skipped, note
cell added to Summary), the Qt-free import guarantee, and the transient
``analyzed_image_path`` leak guard shared with the JSON/CSV builders.
"""

from __future__ import annotations

import ast
import io
import sys
from pathlib import Path

import openpyxl
import pytest
from openpyxl.utils.units import pixels_to_EMU

from qa import qa_xlsx_export
from qa.analysis_types import QAResult
from qa.qa_export import build_metrics_csv, build_single_run_document
from qa.qa_result_flatten import build_metric_rows
from qa.qa_xlsx_export import build_qa_workbook

_SRC_ROOT = Path(__file__).resolve().parent.parent / "src"


def _result(
    *,
    series_uid: str = "1.2.3",
    success: bool = True,
    cnr_details: dict | None = None,
    warnings: list[str] | None = None,
    analyzed_image_path: str | None = None,
    analyzed_module_images: dict[str, str] | None = None,
) -> QAResult:
    metrics: dict = {"input_count": 5, "vanilla_pylinac": False}
    if cnr_details is not None:
        metrics["low_contrast_cnr"] = cnr_details
    return QAResult(
        success=success,
        analysis_type="acr_ct",
        metrics=metrics,
        warnings=warnings or [],
        errors=[] if success else ["boom"],
        study_uid="1.2.3.study",
        series_uid=series_uid,
        modality="CT",
        num_images=5,
        pylinac_version="3.43.2",
        analyzed_image_path=analyzed_image_path,
        analyzed_module_images=analyzed_module_images or {},
    )


_CNR_DETAILS = {
    "cnr": 4.5,
    "object_rois": [{"mean": 100.0, "pixel_value": 100.0, "contrast_to_noise": 3.0}],
    "background": {"means": [10.0], "stds": [2.0], "mean": 10.0, "std": 2.0},
}


def _save_and_reload(wb: openpyxl.Workbook) -> openpyxl.Workbook:
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return openpyxl.load_workbook(buffer)


def test_single_row_workbook_round_trip() -> None:
    result = _result(cnr_details=_CNR_DETAILS, warnings=["w1"])
    wb = build_qa_workbook([result], labels=["Series A"], app_version="9.9.9")
    reloaded = _save_and_reload(wb)

    assert reloaded.sheetnames[:2] == ["Summary", "Detail"]
    summary = reloaded["Summary"]
    header = [c.value for c in summary[1]]
    assert header[0] == "Series/Run ID"
    row2 = [c.value for c in summary[2]]
    assert row2[0] == "Series A"
    assert row2[1] == 100.0  # object ROI mean
    assert row2[2] == 10.0  # background mean
    assert row2[3] == 2.0  # background std
    assert row2[4] == 4.5  # cnr
    assert row2[5] == "success"
    assert row2[6] == "w1"

    detail = reloaded["Detail"]
    detail_values = [c.value for row in detail.iter_rows() for c in row]
    assert "Series A" in detail_values
    assert "low_contrast_cnr.cnr" in detail_values


def test_multi_row_workbook_round_trip() -> None:
    results = [
        _result(series_uid="1.1", cnr_details=_CNR_DETAILS),
        _result(series_uid="1.2", success=False),
    ]
    wb = build_qa_workbook(results)  # no labels -> falls back to series_uid
    reloaded = _save_and_reload(wb)

    summary = reloaded["Summary"]
    ids = [row[0].value for row in summary.iter_rows(min_row=2, max_row=3)]
    assert ids == ["1.1", "1.2"]
    statuses = [row[5].value for row in summary.iter_rows(min_row=2, max_row=3)]
    assert statuses == ["success", "failed"]


def test_labels_length_mismatch_raises() -> None:
    with pytest.raises(ValueError):
        build_qa_workbook([_result(), _result(series_uid="1.4")], labels=["only one"])


def test_no_analyzed_image_skips_images_sheet_with_note() -> None:
    result = _result(cnr_details=_CNR_DETAILS, analyzed_image_path=None)
    wb = build_qa_workbook([result])
    reloaded = _save_and_reload(wb)

    assert "Images" not in reloaded.sheetnames
    summary_values = [c.value for row in reloaded["Summary"].iter_rows() for c in row]
    assert any(
        isinstance(v, str) and "Images sheet skipped" in v for v in summary_values
    )


def test_pillow_missing_skips_images_sheet(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    """
    Simulate a Pillow-less environment via the module's availability guard
    (real Pillow import failure would require uninstalling the dependency).
    """
    png_path = tmp_path / "analyzed.png"
    png_path.write_bytes(b"not a real png but path exists")
    result = _result(cnr_details=_CNR_DETAILS, analyzed_image_path=str(png_path))

    monkeypatch.setattr(qa_xlsx_export, "_PILLOW_AVAILABLE", False)
    wb = build_qa_workbook([result])
    reloaded = _save_and_reload(wb)

    assert "Images" not in reloaded.sheetnames
    summary_values = [c.value for row in reloaded["Summary"].iter_rows() for c in row]
    assert any(
        isinstance(v, str) and "Pillow is not available" in v for v in summary_values
    )
    # Rest of the workbook still writes.
    assert reloaded["Detail"].max_row > 1


def test_images_sheet_deterministic_dimensions_and_stride(tmp_path: Path) -> None:
    """Verify image dimensions are set to 480x480 and row stride is 34 rows per run."""
    try:
        from PIL import Image as PILImage
    except ImportError:
        pytest.skip("Pillow is not installed")

    img_path1 = tmp_path / "img1.png"
    img_path2 = tmp_path / "img2.png"
    PILImage.new("RGB", (200, 100), color="red").save(img_path1)
    PILImage.new("RGB", (300, 300), color="blue").save(img_path2)

    res1 = _result(series_uid="1.1", analyzed_image_path=str(img_path1))
    res2 = _result(series_uid="1.2", analyzed_image_path=str(img_path2))

    workbook = build_qa_workbook([res1, res2], labels=["Run 1", "Run 2"])
    reloaded = _save_and_reload(workbook)
    assert "Images" in reloaded.sheetnames
    ws = reloaded["Images"]

    # Cell positions for labels
    assert ws.cell(row=1, column=1).value == "Run 1"
    # Row 1 (label) -> row 2 (image) -> row += 34 -> row 36 (label 2)
    assert ws.cell(row=36, column=1).value == "Run 2"

    assert len(ws._images) == 2
    expected_extent = pixels_to_EMU(480)
    actual_anchors = [
        (image.anchor._from.row, image.anchor.ext.width, image.anchor.ext.height)
        for image in ws._images
    ]
    assert actual_anchors == [
        (1, expected_extent, expected_extent),
        (36, expected_extent, expected_extent),
    ]

def test_module_has_no_qt_import() -> None:
    """
    Static check: qa_xlsx_export.py must not import any Qt package, so it
    stays usable from headless/batch contexts (see module boundary in the
    plan: qa_export.py / qa_xlsx_export.py are both Qt-free).
    """
    source = (_SRC_ROOT / "qa" / "qa_xlsx_export.py").read_text(encoding="utf-8")
    tree = ast.parse(source)
    qt_markers = ("PySide6", "PyQt5", "PyQt6", "PySide2")
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                assert not any(marker in alias.name for marker in qt_markers)
        elif isinstance(node, ast.ImportFrom) and node.module:
            assert not any(marker in node.module for marker in qt_markers)


def test_module_not_already_imported_qt() -> None:
    """qa_xlsx_export must be importable without pulling in a Qt binding."""
    for mod_name in list(sys.modules):
        if mod_name.startswith(("PySide6", "PyQt5", "PyQt6", "PySide2")):
            pytest.skip("A Qt binding is already imported in this test session.")
    import qa.qa_xlsx_export  # noqa: F401

    assert not any(m.startswith(("PySide6", "PyQt5", "PyQt6", "PySide2")) for m in sys.modules)


def test_analyzed_image_path_does_not_leak_into_json_or_csv() -> None:
    """
    Regression guard for the F3 leak-safety claim: the transient
    QAResult.analyzed_image_path field must never appear in the JSON or CSV
    single-run exports (both name their fields explicitly).
    """
    secret_path = "/tmp/definitely-not-in-exports/analyzed_image_abc123.png"
    result = _result(cnr_details=_CNR_DETAILS, analyzed_image_path=secret_path)

    doc = build_single_run_document(result, app_version="9.9.9", inputs={})
    assert secret_path not in repr(doc)

    csv_text = build_metrics_csv(result)
    assert secret_path not in csv_text


# ---------------------------------------------------------------------------
# P2-X2 — Detail sheet uses full flatten + formula neutralization
# ---------------------------------------------------------------------------


def _raw_pylinac_result() -> QAResult:
    """Synthetic run with a nested raw_pylinac tree + curated metrics."""
    return QAResult(
        success=True,
        analysis_type="acr_ct",
        metrics={
            "low_contrast_cnr": {
                "cnr": 4.25,
                "object_rois": [{"mean": 105.0}, {"mean": 95.0}],
                "background": {"mean": 12.0, "std": 1.5},
            },
            "low_contrast_score": 1,
            "num_images": 40,
            "phantom_roll": 0.31,
            "origin_slice": 5,
        },
        warnings=["=cmd"],
        errors=[],
        raw_pylinac={
            "phantom_model": "ACR CT 464",
            "phantom_roll_deg": 0.31,
            "origin_slice": 5,
            "num_images": 40,
            "ct_module": {
                "offset": 0.0,
                "rois": {"Air": -987.1, "Water": 5.2},
            },
            "uniformity_module": {
                "offset": 70.0,
                "rois": {"Center": 3.0},
                "center_roi_stdev": 4.2,
            },
        },
        study_uid="1.2.3.4",
        series_uid="1.2.3.4.5",
        modality="CT",
        num_images=40,
        pylinac_version="3.43.2",
        analyzed_image_path="/tmp/analyzed_image.png",
    )


def test_detail_row_count_matches_build_metric_rows() -> None:
    """Detail body rows must match build_metric_rows for a synthetic raw_pylinac tree."""
    result = _raw_pylinac_result()
    expected_rows = list(build_metric_rows(result))
    wb = build_qa_workbook([result], labels=["Run A"])
    detail = wb["Detail"]

    # Count rows that carry a metric key (column 1 non-blank) AND a non-empty
    # value (column 2 not blank). Label rows have a blank column 2, so they
    # are excluded; blank separator rows have both blank. This count must
    # equal the number of build_metric_rows entries.
    metric_row_count = sum(
        1
        for row in detail.iter_rows(min_row=2)
        if row[0].value and row[1].value not in (None, "")
    )
    assert metric_row_count == len(expected_rows)


def test_detail_row_count_exceeds_metrics_only_flatten() -> None:
    """Full flatten Detail must have more metric rows than metrics-only walk."""
    result = _raw_pylinac_result()
    from qa.qa_export import flatten_metrics

    metrics_only_count = len(list(flatten_metrics(result.metrics or {})))
    full_count = len(list(build_metric_rows(result)))
    assert full_count > metrics_only_count


def test_detail_formula_like_series_label_neutralized() -> None:
    """A Series/Run label beginning with '=' must be neutralized in Detail."""
    result = _raw_pylinac_result()
    wb = build_qa_workbook([result], labels=["=Run1"])
    detail = wb["Detail"]
    # The label row is the first row after the header.
    label_cell = detail.cell(row=2, column=1).value
    assert label_cell == "'=Run1"


def test_summary_formula_like_series_label_neutralized() -> None:
    """A Series/Run label beginning with '=' must be neutralized in Summary."""
    result = _raw_pylinac_result()
    wb = build_qa_workbook([result], labels=["=Run1"])
    summary = wb["Summary"]
    # Row 2, column 1 = first data row, Series/Run ID.
    assert summary.cell(row=2, column=1).value == "'=Run1"


def test_summary_formula_like_warning_neutralized() -> None:
    """A warning beginning with '=' must be neutralized in Summary."""
    result = _raw_pylinac_result()
    wb = build_qa_workbook([result], labels=["Run1"])
    summary = wb["Summary"]
    # Column 7 = Warnings.
    assert summary.cell(row=2, column=7).value == "'=cmd"


def test_detail_analyzed_image_path_not_present() -> None:
    """The path denylist must hold in the Detail sheet."""
    result = _raw_pylinac_result()
    wb = build_qa_workbook([result], labels=["Run1"])
    detail_values = [c.value for row in wb["Detail"].iter_rows() for c in row]
    assert "analyzed_image_path" not in detail_values
    assert "/tmp/analyzed_image.png" not in [str(v) for v in detail_values if v]


def test_detail_formula_like_metric_value_neutralized() -> None:
    """Formula-like metric values in raw_pylinac must be neutralized in Detail."""
    result = QAResult(
        success=True,
        analysis_type="acr_ct",
        metrics={"safe": "ok"},
        raw_pylinac={"phantom_model": "=SUM(A1)", "origin_slice": 3},
    )
    wb = build_qa_workbook([result], labels=["Run1"])
    detail = wb["Detail"]
    # Find the phantom_model row and confirm neutralization.
    model_rows = [
        row for row in detail.iter_rows(min_row=2)
        if row[0].value == "phantom_model"
    ]
    assert len(model_rows) == 1
    assert model_rows[0][1].value == "'=SUM(A1)"


def test_images_formula_like_series_label_neutralized(tmp_path: Path) -> None:
    """A Series/Run label beginning with '=' must be neutralized on Images."""
    try:
        from PIL import Image as PILImage
    except ImportError:
        pytest.skip("Pillow is not installed")

    png_path = tmp_path / "analyzed.png"
    PILImage.new("RGB", (16, 16), color="red").save(png_path)
    result = _result(analyzed_image_path=str(png_path))
    wb = build_qa_workbook([result], labels=["=Run1"])
    images = wb["Images"]
    assert images.cell(row=1, column=1).value == "'=Run1"


# ---------------------------------------------------------------------------
# P2-X3 — Images sheet multi-module embed from analyzed_module_images
# ---------------------------------------------------------------------------


def _write_png(path: Path, size: tuple[int, int] = (16, 16)) -> None:
    """Write a tiny valid PNG (Pillow required; skip caller if missing)."""
    from PIL import Image as PILImage

    PILImage.new("RGB", size, color="red").save(path)


def test_images_sheet_embeds_multiple_module_pngs(tmp_path: Path) -> None:
    """P2-X3: multiple module PNGs per run embed in stable key order."""
    try:
        from PIL import Image as PILImage  # noqa: F401
    except ImportError:
        pytest.skip("Pillow is not installed")

    hu = tmp_path / "hu.png"
    mtf = tmp_path / "mtf.png"
    uniformity = tmp_path / "uniformity.png"
    _write_png(hu)
    _write_png(mtf)
    _write_png(uniformity)

    # Deliberately unsorted keys to verify stable str sort.
    module_images = {
        "uniformity": str(uniformity),
        "hu": str(hu),
        "mtf": str(mtf),
    }
    result = _result(analyzed_module_images=module_images)
    wb = build_qa_workbook([result], labels=["Run 1"])
    reloaded = _save_and_reload(wb)
    assert "Images" in reloaded.sheetnames
    ws = reloaded["Images"]

    # Stable str sort -> hu, mtf, uniformity.
    # Row 1: Series/Run label. Row 2: "hu" label. Row 3: hu image (stride 34).
    # Row 37: "mtf" label. Row 38: mtf image. Row 72: "uniformity" label.
    # Row 73: uniformity image.
    assert ws.cell(row=1, column=1).value == "Run 1"
    assert ws.cell(row=2, column=1).value == "hu"
    assert ws.cell(row=37, column=1).value == "mtf"
    assert ws.cell(row=72, column=1).value == "uniformity"
    assert len(ws._images) == 3


def test_images_sheet_module_label_formula_like_neutralized(tmp_path: Path) -> None:
    """P2-X3: a formula-like module label must be neutralized on Images."""
    try:
        from PIL import Image as PILImage  # noqa: F401
    except ImportError:
        pytest.skip("Pillow is not installed")

    p = tmp_path / "=cmd.png"
    _write_png(p)
    result = _result(analyzed_module_images={"=SUM(A1)": str(p)})
    wb = build_qa_workbook([result], labels=["Run1"])
    ws = wb["Images"]
    # Row 1: Series/Run label; Row 2: the (neutralized) module label.
    assert ws.cell(row=2, column=1).value == "'=SUM(A1)"


def test_images_sheet_skips_missing_module_files(tmp_path: Path) -> None:
    """P2-X3: module paths that don't exist on disk are dropped."""
    try:
        from PIL import Image as PILImage  # noqa: F401
    except ImportError:
        pytest.skip("Pillow is not installed")

    good = tmp_path / "hu.png"
    _write_png(good)
    missing = tmp_path / "gone.png"  # never created
    module_images = {"hu": str(good), "missing": str(missing)}
    result = _result(analyzed_module_images=module_images)
    wb = build_qa_workbook([result], labels=["Run1"])
    ws = wb["Images"]
    # Only the existing file embeds.
    assert len(ws._images) == 1
    assert ws.cell(row=2, column=1).value == "hu"


def test_images_sheet_empty_module_dict_skips_with_note(tmp_path: Path) -> None:
    """P2-X3: empty analyzed_module_images + no composite -> skip + note."""
    result = _result(analyzed_module_images={}, analyzed_image_path=None)
    wb = build_qa_workbook([result], labels=["Run1"])
    reloaded = _save_and_reload(wb)
    assert "Images" not in reloaded.sheetnames
    summary_values = [c.value for row in reloaded["Summary"].iter_rows() for c in row]
    assert any(
        isinstance(v, str) and "Images sheet skipped" in v for v in summary_values
    )


def test_images_sheet_falls_back_to_composite_when_no_module_images(
    tmp_path: Path,
) -> None:
    """P2-X3: legacy composite embed when module dict empty but composite set."""
    try:
        from PIL import Image as PILImage  # noqa: F401
    except ImportError:
        pytest.skip("Pillow is not installed")

    composite = tmp_path / "composite.png"
    _write_png(composite)
    result = _result(
        analyzed_module_images={},
        analyzed_image_path=str(composite),
    )
    wb = build_qa_workbook([result], labels=["Run 1"])
    ws = wb["Images"]
    # Composite fallback: one image, no per-module label row between run label
    # and the image (module_label is "" -> no label row). Row 1 = run label,
    # row 2 = image.
    assert len(ws._images) == 1
    assert ws.cell(row=1, column=1).value == "Run 1"


def test_images_sheet_mixed_modules_and_composite(tmp_path: Path) -> None:
    """P2-X3: modules on run 1 and composite-only on run 2; stride stays 34."""
    try:
        from PIL import Image as PILImage  # noqa: F401
    except ImportError:
        pytest.skip("Pillow is not installed")

    hu = tmp_path / "hu.png"
    composite = tmp_path / "composite.png"
    _write_png(hu)
    _write_png(composite)
    res1 = _result(analyzed_module_images={"hu": str(hu)})
    res2 = _result(
        analyzed_module_images={},
        analyzed_image_path=str(composite),
    )
    wb = build_qa_workbook([res1, res2], labels=["Run 1", "Run 2"])
    ws = wb["Images"]
    # Run 1: row 1 label, row 2 "hu", row 3 image, row += 34 -> 37.
    # Run 2 composite: no module-label row; row 37 label, row 38 image.
    assert ws.cell(row=1, column=1).value == "Run 1"
    assert ws.cell(row=2, column=1).value == "hu"
    assert ws.cell(row=37, column=1).value == "Run 2"
    assert len(ws._images) == 2


def test_images_sheet_placeholder_for_run_without_images(tmp_path: Path) -> None:
    """P2-X3: when the sheet exists, a run with no image still gets a placeholder."""
    try:
        from PIL import Image as PILImage  # noqa: F401
    except ImportError:
        pytest.skip("Pillow is not installed")

    hu = tmp_path / "hu.png"
    _write_png(hu)
    res1 = _result(analyzed_module_images={"hu": str(hu)})
    res2 = _result(analyzed_module_images={}, analyzed_image_path=None)
    wb = build_qa_workbook([res1, res2], labels=["Run 1", "Run 2"])
    ws = wb["Images"]
    assert ws.cell(row=1, column=1).value == "Run 1"
    assert ws.cell(row=2, column=1).value == "hu"
    # Image at row 3, stride 34 -> next run label at row 37.
    assert ws.cell(row=37, column=1).value == "Run 2"
    assert ws.cell(row=38, column=1).value == "(no analyzed image for this run)"
    assert len(ws._images) == 1


def test_images_formula_like_run_label_neutralized_on_images(tmp_path: Path) -> None:
    """P2-X4: a formula-like run label is neutralized on the Images sheet."""
    try:
        from PIL import Image as PILImage  # noqa: F401
    except ImportError:
        pytest.skip("Pillow is not installed")

    p = tmp_path / "hu.png"
    _write_png(p)
    result = _result(analyzed_module_images={"hu": str(p)})
    wb = build_qa_workbook([result], labels=["=Run1"])
    ws = wb["Images"]
    assert ws.cell(row=1, column=1).value == "'=Run1"


def test_module_images_paths_do_not_leak_into_json_or_csv(tmp_path: Path) -> None:
    """P2-X4: module-image paths must not appear in JSON/CSV exports."""
    try:
        from PIL import Image as PILImage  # noqa: F401
    except ImportError:
        pytest.skip("Pillow is not installed")

    hu = tmp_path / "hu.png"
    mtf = tmp_path / "mtf.png"
    _write_png(hu)
    _write_png(mtf)
    module_images = {"hu": str(hu), "mtf": str(mtf)}
    result = _result(analyzed_module_images=module_images)

    doc = build_single_run_document(result, app_version="9.9.9", inputs={})
    doc_blob = repr(doc)
    for path in module_images.values():
        assert path not in doc_blob

    csv_text = build_metrics_csv(result)
    for path in module_images.values():
        assert path not in csv_text
    assert "analyzed_module_images" not in csv_text


def test_images_composite_stride_still_passes(tmp_path: Path) -> None:
    """P2-X4: existing composite stride test still passes (regression guard)."""
    try:
        from PIL import Image as PILImage
    except ImportError:
        pytest.skip("Pillow is not installed")

    img_path1 = tmp_path / "img1.png"
    img_path2 = tmp_path / "img2.png"
    PILImage.new("RGB", (200, 100), color="red").save(img_path1)
    PILImage.new("RGB", (300, 300), color="blue").save(img_path2)

    res1 = _result(series_uid="1.1", analyzed_image_path=str(img_path1))
    res2 = _result(series_uid="1.2", analyzed_image_path=str(img_path2))

    workbook = build_qa_workbook([res1, res2], labels=["Run 1", "Run 2"])
    reloaded = _save_and_reload(workbook)
    assert "Images" in reloaded.sheetnames
    ws = reloaded["Images"]

    assert ws.cell(row=1, column=1).value == "Run 1"
    assert ws.cell(row=36, column=1).value == "Run 2"
    assert len(ws._images) == 2
