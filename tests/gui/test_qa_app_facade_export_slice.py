"""Focused QA facade tests for preflight, summaries, and local exports."""

from __future__ import annotations

import json
from types import SimpleNamespace
from unittest.mock import MagicMock

from PySide6.QtWidgets import QMessageBox

from gui.qa_app_facade import QAAppFacade
from qa.analysis_types import QAResult


def _app(save_path: str = "") -> SimpleNamespace:
    return SimpleNamespace(
        main_window=MagicMock(),
        _prompt_save_path=MagicMock(return_value=save_path),
    )


def _message_box(monkeypatch, response: int) -> MagicMock:
    """Install a modal-message fake while retaining real enum values."""
    box = MagicMock()
    box.windowFlags.return_value = 0
    box.exec.return_value = response
    message_box = MagicMock(return_value=box)
    message_box.Icon = QMessageBox.Icon
    message_box.StandardButton = QMessageBox.StandardButton
    monkeypatch.setattr("gui.qa_app_facade.QMessageBox", message_box)
    return box


def test_preflight_warnings_include_folder_and_modality_context() -> None:
    facade = QAAppFacade(_app())

    warnings = facade.build_preflight_warnings(
        expected_modality="CT",
        use_focused=False,
        folder_path="/synthetic/input",
        datasets=[],
        modality="MR",
    )

    assert len(warnings) == 2
    assert "folder input" in warnings[0]
    assert "targets CT" in warnings[1]


def test_user_confirms_preflight_only_when_yes(monkeypatch) -> None:
    app = _app()
    facade = QAAppFacade(app)
    box = _message_box(monkeypatch, int(QMessageBox.StandardButton.No))

    assert facade.user_confirms_preflight(["Synthetic warning"]) is False
    assert "Synthetic warning" in box.setText.call_args.args[0]
    assert facade.user_confirms_preflight([]) is True


def test_show_result_dialog_includes_failure_details(monkeypatch) -> None:
    facade = QAAppFacade(_app())
    box = _message_box(monkeypatch, int(QMessageBox.StandardButton.Ok))
    result = QAResult(
        success=False,
        analysis_type="acr_ct",
        study_uid="1.2.3",
        series_uid="1.2.4",
        num_images=4,
        pylinac_version="3.0",
        warnings=["Synthetic warning"],
        errors=["Synthetic failure"],
        pylinac_analysis_profile={"vanilla_pylinac": False, "vanilla_equivalent": False},
    )

    facade.show_qa_result_dialog("QA result", result)

    text = box.setText.call_args.args[0]
    assert "Analysis failed." in text
    assert "Synthetic warning" in text
    assert "Synthetic failure" in text
    box.setIcon.assert_called_once_with(QMessageBox.Icon.Warning)


def test_offer_pdf_opens_only_after_confirmation(monkeypatch) -> None:
    facade = QAAppFacade(_app())
    _message_box(monkeypatch, int(QMessageBox.StandardButton.Yes))
    open_path = MagicMock()
    monkeypatch.setattr(QAAppFacade, "open_path_in_system_viewer", open_path)
    result = QAResult(success=True, analysis_type="acr_ct", pdf_report_path="report.pdf")

    facade.offer_open_single_run_pdf(result)

    open_path.assert_called_once_with("report.pdf")


def test_export_qa_results_writes_json_and_updates_status(tmp_path) -> None:
    output = tmp_path / "qa-result.json"
    app = _app(str(output))
    result = QAResult(success=True, analysis_type="acr_ct", metrics={"score": 1})

    QAAppFacade(app).export_qa_results(result, "qa-result", inputs={"synthetic": True})

    assert json.loads(output.read_text(encoding="utf-8"))["run"]["status"] == "success"
    app.main_window.update_status.assert_called_once_with(f"Saved QA JSON: {output}")


def test_export_qa_results_writes_csv_and_updates_status(tmp_path) -> None:
    output = tmp_path / "qa-result.csv"
    app = _app(str(output))
    result = QAResult(success=True, analysis_type="acr_ct", metrics={"score": 1})

    QAAppFacade(app).export_qa_results(result, "qa-result")

    assert "score,1" in output.read_text(encoding="utf-8")
    app.main_window.update_status.assert_called_once_with(f"Saved QA CSV: {output}")


# ---------------------------------------------------------------------------
# P3-C2 — single-run save-dialog CSV/XLSX pick up the full flatten (no UI
# change). Synthetic QAResult fixtures carry a curated-metrics overlay that
# must win on collision, a nested raw_pylinac tree with a dotted leaf, and
# denylisted path fields that must never reach CSV / XLSX Detail.
# ---------------------------------------------------------------------------


def _acr_ct_result() -> QAResult:
    """Synthetic ACR CT run: curated overlay + nested raw_pylinac + denied paths."""
    return QAResult(
        success=True,
        analysis_type="acr_ct",
        metrics={
            "low_contrast_cnr": 4.25,
            "low_contrast_score": 1,
            "num_images": 40,
        },
        raw_pylinac={
            "phantom_model": "ACR CT 464",
            # Same top-level key as metrics — flatten overlay must keep 4.25.
            "low_contrast_cnr": 999.0,
            "uniformity_module": {
                "offset": 70.0,
                "piu": 0.98,
                "rois": {"Center": 3.0},
            },
            "ct_module": {
                "offset": 0.0,
                "rois": {"Air": -987.1},
            },
            # Denylist must skip this even when it lives under raw_pylinac.
            "analyzed_image_path": "/tmp/raw_pylinac_ct.png",
        },
        num_images=40,
        analyzed_image_path="/tmp/ct_analyzed.png",
        analyzed_module_images={"uniformity": "/tmp/ct_uniformity.png"},
    )


def _acr_mri_result() -> QAResult:
    """Synthetic ACR MRI Large run: curated overlay + nested raw_pylinac + denied paths."""
    return QAResult(
        success=True,
        analysis_type="acr_mri_large",
        metrics={
            "low_contrast_cnr": 3.10,
            "low_contrast_score": 2,
            "num_images": 11,
        },
        raw_pylinac={
            "phantom_model": "ACR MRI Large",
            "low_contrast_score": 99,
            "uniformity_module": {
                "offset": 50.0,
                "piu": 0.95,
            },
            "geometry_module": {
                "horizontal_distortion_mm": 1.2,
            },
            "analyzed_module_images": {"hu": "/tmp/raw_pylinac_mri.png"},
        },
        num_images=11,
        analyzed_image_path="/tmp/mri_analyzed.png",
        analyzed_module_images={"uniformity": "/tmp/mri_uniformity.png"},
    )


def test_export_qa_results_csv_full_flatten_ct(tmp_path) -> None:
    """CSV via the save-dialog path carries the nested raw_pylinac leaf for CT."""
    output = tmp_path / "ct.csv"
    app = _app(str(output))

    QAAppFacade(app).export_qa_results(_acr_ct_result(), "ct")

    text = output.read_text(encoding="utf-8")
    assert text.startswith("metric,value\n")
    assert "uniformity_module.piu,0.98" in text
    assert "ct_module.rois.Air,-987.1" in text
    # Curated overlay wins over the colliding raw_pylinac leaf (999.0).
    assert "low_contrast_cnr,4.25" in text
    assert "low_contrast_cnr,999" not in text
    # Denylisted path fields must not leak into CSV (QAResult or raw_pylinac).
    assert "analyzed_image_path" not in text
    assert "/tmp/ct_analyzed.png" not in text
    assert "/tmp/raw_pylinac_ct.png" not in text
    assert "analyzed_module_images" not in text
    app.main_window.update_status.assert_called_once_with(f"Saved QA CSV: {output}")


def test_export_qa_results_csv_full_flatten_mri(tmp_path) -> None:
    """CSV via the save-dialog path carries the nested raw_pylinac leaf for MRI."""
    output = tmp_path / "mri.csv"
    app = _app(str(output))

    QAAppFacade(app).export_qa_results(_acr_mri_result(), "mri")

    text = output.read_text(encoding="utf-8")
    assert "uniformity_module.piu,0.95" in text
    assert "geometry_module.horizontal_distortion_mm,1.2" in text
    assert "low_contrast_score,2" in text
    assert "low_contrast_score,99" not in text
    assert "analyzed_image_path" not in text
    assert "/tmp/mri_analyzed.png" not in text
    assert "/tmp/raw_pylinac_mri.png" not in text
    assert "analyzed_module_images" not in text
    app.main_window.update_status.assert_called_once_with(f"Saved QA CSV: {output}")


def test_export_qa_results_xlsx_detail_full_flatten_ct(tmp_path) -> None:
    """XLSX Detail sheet (save-dialog path) carries the nested raw_pylinac leaf for CT."""
    import openpyxl

    output = tmp_path / "ct.xlsx"
    app = _app(str(output))

    QAAppFacade(app).export_qa_results(_acr_ct_result(), "ct")

    wb = openpyxl.load_workbook(output)
    assert "Detail" in wb.sheetnames
    detail = wb["Detail"]
    pairs = {row[0]: row[1] for row in detail.iter_rows(values_only=True) if row[0]}
    assert "uniformity_module.piu" in pairs
    assert "ct_module.rois.Air" in pairs
    assert pairs["low_contrast_cnr"] == 4.25
    assert "analyzed_image_path" not in pairs
    assert "analyzed_module_images" not in pairs
    app.main_window.update_status.assert_called_once_with(f"Saved QA XLSX: {output}")


def test_export_qa_results_xlsx_detail_full_flatten_mri(tmp_path) -> None:
    """XLSX Detail sheet (save-dialog path) carries the nested raw_pylinac leaf for MRI."""
    import openpyxl

    output = tmp_path / "mri.xlsx"
    app = _app(str(output))

    QAAppFacade(app).export_qa_results(_acr_mri_result(), "mri")

    wb = openpyxl.load_workbook(output)
    detail = wb["Detail"]
    pairs = {row[0]: row[1] for row in detail.iter_rows(values_only=True) if row[0]}
    assert "uniformity_module.piu" in pairs
    assert "geometry_module.horizontal_distortion_mm" in pairs
    assert pairs["low_contrast_score"] == 2
    assert "analyzed_image_path" not in pairs
    assert "analyzed_module_images" not in pairs
    app.main_window.update_status.assert_called_once_with(f"Saved QA XLSX: {output}")


def test_export_qa_results_json_has_no_metrics_flat(tmp_path) -> None:
    """JSON schema must not gain metrics_flat (no schema churn)."""
    output = tmp_path / "ct.json"
    app = _app(str(output))

    QAAppFacade(app).export_qa_results(_acr_ct_result(), "ct")

    doc = json.loads(output.read_text(encoding="utf-8"))
    assert "metrics_flat" not in doc
    assert "metrics_flat" not in json.dumps(doc)
    app.main_window.update_status.assert_called_once_with(f"Saved QA JSON: {output}")
