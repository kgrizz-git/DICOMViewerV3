"""
DICOM Tag Export Dialog

This module provides a dialog for selecting and exporting DICOM tags to Excel or CSV.
Features multi-series selection and hierarchical tag selection with search.

Inputs:
    - Studies and series data (dict)
    - pydicom.Dataset objects
    
Outputs:
    - Excel or CSV files with exported tags
    
Requirements:
    - PySide6 for dialog components
    - openpyxl for Excel export
    - csv module (standard library) for CSV export
    - DICOMParser for tag extraction
"""

from PySide6.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QLabel,
                                QTreeWidget, QTreeWidgetItem, QLineEdit,
                                QPushButton, QCheckBox, QGroupBox, QSplitter,
                                QFileDialog, QMessageBox, QComboBox)
from PySide6.QtCore import Qt, Signal
from typing import Optional, Dict, Any, List
import pydicom
from pydicom.dataset import Dataset
from pathlib import Path
import csv
import os

from core.dicom_parser import DICOMParser


class TagExportDialog(QDialog):
    """
    Dialog for exporting DICOM tags to Excel or CSV with series and tag selection.
    
    Features:
    - Multi-series selection grouped by study
    - Hierarchical tag selection (by group)
    - Search/filter tags
    - Export to Excel (one tab per study, one row per tag) or CSV (one file per study)
    - Tag selection presets (save/load/delete)
    """
    
    def __init__(self, studies: Dict[str, Dict[str, List[Dataset]]], config_manager=None, parent=None):
        """
        Initialize the tag export dialog.
        
        Args:
            studies: Dictionary of studies {study_uid: {series_uid: [datasets]}}
            config_manager: Optional ConfigManager instance for preset storage
            parent: Parent widget
        """
        super().__init__(parent)
        
        self.setWindowTitle("Export DICOM Tags")
        self.setModal(True)
        self.resize(1000, 700)
        
        self.studies = studies
        self.config_manager = config_manager
        self.selected_series: Dict[str, Dict[str, List[int]]] = {}  # {study_uid: {series_uid: [instance_indices]}}
        self.selected_tags: List[str] = []  # List of selected tag strings
        
        self._create_ui()
        self._populate_series()
        self._populate_tags()
        self._load_presets_list()
    
    def _create_ui(self) -> None:
        """Create the UI components."""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        
        # Main splitter for two panels
        splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Left panel: Series selection
        series_panel = self._create_series_panel()
        splitter.addWidget(series_panel)
        
        # Right panel: Tag selection
        tag_panel = self._create_tag_panel()
        splitter.addWidget(tag_panel)
        
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 2)
        
        layout.addWidget(splitter)
        
        # Bottom buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        export_button = QPushButton("Export Tags...")
        export_button.clicked.connect(self._export_to_excel)
        button_layout.addWidget(export_button)
        
        cancel_button = QPushButton("Cancel")
        cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(cancel_button)
        
        layout.addLayout(button_layout)
    
    def _create_series_panel(self) -> QGroupBox:
        """Create the series selection panel."""
        group = QGroupBox("Select Series")
        layout = QVBoxLayout()
        
        # Select/Deselect buttons
        button_layout = QHBoxLayout()
        select_all_btn = QPushButton("Select All")
        select_all_btn.clicked.connect(lambda: self._toggle_all_series(True))
        button_layout.addWidget(select_all_btn)
        
        deselect_all_btn = QPushButton("Deselect All")
        deselect_all_btn.clicked.connect(lambda: self._toggle_all_series(False))
        button_layout.addWidget(deselect_all_btn)
        button_layout.addStretch()
        
        layout.addLayout(button_layout)
        
        # Series tree
        self.series_tree = QTreeWidget()
        self.series_tree.setHeaderLabels(["Series"])
        self.series_tree.setColumnWidth(0, 350)
        self.series_tree.itemChanged.connect(self._on_series_selection_changed)
        layout.addWidget(self.series_tree)
        
        group.setLayout(layout)
        return group
    
    def _create_tag_panel(self) -> QGroupBox:
        """Create the tag selection panel."""
        group = QGroupBox("Select Tags to Export")
        layout = QVBoxLayout()
        
        # Preset management section
        if self.config_manager:
            preset_layout = QHBoxLayout()
            preset_label = QLabel("Preset:")
            self.preset_combo = QComboBox()
            self.preset_combo.setEditable(False)
            self.preset_combo.currentTextChanged.connect(self._on_preset_selected)
            preset_layout.addWidget(preset_label)
            preset_layout.addWidget(self.preset_combo)
            
            save_preset_btn = QPushButton("Save As...")
            save_preset_btn.clicked.connect(self._save_preset)
            preset_layout.addWidget(save_preset_btn)
            
            load_preset_btn = QPushButton("Load")
            load_preset_btn.clicked.connect(self._load_preset)
            preset_layout.addWidget(load_preset_btn)
            
            delete_preset_btn = QPushButton("Delete")
            delete_preset_btn.clicked.connect(self._delete_preset)
            preset_layout.addWidget(delete_preset_btn)

            # Export/Import presets buttons (JSON)
            export_presets_btn = QPushButton("Export...")
            export_presets_btn.clicked.connect(self._export_presets)
            preset_layout.addWidget(export_presets_btn)

            import_presets_btn = QPushButton("Import...")
            import_presets_btn.clicked.connect(self._import_presets)
            preset_layout.addWidget(import_presets_btn)
            
            preset_layout.addStretch()
            layout.addLayout(preset_layout)
        
        # Search box
        search_layout = QHBoxLayout()
        search_label = QLabel("Filter:")
        self.tag_search = QLineEdit()
        self.tag_search.setPlaceholderText("Search tags...")
        self.tag_search.textChanged.connect(self._filter_tags)
        search_layout.addWidget(search_label)
        search_layout.addWidget(self.tag_search)
        layout.addLayout(search_layout)
        
        # Select/Deselect buttons
        button_layout = QHBoxLayout()
        select_all_btn = QPushButton("Select All")
        select_all_btn.clicked.connect(lambda: self._toggle_all_tags(True))
        button_layout.addWidget(select_all_btn)
        
        deselect_all_btn = QPushButton("Deselect All")
        deselect_all_btn.clicked.connect(lambda: self._toggle_all_tags(False))
        button_layout.addWidget(deselect_all_btn)
        button_layout.addStretch()
        
        # Show private tags checkbox
        self.private_tags_checkbox = QCheckBox("Include Private Tags")
        self.private_tags_checkbox.setChecked(True)
        self.private_tags_checkbox.toggled.connect(self._populate_tags)
        button_layout.addWidget(self.private_tags_checkbox)
        
        layout.addLayout(button_layout)
        
        # Tags tree
        self.tags_tree = QTreeWidget()
        self.tags_tree.setHeaderLabels(["Tag", "Name"])
        self.tags_tree.setColumnWidth(0, 120)
        self.tags_tree.setColumnWidth(1, 300)
        self.tags_tree.itemChanged.connect(self._on_tag_selection_changed)
        layout.addWidget(self.tags_tree)
        
        group.setLayout(layout)
        return group
    
    def _populate_series(self) -> None:
        """Populate the series tree with available series and instances."""
        self.series_tree.clear()
        self.series_tree.blockSignals(True)
        
        for study_uid, series_dict in self.studies.items():
            # Get first dataset to extract study info
            first_series_uid = list(series_dict.keys())[0]
            first_dataset = series_dict[first_series_uid][0]
            
            study_desc = getattr(first_dataset, 'StudyDescription', 'Unknown Study')
            study_date = getattr(first_dataset, 'StudyDate', '')
            if study_date:
                study_date = f" ({study_date})"
            
            # Create study item
            study_item = QTreeWidgetItem(self.series_tree)
            study_item.setText(0, f"{study_desc}{study_date}")
            study_item.setData(0, Qt.ItemDataRole.UserRole, study_uid)
            study_item.setFlags(study_item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            study_item.setCheckState(0, Qt.CheckState.Unchecked)
            study_item.setExpanded(True)
            
            # Add series items
            for series_uid, datasets in series_dict.items():
                first_ds = datasets[0]
                series_num = getattr(first_ds, 'SeriesNumber', '')
                series_desc = getattr(first_ds, 'SeriesDescription', 'Unknown Series')
                modality = getattr(first_ds, 'Modality', '')
                
                series_item = QTreeWidgetItem(study_item)
                series_text = f"Series {series_num}: {series_desc} ({modality}) - {len(datasets)} images"
                series_item.setText(0, series_text)
                series_item.setData(0, Qt.ItemDataRole.UserRole, series_uid)
                series_item.setData(0, Qt.ItemDataRole.UserRole + 1, study_uid)
                series_item.setFlags(series_item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                series_item.setCheckState(0, Qt.CheckState.Unchecked)
                series_item.setExpanded(True)
                
                # Add instance/slice items
                for idx, dataset in enumerate(datasets):
                    instance_num = getattr(dataset, 'InstanceNumber', None)
                    slice_location = getattr(dataset, 'SliceLocation', None)
                    
                    if instance_num is not None:
                        instance_text = f"Instance {instance_num}"
                        if slice_location is not None:
                            instance_text += f" (Slice: {slice_location:.2f})"
                    else:
                        instance_text = f"Instance {idx + 1}"
                        if slice_location is not None:
                            instance_text += f" (Slice: {slice_location:.2f})"
                    
                    instance_item = QTreeWidgetItem(series_item)
                    instance_item.setText(0, instance_text)
                    instance_item.setData(0, Qt.ItemDataRole.UserRole, idx)  # Store instance index
                    instance_item.setData(0, Qt.ItemDataRole.UserRole + 1, series_uid)
                    instance_item.setData(0, Qt.ItemDataRole.UserRole + 2, study_uid)
                    instance_item.setFlags(instance_item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                    instance_item.setCheckState(0, Qt.CheckState.Unchecked)
        
        self.series_tree.blockSignals(False)
    
    def _populate_tags(self) -> None:
        """Populate the tags tree with DICOM tags from a representative dataset."""
        self.tags_tree.clear()
        self.tags_tree.blockSignals(True)
        
        # Get a representative dataset (first series, first image)
        if not self.studies:
            self.tags_tree.blockSignals(False)
            return
        
        first_study_uid = list(self.studies.keys())[0]
        first_series_uid = list(self.studies[first_study_uid].keys())[0]
        first_dataset = self.studies[first_study_uid][first_series_uid][0]
        
        parser = DICOMParser(first_dataset)
        tags = parser.get_all_tags(include_private=self.private_tags_checkbox.isChecked())
        
        # Sort tags by tag number
        sorted_tags = sorted(tags.items(), key=lambda x: x[0])
        
        # Group by tag group (first 4 hex digits)
        groups: Dict[str, list] = {}
        for tag_str, tag_data in sorted_tags:
            group = tag_str[:6]  # e.g., "(0008," for group 0008
            if group not in groups:
                groups[group] = []
            groups[group].append((tag_str, tag_data))
        
        # Create tree items
        for group, tag_list in sorted(groups.items()):
            group_item = QTreeWidgetItem(self.tags_tree)
            group_item.setText(0, f"Group {group[1:5]}")  # Remove parentheses
            group_item.setText(1, f"{len(tag_list)} tags")
            group_item.setFlags(group_item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            group_item.setCheckState(0, Qt.CheckState.Unchecked)
            group_item.setExpanded(False)
            
            for tag_str, tag_data in tag_list:
                tag_item = QTreeWidgetItem(group_item)
                tag_item.setText(0, tag_data.get("tag", tag_str))
                tag_item.setText(1, tag_data.get("name", ""))
                tag_item.setData(0, Qt.ItemDataRole.UserRole, tag_str)
                tag_item.setFlags(tag_item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                tag_item.setCheckState(0, Qt.CheckState.Unchecked)
        
        self.tags_tree.blockSignals(False)
    
    def _toggle_all_series(self, checked: bool) -> None:
        """Toggle all series and instance selection."""
        self.series_tree.blockSignals(True)
        root = self.series_tree.invisibleRootItem()
        for i in range(root.childCount()):
            study_item = root.child(i)
            study_item.setCheckState(0, Qt.CheckState.Checked if checked else Qt.CheckState.Unchecked)
            for j in range(study_item.childCount()):
                series_item = study_item.child(j)
                series_item.setCheckState(0, Qt.CheckState.Checked if checked else Qt.CheckState.Unchecked)
                for k in range(series_item.childCount()):
                    instance_item = series_item.child(k)
                    instance_item.setCheckState(0, Qt.CheckState.Checked if checked else Qt.CheckState.Unchecked)
        self.series_tree.blockSignals(False)
        self._update_selected_series()
    
    def _toggle_all_tags(self, checked: bool) -> None:
        """Toggle all tag selection."""
        self.tags_tree.blockSignals(True)
        root = self.tags_tree.invisibleRootItem()
        for i in range(root.childCount()):
            group_item = root.child(i)
            if not group_item.isHidden():
                group_item.setCheckState(0, Qt.CheckState.Checked if checked else Qt.CheckState.Unchecked)
                for j in range(group_item.childCount()):
                    tag_item = group_item.child(j)
                    if not tag_item.isHidden():
                        tag_item.setCheckState(0, Qt.CheckState.Checked if checked else Qt.CheckState.Unchecked)
        self.tags_tree.blockSignals(False)
        self._update_selected_tags()
    
    def _filter_tags(self, search_text: str) -> None:
        """Filter tags based on search text."""
        search_lower = search_text.lower()
        root = self.tags_tree.invisibleRootItem()
        
        for i in range(root.childCount()):
            group_item = root.child(i)
            group_has_visible_child = False
            
            for j in range(group_item.childCount()):
                tag_item = group_item.child(j)
                tag_text = tag_item.text(0).lower() + " " + tag_item.text(1).lower()
                
                if not search_text or search_lower in tag_text:
                    tag_item.setHidden(False)
                    group_has_visible_child = True
                else:
                    tag_item.setHidden(True)
            
            group_item.setHidden(not group_has_visible_child)
    
    def _on_series_selection_changed(self, item: QTreeWidgetItem, column: int) -> None:
        """Handle series and instance selection changes."""
        # Block signals to prevent recursive calls
        self.series_tree.blockSignals(True)
        
        parent = item.parent()
        
        # If this is a study item, update all series and instances under it
        if parent is None:
            check_state = item.checkState(0)
            for i in range(item.childCount()):
                series_item = item.child(i)
                series_item.setCheckState(0, check_state)
                for j in range(series_item.childCount()):
                    instance_item = series_item.child(j)
                    instance_item.setCheckState(0, check_state)
        # If this is a series item, update all instances under it and parent study
        elif parent.parent() is None:
            check_state = item.checkState(0)
            for i in range(item.childCount()):
                instance_item = item.child(i)
                instance_item.setCheckState(0, check_state)
            
            # Update parent study's check state
            study_item = parent
            all_checked = True
            any_checked = False
            for i in range(study_item.childCount()):
                series_item = study_item.child(i)
                series_state = series_item.checkState(0)
                if series_state == Qt.CheckState.Unchecked:
                    all_checked = False
                else:
                    any_checked = True
            
            if all_checked:
                study_item.setCheckState(0, Qt.CheckState.Checked)
            elif any_checked:
                study_item.setCheckState(0, Qt.CheckState.PartiallyChecked)
            else:
                study_item.setCheckState(0, Qt.CheckState.Unchecked)
        # If this is an instance item, update parent series and study
        else:
            check_state = item.checkState(0)
            
            # Update parent series's check state
            series_item = parent
            all_checked = True
            any_checked = False
            for i in range(series_item.childCount()):
                instance_state = series_item.child(i).checkState(0)
                if instance_state == Qt.CheckState.Unchecked:
                    all_checked = False
                else:
                    any_checked = True
            
            if all_checked:
                series_item.setCheckState(0, Qt.CheckState.Checked)
            elif any_checked:
                series_item.setCheckState(0, Qt.CheckState.PartiallyChecked)
            else:
                series_item.setCheckState(0, Qt.CheckState.Unchecked)
            
            # Update parent study's check state
            study_item = series_item.parent()
            all_checked = True
            any_checked = False
            for i in range(study_item.childCount()):
                series_state = study_item.child(i).checkState(0)
                if series_state == Qt.CheckState.Unchecked:
                    all_checked = False
                else:
                    any_checked = True
            
            if all_checked:
                study_item.setCheckState(0, Qt.CheckState.Checked)
            elif any_checked:
                study_item.setCheckState(0, Qt.CheckState.PartiallyChecked)
            else:
                study_item.setCheckState(0, Qt.CheckState.Unchecked)
        
        self.series_tree.blockSignals(False)
        self._update_selected_series()
    
    def _on_tag_selection_changed(self, item: QTreeWidgetItem, column: int) -> None:
        """Handle tag selection changes."""
        # Block signals to prevent recursive calls
        self.tags_tree.blockSignals(True)
        
        # If this is a group item, update all tags under it
        if item.parent() is None:
            check_state = item.checkState(0)
            for i in range(item.childCount()):
                if not item.child(i).isHidden():
                    item.child(i).setCheckState(0, check_state)
        else:
            # If this is a tag item, update parent group's check state
            parent = item.parent()
            all_checked = True
            any_checked = False
            for i in range(parent.childCount()):
                child = parent.child(i)
                if not child.isHidden():
                    child_state = child.checkState(0)
                    if child_state == Qt.CheckState.Unchecked:
                        all_checked = False
                    else:
                        any_checked = True
            
            if all_checked:
                parent.setCheckState(0, Qt.CheckState.Checked)
            elif any_checked:
                parent.setCheckState(0, Qt.CheckState.PartiallyChecked)
            else:
                parent.setCheckState(0, Qt.CheckState.Unchecked)
        
        self.tags_tree.blockSignals(False)
        self._update_selected_tags()
    
    def _update_selected_series(self) -> None:
        """Update the list of selected series and instances."""
        self.selected_series = {}
        root = self.series_tree.invisibleRootItem()
        
        for i in range(root.childCount()):
            study_item = root.child(i)
            study_uid = study_item.data(0, Qt.ItemDataRole.UserRole)
            
            series_dict = {}
            for j in range(study_item.childCount()):
                series_item = study_item.child(j)
                series_uid = series_item.data(0, Qt.ItemDataRole.UserRole)
                
                instance_list = []
                for k in range(series_item.childCount()):
                    instance_item = series_item.child(k)
                    if instance_item.checkState(0) == Qt.CheckState.Checked:
                        instance_idx = instance_item.data(0, Qt.ItemDataRole.UserRole)
                        instance_list.append(instance_idx)
                
                if instance_list:
                    series_dict[series_uid] = instance_list
            
            if series_dict:
                self.selected_series[study_uid] = series_dict
    
    def _update_selected_tags(self) -> None:
        """Update the list of selected tags."""
        self.selected_tags = []
        root = self.tags_tree.invisibleRootItem()
        
        for i in range(root.childCount()):
            group_item = root.child(i)
            for j in range(group_item.childCount()):
                tag_item = group_item.child(j)
                # Include all checked tags regardless of visibility (filter state)
                if tag_item.checkState(0) == Qt.CheckState.Checked:
                    tag_str = tag_item.data(0, Qt.ItemDataRole.UserRole)
                    self.selected_tags.append(tag_str)
    
    def _export_to_excel(self) -> None:
        """Export selected tags to Excel or CSV file."""
        # Update selected tags and series to ensure they're current
        self._update_selected_tags()
        self._update_selected_series()
        
        # Validate selections
        if not self.selected_series:
            QMessageBox.warning(self, "No Series Selected", 
                              "Please select at least one series to export.")
            return
        
        if not self.selected_tags:
            QMessageBox.warning(self, "No Tags Selected",
                              "Please select at least one tag to export.")
            return
        
        # Analyze tag variations
        variation_analysis = self._analyze_tag_variations()
        
        # Show variation analysis dialog
        if not self._show_variation_analysis_dialog(variation_analysis):
            return  # User cancelled
        
        # Generate default filename
        default_filename = self._generate_default_filename()
        
        # Get last export path if available
        if self.config_manager:
            last_path = self.config_manager.get_last_export_path()
            if last_path:
                # last_path is a directory, use it as the initial directory
                last_path_obj = Path(last_path)
                if last_path_obj.is_dir():
                    default_filename = str(last_path_obj / Path(default_filename).name)
                else:
                    # If it's a file, use its parent directory
                    default_filename = str(last_path_obj.parent / Path(default_filename).name)
        
        # Show file save dialog
        file_path, selected_filter = QFileDialog.getSaveFileName(
            self,
            "Save DICOM Tag Export",
            default_filename,
            "Excel Files (*.xlsx);;CSV Files (*.csv);;All Files (*)"
        )
        
        if not file_path:
            return
        
        # Determine format and ensure correct extension
        is_csv = selected_filter.startswith("CSV") or file_path.endswith('.csv')
        if is_csv:
            if not file_path.endswith('.csv'):
                file_path += '.csv'
        else:
            if not file_path.endswith('.xlsx'):
                file_path += '.xlsx'
        
        # Save export location
        if self.config_manager:
            export_dir = str(Path(file_path).parent)
            self.config_manager.set_last_export_path(export_dir)
        
        # Perform export
        try:
            if is_csv:
                exported_files = self._write_csv_files(file_path, variation_analysis)
                if len(exported_files) > 1:
                    file_list = "\n".join(str(f) for f in exported_files)
                    QMessageBox.information(self, "Export Complete",
                                          f"Tags exported successfully to {len(exported_files)} files:\n{file_list}")
                else:
                    QMessageBox.information(self, "Export Complete",
                                          f"Tags exported successfully to:\n{exported_files[0]}")
            else:
                self._write_excel_file(file_path, variation_analysis)
                QMessageBox.information(self, "Export Complete",
                                  f"Tags exported successfully to:\n{file_path}")
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Export Failed",
                               f"Failed to export tags:\n{str(e)}")
    
    def _analyze_tag_variations(self) -> Dict[str, Dict[str, List[str]]]:
        """
        Analyze which tags vary across instances within each selected series.
        
        Returns:
            Dictionary mapping series_uid to dict with 'varying_tags' and 'constant_tags' lists
        """
        variation_analysis = {}
        
        for study_uid, series_dict in self.selected_series.items():
            for series_uid, instance_indices in series_dict.items():
                if not instance_indices:
                    continue
                
                datasets = self.studies[study_uid][series_uid]
                varying_tags = []
                constant_tags = []
                
                # For each selected tag, check if it varies across instances
                for tag_str in self.selected_tags:
                    tag_values = []
                    
                    # Collect values for this tag across all selected instances
                    for instance_idx in instance_indices:
                        if instance_idx >= len(datasets):
                            continue
                        
                        dataset = datasets[instance_idx]
                        parser = DICOMParser(dataset)
                        all_tags = parser.get_all_tags(include_private=self.private_tags_checkbox.isChecked())
                        
                        if tag_str in all_tags:
                            tag_data = all_tags[tag_str]
                            value = tag_data.get('value', '')
                            
                            # Convert to string for comparison
                            if isinstance(value, list):
                                value_str = ', '.join(str(v) for v in value)
                            elif value is None:
                                value_str = ''
                            else:
                                value_str = str(value)
                            
                            tag_values.append(value_str)
                    
                    # Check if values vary
                    if len(tag_values) > 1:
                        first_value = tag_values[0]
                        varies = any(val != first_value for val in tag_values[1:])
                        if varies:
                            varying_tags.append(tag_str)
                        else:
                            constant_tags.append(tag_str)
                    else:
                        # Single instance or no values - treat as constant
                        constant_tags.append(tag_str)
                
                variation_analysis[series_uid] = {
                    'varying_tags': varying_tags,
                    'constant_tags': constant_tags
                }
        
        return variation_analysis
    
    def _show_variation_analysis_dialog(self, variation_analysis: Dict[str, Dict[str, List[str]]]) -> bool:
        """
        Show dialog displaying tag variation analysis.
        
        Args:
            variation_analysis: Dictionary from _analyze_tag_variations()
            
        Returns:
            True if user confirms export, False if cancelled
        """
        dialog = QDialog(self)
        dialog.setWindowTitle("Tag Variation Analysis")
        dialog.setModal(True)
        dialog.resize(700, 500)
        
        layout = QVBoxLayout(dialog)
        
        # Info label
        info_label = QLabel(
            "Tags varying by instance will be exported per-instance.\n"
            "Constant tags will be exported once per series."
        )
        layout.addWidget(info_label)
        
        # Tree widget to show analysis
        tree = QTreeWidget()
        tree.setHeaderLabels(["Series/Tag", "Status"])
        tree.setColumnWidth(0, 400)
        tree.setColumnWidth(1, 200)
        layout.addWidget(tree)
        
        # Populate tree
        for study_uid, series_dict in self.selected_series.items():
            for series_uid, instance_indices in series_dict.items():
                if series_uid not in variation_analysis:
                    continue
                
                # Get series info
                datasets = self.studies[study_uid][series_uid]
                first_ds = datasets[0] if datasets else None
                if not first_ds:
                    continue
                
                series_num = getattr(first_ds, 'SeriesNumber', '')
                series_desc = getattr(first_ds, 'SeriesDescription', 'Unknown Series')
                series_text = f"Series {series_num}: {series_desc} ({len(instance_indices)} instances selected)"
                
                series_item = QTreeWidgetItem(tree)
                series_item.setText(0, series_text)
                series_item.setExpanded(True)
                
                analysis = variation_analysis[series_uid]
                varying_tags = analysis['varying_tags']
                constant_tags = analysis['constant_tags']
                
                # Varying tags
                if varying_tags:
                    varying_item = QTreeWidgetItem(series_item)
                    varying_item.setText(0, f"Varying Tags ({len(varying_tags)})")
                    varying_item.setText(1, "Per-instance export")
                    
                    parser = DICOMParser(first_ds)
                    all_tags = parser.get_all_tags(include_private=self.private_tags_checkbox.isChecked())
                    
                    for tag_str in sorted(varying_tags):
                        tag_item = QTreeWidgetItem(varying_item)
                        if tag_str in all_tags:
                            tag_data = all_tags[tag_str]
                            tag_name = tag_data.get('name', tag_str)
                            tag_item.setText(0, f"{tag_str} - {tag_name}")
                        else:
                            tag_item.setText(0, tag_str)
                
                # Constant tags
                if constant_tags:
                    constant_item = QTreeWidgetItem(series_item)
                    constant_item.setText(0, f"Constant Tags ({len(constant_tags)})")
                    constant_item.setText(1, "Per-series export")
                    
                    parser = DICOMParser(first_ds)
                    all_tags = parser.get_all_tags(include_private=self.private_tags_checkbox.isChecked())
                    
                    for tag_str in sorted(constant_tags):
                        tag_item = QTreeWidgetItem(constant_item)
                        if tag_str in all_tags:
                            tag_data = all_tags[tag_str]
                            tag_name = tag_data.get('name', tag_str)
                            tag_item.setText(0, f"{tag_str} - {tag_name}")
                        else:
                            tag_item.setText(0, tag_str)
        
        # Buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        export_btn = QPushButton("Continue Export")
        export_btn.clicked.connect(dialog.accept)
        button_layout.addWidget(export_btn)
        
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(dialog.reject)
        button_layout.addWidget(cancel_btn)
        
        layout.addLayout(button_layout)
        
        return dialog.exec() == QDialog.DialogCode.Accepted
    
    def _generate_default_filename(self) -> str:
        """Generate default filename for export."""
        # Get first selected series to extract info
        first_study_uid = list(self.selected_series.keys())[0]
        first_series_uid = list(self.selected_series[first_study_uid].keys())[0]
        first_instance_idx = self.selected_series[first_study_uid][first_series_uid][0]
        first_dataset = self.studies[first_study_uid][first_series_uid][first_instance_idx]
        
        modality = getattr(first_dataset, 'Modality', 'Unknown')
        accession = getattr(first_dataset, 'AccessionNumber', 'Unknown')
        
        # Default to Excel format (user can change in file dialog)
        return f"{modality} DICOM Tag Export {accession}.xlsx"
    
    def _write_excel_file(self, file_path: str, variation_analysis: Dict[str, Dict[str, List[str]]]) -> None:
        """Write selected tags to Excel file with per-instance export for varying tags."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
        except ImportError:
            raise ImportError("openpyxl library is required for Excel export. "
                            "Install it with: pip install openpyxl")
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create one sheet per study
        for study_uid, series_dict in self.selected_series.items():
            # Get study info for sheet name
            first_series_uid = list(series_dict.keys())[0]
            first_instance_idx = series_dict[first_series_uid][0]
            first_dataset = self.studies[study_uid][first_series_uid][first_instance_idx]
            study_desc = getattr(first_dataset, 'StudyDescription', 'Study')
            # Sanitize sheet name (max 31 chars, no special chars)
            sheet_name = study_desc[:31].replace('/', '-').replace('\\', '-').replace(':', '-')
            
            ws = wb.create_sheet(title=sheet_name)
            
            # Write header
            ws['A1'] = 'Instance'
            ws['B1'] = 'Tag Number'
            ws['C1'] = 'Name'
            ws['D1'] = 'Value'
            ws['A1'].font = Font(bold=True)
            ws['B1'].font = Font(bold=True)
            ws['C1'].font = Font(bold=True)
            ws['D1'].font = Font(bold=True)
            
            row = 2
            
            # Export each selected series
            for series_uid, instance_indices in series_dict.items():
                datasets = self.studies[study_uid][series_uid]
                if not datasets:
                    continue
                
                first_ds = datasets[0]
                
                # Write series header
                series_num = getattr(first_ds, 'SeriesNumber', '')
                series_desc = getattr(first_ds, 'SeriesDescription', 'Unknown')
                ws[f'A{row}'] = f"Series {series_num}: {series_desc}"
                ws[f'A{row}'].font = Font(bold=True, italic=True)
                ws.merge_cells(f'A{row}:D{row}')
                row += 1
                
                # Get variation analysis for this series
                analysis = variation_analysis.get(series_uid, {'varying_tags': [], 'constant_tags': self.selected_tags})
                varying_tags = analysis['varying_tags']
                constant_tags = analysis['constant_tags']
                
                # Export constant tags (once per series, using first instance)
                if constant_tags and instance_indices:
                    first_instance_idx = instance_indices[0]
                    if first_instance_idx < len(datasets):
                        dataset = datasets[first_instance_idx]
                        parser = DICOMParser(dataset)
                        all_tags = parser.get_all_tags(include_private=self.private_tags_checkbox.isChecked())
                        
                        for tag_str in constant_tags:
                            if tag_str in all_tags:
                                tag_data = all_tags[tag_str]
                                ws[f'A{row}'] = 'All'  # Indicates all instances
                                ws[f'B{row}'] = tag_data.get('tag', tag_str)
                                ws[f'C{row}'] = tag_data.get('name', '')
                                
                                # Format value
                                value = tag_data.get('value', '')
                                if isinstance(value, list):
                                    value_str = ', '.join(str(v) for v in value)
                                else:
                                    value_str = str(value)
                                ws[f'D{row}'] = value_str
                                
                                row += 1
                
                # Export varying tags (per instance)
                if varying_tags:
                    for instance_idx in instance_indices:
                        if instance_idx >= len(datasets):
                            continue
                        
                        dataset = datasets[instance_idx]
                        parser = DICOMParser(dataset)
                        all_tags = parser.get_all_tags(include_private=self.private_tags_checkbox.isChecked())
                        
                        # Get instance identifier
                        instance_num = getattr(dataset, 'InstanceNumber', None)
                        instance_id = f"Instance {instance_num}" if instance_num is not None else f"Instance {instance_idx + 1}"
                        
                        for tag_str in varying_tags:
                            if tag_str in all_tags:
                                tag_data = all_tags[tag_str]
                                ws[f'A{row}'] = instance_id
                                ws[f'B{row}'] = tag_data.get('tag', tag_str)
                                ws[f'C{row}'] = tag_data.get('name', '')
                                
                                # Format value
                                value = tag_data.get('value', '')
                                if isinstance(value, list):
                                    value_str = ', '.join(str(v) for v in value)
                                else:
                                    value_str = str(value)
                                ws[f'D{row}'] = value_str
                                
                                row += 1
                
                # Add blank row between series
                row += 1
        
        # Adjust column widths
        for sheet in wb.worksheets:
            sheet.column_dimensions['A'].width = 15
            sheet.column_dimensions['B'].width = 15
            sheet.column_dimensions['C'].width = 40
            sheet.column_dimensions['D'].width = 60
        
        # Save file
        wb.save(file_path)
    
    def _write_csv_files(self, base_file_path: str, variation_analysis: Dict[str, Dict[str, List[str]]]) -> List[Path]:
        """
        Write selected tags to CSV files (one file per study) with per-instance export for varying tags.
        
        Returns:
            List of created file paths
        """
        base_path = Path(base_file_path)
        base_name = base_path.stem
        base_dir = base_path.parent
        exported_files = []
        
        # Create one CSV file per study
        for study_uid, series_dict in self.selected_series.items():
            # Get study info for filename
            first_series_uid = list(series_dict.keys())[0]
            first_instance_idx = series_dict[first_series_uid][0]
            first_dataset = self.studies[study_uid][first_series_uid][first_instance_idx]
            study_desc = getattr(first_dataset, 'StudyDescription', 'Study')
            # Sanitize filename
            safe_study_desc = study_desc.replace('/', '-').replace('\\', '-').replace(':', '-')[:50]
            
            # Create filename: base_name_StudyDescription.csv
            if len(self.selected_series) > 1:
                csv_filename = f"{base_name}_{safe_study_desc}.csv"
            else:
                csv_filename = f"{base_name}.csv"
            
            csv_path = base_dir / csv_filename
            
            # Write CSV file
            with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write header
                writer.writerow(['Instance', 'Tag Number', 'Name', 'Value'])
                
                # Export each selected series
                for series_uid, instance_indices in series_dict.items():
                    datasets = self.studies[study_uid][series_uid]
                    if not datasets:
                        continue
                    
                    first_ds = datasets[0]
                    
                    # Write series header
                    series_num = getattr(first_ds, 'SeriesNumber', '')
                    series_desc = getattr(first_ds, 'SeriesDescription', 'Unknown')
                    writer.writerow([f"Series {series_num}: {series_desc}", '', '', ''])
                    
                    # Get variation analysis for this series
                    analysis = variation_analysis.get(series_uid, {'varying_tags': [], 'constant_tags': self.selected_tags})
                    varying_tags = analysis['varying_tags']
                    constant_tags = analysis['constant_tags']
                    
                    # Export constant tags (once per series, using first instance)
                    if constant_tags and instance_indices:
                        first_instance_idx = instance_indices[0]
                        if first_instance_idx < len(datasets):
                            dataset = datasets[first_instance_idx]
                            parser = DICOMParser(dataset)
                            all_tags = parser.get_all_tags(include_private=self.private_tags_checkbox.isChecked())
                            
                            for tag_str in constant_tags:
                                if tag_str in all_tags:
                                    tag_data = all_tags[tag_str]
                                    tag_num = tag_data.get('tag', tag_str)
                                    tag_name = tag_data.get('name', '')
                                    
                                    # Format value
                                    value = tag_data.get('value', '')
                                    if isinstance(value, list):
                                        value_str = ', '.join(str(v) for v in value)
                                    else:
                                        value_str = str(value)
                                    
                                    writer.writerow(['All', tag_num, tag_name, value_str])
                    
                    # Export varying tags (per instance)
                    if varying_tags:
                        for instance_idx in instance_indices:
                            if instance_idx >= len(datasets):
                                continue
                            
                            dataset = datasets[instance_idx]
                            parser = DICOMParser(dataset)
                            all_tags = parser.get_all_tags(include_private=self.private_tags_checkbox.isChecked())
                            
                            # Get instance identifier
                            instance_num = getattr(dataset, 'InstanceNumber', None)
                            instance_id = f"Instance {instance_num}" if instance_num is not None else f"Instance {instance_idx + 1}"
                            
                            for tag_str in varying_tags:
                                if tag_str in all_tags:
                                    tag_data = all_tags[tag_str]
                                    tag_num = tag_data.get('tag', tag_str)
                                    tag_name = tag_data.get('name', '')
                                    
                                    # Format value
                                    value = tag_data.get('value', '')
                                    if isinstance(value, list):
                                        value_str = ', '.join(str(v) for v in value)
                                    else:
                                        value_str = str(value)
                                    
                                    writer.writerow([instance_id, tag_num, tag_name, value_str])
                    
                    # Add blank row between series
                    writer.writerow([])
            
            exported_files.append(csv_path)
        
        return exported_files
    
    def _load_presets_list(self) -> None:
        """Load list of presets into combo box."""
        if not self.config_manager:
            return
        
        self.preset_combo.clear()
        presets = self.config_manager.get_tag_export_presets()
        if presets:
            self.preset_combo.addItems(sorted(presets.keys()))
        self.preset_combo.addItem("(No preset)")
        self.preset_combo.setCurrentIndex(self.preset_combo.count() - 1)
    
    def _on_preset_selected(self, preset_name: str) -> None:
        """Handle preset selection (for future use, e.g., auto-load on selection)."""
        pass
    
    def _save_preset(self) -> None:
        """Save current tag selections as a preset."""
        if not self.config_manager:
            QMessageBox.warning(self, "No Config Manager",
                              "Preset saving is not available.")
            return
        
        # Update selected tags first
        self._update_selected_tags()
        
        if not self.selected_tags:
            QMessageBox.warning(self, "No Tags Selected",
                              "Please select at least one tag to save as a preset.")
            return
        
        # Get preset name from user
        from PySide6.QtWidgets import QInputDialog
        preset_name, ok = QInputDialog.getText(
            self,
            "Save Preset",
            "Enter preset name:",
            text=""
        )
        
        if not ok or not preset_name.strip():
            return
        
        preset_name = preset_name.strip()
        
        # Save preset
        self.config_manager.save_tag_export_preset(preset_name, self.selected_tags)
        self._load_presets_list()
        
        # Select the newly saved preset
        index = self.preset_combo.findText(preset_name)
        if index >= 0:
            self.preset_combo.setCurrentIndex(index)
        
        QMessageBox.information(self, "Preset Saved",
                               f"Preset '{preset_name}' saved successfully.")
    
    def _load_preset(self) -> None:
        """Load a preset and apply tag selections."""
        if not self.config_manager:
            QMessageBox.warning(self, "No Config Manager",
                              "Preset loading is not available.")
            return
        
        preset_name = self.preset_combo.currentText()
        if not preset_name or preset_name == "(No preset)":
            QMessageBox.warning(self, "No Preset Selected",
                              "Please select a preset to load.")
            return
        
        # Get preset tags
        presets = self.config_manager.get_tag_export_presets()
        if preset_name not in presets:
            QMessageBox.warning(self, "Preset Not Found",
                              f"Preset '{preset_name}' not found.")
            return
        
        preset_tags = presets[preset_name]
        
        # Apply preset to tag tree
        self.tags_tree.blockSignals(True)
        root = self.tags_tree.invisibleRootItem()
        
        # First, uncheck all tags
        for i in range(root.childCount()):
            group_item = root.child(i)
            group_item.setCheckState(0, Qt.CheckState.Unchecked)
            for j in range(group_item.childCount()):
                tag_item = group_item.child(j)
                tag_item.setCheckState(0, Qt.CheckState.Unchecked)
        
        # Then check tags that are in the preset
        for i in range(root.childCount()):
            group_item = root.child(i)
            group_has_checked = False
            for j in range(group_item.childCount()):
                tag_item = group_item.child(j)
                tag_str = tag_item.data(0, Qt.ItemDataRole.UserRole)
                if tag_str in preset_tags:
                    tag_item.setCheckState(0, Qt.CheckState.Checked)
                    group_has_checked = True
            
            # Update group check state
            if group_has_checked:
                # Check if all tags in group are checked
                all_checked = True
                for j in range(group_item.childCount()):
                    if group_item.child(j).checkState(0) != Qt.CheckState.Checked:
                        all_checked = False
                        break
                if all_checked:
                    group_item.setCheckState(0, Qt.CheckState.Checked)
                else:
                    group_item.setCheckState(0, Qt.CheckState.PartiallyChecked)
        
        self.tags_tree.blockSignals(False)
        self._update_selected_tags()
        
        QMessageBox.information(self, "Preset Loaded",
                               f"Preset '{preset_name}' loaded successfully.")
    
    def _delete_preset(self) -> None:
        """Delete the selected preset."""
        if not self.config_manager:
            QMessageBox.warning(self, "No Config Manager",
                              "Preset deletion is not available.")
            return
        
        preset_name = self.preset_combo.currentText()
        if not preset_name or preset_name == "(No preset)":
            QMessageBox.warning(self, "No Preset Selected",
                              "Please select a preset to delete.")
            return
        
        # Confirm deletion
        reply = QMessageBox.question(
            self,
            "Delete Preset",
            f"Are you sure you want to delete preset '{preset_name}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.config_manager.delete_tag_export_preset(preset_name)
            self._load_presets_list()
            QMessageBox.information(self, "Preset Deleted",
                                   f"Preset '{preset_name}' deleted successfully.")

    def _export_presets(self) -> None:
        """Export all tag export presets to a JSON file."""
        if not self.config_manager:
            QMessageBox.warning(
                self,
                "No Config Manager",
                "Preset export is not available."
            )
            return

        presets = self.config_manager.get_tag_export_presets()
        if not presets:
            QMessageBox.information(
                self,
                "No Tag Presets",
                "There are no tag export presets to export."
            )
            return

        # Determine initial directory (reuse last export path behaviour)
        last_export_path = self.config_manager.get_last_export_path()
        if not last_export_path or not os.path.exists(last_export_path):
            last_export_path = os.getcwd()

        if os.path.isfile(last_export_path):
            last_export_path = os.path.dirname(last_export_path)

        default_filename = str(Path(last_export_path) / "tag_export_presets.json")

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Tag Presets",
            default_filename,
            "JSON Files (*.json);;All Files (*)"
        )

        if not file_path:
            return

        if not file_path.endswith(".json"):
            file_path += ".json"

        if self.config_manager.export_tag_export_presets(file_path):
            # Remember export directory
            self.config_manager.set_last_export_path(str(Path(file_path).parent))
            QMessageBox.information(
                self,
                "Export Successful",
                f"Tag export presets exported successfully to:\n{file_path}"
            )
        else:
            QMessageBox.warning(
                self,
                "Export Failed",
                f"Failed to export tag export presets to:\n{file_path}\n\n"
                "Please check file permissions and try again."
            )

    def _import_presets(self) -> None:
        """Import tag export presets from a JSON file."""
        if not self.config_manager:
            QMessageBox.warning(
                self,
                "No Config Manager",
                "Preset import is not available."
            )
            return

        # Use last path if available for initial directory
        last_path = self.config_manager.get_last_path()
        if not last_path or not os.path.exists(last_path):
            last_path = os.getcwd()

        if os.path.isfile(last_path):
            last_path = os.path.dirname(last_path)

        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Import Tag Presets",
            last_path,
            "JSON Files (*.json);;All Files (*)"
        )

        if not file_path:
            return

        result = self.config_manager.import_tag_export_presets(file_path)
        if result is None:
            QMessageBox.critical(
                self,
                "Import Failed",
                "Failed to import tag export presets.\n\n"
                "Please verify that the file is a valid DICOM Viewer V3 tag presets file."
            )
            return

        imported = result.get("imported", 0)
        skipped = result.get("skipped_conflicts", 0)

        # Refresh presets list in combo box
        self._load_presets_list()

        if imported == 0 and skipped == 0:
            QMessageBox.information(
                self,
                "No Presets Imported",
                "The selected file did not contain any tag export presets."
            )
        else:
            details_lines = [f"Presets imported: {imported}"]
            if skipped > 0:
                details_lines.append(
                    f"Presets skipped (already exist and were not overwritten): {skipped}"
                )
            details = "\n".join(details_lines)
            QMessageBox.information(
                self,
                "Import Complete",
                f"Tag export presets import completed.\n\n{details}"
            )

