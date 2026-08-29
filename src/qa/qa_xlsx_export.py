"""
XLSX export builder for QA results (no Qt).

Kept separate from ``qa_export.py`` (stdlib ``csv``/``json`` only) so the
heavier ``openpyxl`` + ``Pillow`` imports and the image-embed logic stay
isolated. Both modules remain Qt-free and independently testable.

Public:
    build_qa_workbook -- Summary/Detail/Images workbook for one or more QAResults
"""

from __future__ import annotations

import os
from typing import Any, cast

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from core.spreadsheet_safety import neutralize_spreadsheet_value
from qa.analysis_types import QAResult
from qa.qa_export import extract_low_contrast_cnr_values
from qa.qa_result_flatten import build_metric_rows


# Guarded at import time so the Images-sheet builder can cheaply detect a
# missing Pillow install without repeatedly trying/catching per image. Pillow
# is a declared/installed dependency (see requirements.txt); this guard only
# protects against an unusual environment where it is absent.
def _detect_pillow() -> bool:
    """Probe for Pillow once at import (single assignment keeps pyright happy)."""
    try:
        from PIL import Image  # noqa: F401  # pyright: ignore[reportUnusedImport]

        return True
    except Exception:
        return False


_PILLOW_AVAILABLE = _detect_pillow()

_SUMMARY_HEADERS = (
    "Series/Run ID",
    "Object ROI Mean",
    "Background Mean",
    "Background Std",
    "CNR",
    "Status",
    "Warnings",
    "PIU (%)",
    "PSG",
    "LC Score",
    "MTF@50% Row",
    "MTF@50% Col",
    "Slice Thickness (mm)",
    "Slice Shift (mm)",
)

# Modality-aware Summary columns pulled from the canonical flatten rows
# (``build_metric_rows``). Each entry is a human header paired with the exact
# flatten key; a column stays blank when its key is absent for a run (CT rows
# leave the MRI-only fields blank and vice versa — shared header, best-effort
# fill). Locked gaps (CT slice thickness, CT SNR, ``mri_snr``) are intentionally
# excluded; ``mri_snr`` lands with Phase 6.
_SUMMARY_KEY_COLUMNS: tuple[tuple[str, str], ...] = (
    ("PIU (%)", "uniformity_module.piu"),
    ("PSG", "uniformity_module.psg"),
    ("LC Score", "low_contrast_score"),
    ("MTF@50% Row", "slice1.row_mtf_50"),
    ("MTF@50% Col", "slice1.col_mtf_50"),
    ("Slice Thickness (mm)", "slice1.measured_slice_thickness_mm"),
    ("Slice Shift (mm)", "slice1.slice_shift_mm"),
)


def _row_label(result: QAResult, label: str | None) -> str:
    """Series/Run ID column value: explicit label, else the row's series_uid."""
    if label:
        return label
    return result.series_uid or ""


def _cnr_summary_values(result: QAResult) -> tuple[Any, Any, Any, Any]:
    """
    Pull the F1 CNR intermediates for the Summary sheet.

    Reads the canonical ``metrics["low_contrast_cnr"]`` shape (see F1 in
    PYLINAC_CT_CNR_BATCH_XLSX_PLAN.md): object ROI mean is the average of
    ``object_rois[*].mean``; background mean/std and module ``cnr`` are read
    with ``.get()``. Any missing key degrades to a blank cell.
    """
    obj_mean, bg_mean, bg_std, cnr = extract_low_contrast_cnr_values(result.metrics)
    return (
        "" if obj_mean is None else obj_mean,
        "" if bg_mean is None else bg_mean,
        "" if bg_std is None else bg_std,
        "" if cnr is None else cnr,
    )


def _xlsx_cell(value: Any) -> Any:
    """Join list/tuple cells like CSV, then neutralize formula-like strings.

    List/tuple values are joined with ``"; "`` (same separator as the CSV
    builders) so openpyxl never emits a Python list repr. Every string cell
    then passes through ``neutralize_spreadsheet_value`` so leading
    ``= + - @`` values are treated as literal text, not formulas.
    """
    if isinstance(value, (list, tuple)):
        value = "; ".join(str(item) for item in value)
    return neutralize_spreadsheet_value(value)


def _summary_extra_values(result: QAResult) -> list[Any]:
    """
    Pull the modality-aware Summary columns from the canonical flatten rows.

    Reads ``build_metric_rows(result)`` once and looks up each key in
    ``_SUMMARY_KEY_COLUMNS``; missing keys degrade to a blank cell. Numeric
    scalars (PIU, PSG, MTF, thickness, shift, LC score) stay numbers so the
    Summary sheet stays sortable/filterable — only the caller neutralizes the
    string cells (Series/Run, status, warnings).
    """
    flat = dict(build_metric_rows(result))
    return [flat.get(key) for _, key in _SUMMARY_KEY_COLUMNS]


def _build_summary_sheet(
    ws: Worksheet, results: list[QAResult], row_labels: list[str | None]
) -> None:
    ws.append(list(_SUMMARY_HEADERS))
    for result, label in zip(results, row_labels, strict=True):
        obj_mean, bg_mean, bg_std, cnr = _cnr_summary_values(result)
        status = "success" if result.success else "failed"
        warnings_text = "; ".join(result.warnings or [])
        extra = _summary_extra_values(result)
        ws.append(
            [
                _xlsx_cell(_row_label(result, label)),
                obj_mean,
                bg_mean,
                bg_std,
                cnr,
                _xlsx_cell(status),
                _xlsx_cell(warnings_text),
                *extra,
            ]
        )


def _build_detail_sheet(
    ws: Worksheet, results: list[QAResult], row_labels: list[str | None]
) -> None:
    """
    Full flatten metric rows, one block per run.

    Uses :func:`qa.qa_result_flatten.build_metric_rows` (the CSV source of
    truth) which walks ``result.raw_pylinac`` into dotted keys then overlays
    curated ``result.metrics`` (metrics wins on collision, curated keys stay
    top-level). The path denylist in that builder still applies —
    ``analyzed_image_path`` / ``pdf_report_path`` never reach Detail.

    Every string cell is neutralized via :func:`neutralize_spreadsheet_value`
    so formula-like values (e.g. DICOM-derived Series/Run labels or warnings)
    are written as literal text, not live formulas. List/tuple cells are
    joined with ``"; "`` first, matching the CSV export.
    """
    ws.append(["Metric", "Value"])
    for result, label in zip(results, row_labels, strict=True):
        ws.append([_xlsx_cell(_row_label(result, label)), ""])
        for key, value in build_metric_rows(result):
            ws.append([_xlsx_cell(key), _xlsx_cell(value)])
        ws.append([])  # blank separator row between run blocks


def _append_note(ws: Worksheet, note: str) -> None:
    ws.append([])
    ws.append([note])


def _module_images(result: QAResult) -> list[tuple[str, str]]:
    """
    Sorted (module label, absolute path) pairs for a run's per-module images.

    Reads ``analyzed_module_images`` (module label → abs PNG path), drops any
    path that is not a file on disk, and returns the pairs in stable ``str``
    key order. An empty dict (embed off / save failed) yields an empty list.
    A non-empty dict whose paths are all missing on disk also yields an empty
    list, so :func:`_embeddable_images` can fall back to the legacy composite.
    """
    module_images = getattr(result, "analyzed_module_images", None) or {}
    pairs = [
        (str(label), path)
        for label, path in module_images.items()
        if path and os.path.isfile(path)
    ]
    pairs.sort(key=lambda pair: pair[0])
    return pairs


def _composite_image_path(result: QAResult) -> str | None:
    """Absolute composite PNG path when the file exists on disk, else None."""
    path = getattr(result, "analyzed_image_path", None)
    return path if path and os.path.isfile(path) else None


def _embeddable_images(result: QAResult) -> list[tuple[str, str]]:
    """
    Image list for one run: per-module PNGs first, else legacy composite.

    Returns ``(module_label, abs_path)`` pairs. Composite fallback uses an
    empty module label so the Images sheet does not emit a second label row
    between the Series/Run ID and the image (same layout as the pre-P2-X3
    one-image-per-run sheet). An empty list means this run has nothing
    embeddable — the caller writes a per-run placeholder when the sheet
    exists, or skips the sheet when every run is empty.
    """
    module_imgs = _module_images(result)
    if module_imgs:
        return module_imgs
    composite = _composite_image_path(result)
    if composite is not None:
        return [("", composite)]
    return []


def _build_images_sheet(
    wb: Workbook,
    summary_ws: Worksheet,
    results: list[QAResult],
    row_labels: list[str | None],
) -> None:
    """
    Per-module embedded PNGs, stacked vertically, from ``analyzed_module_images``.

    Per run: a Series/Run ID label row, then for each module a label row
    followed by its embedded image. When a run has no module images but carries
    a legacy ``analyzed_image_path`` (composite), that single image is embedded
    as today (backward compat). When the sheet exists because *some* run has
    an embeddable image, runs with nothing embeddable still get a
    ``(no analyzed image for this run)`` placeholder so batch row identity
    is preserved. Degrades gracefully: if Pillow is unavailable, or none of
    the runs yield an embeddable image from either source, the Images sheet
    is skipped entirely and a note cell is appended to the Summary sheet
    instead -- the rest of the workbook (Summary, Detail) still writes.
    """
    if not _PILLOW_AVAILABLE:
        _append_note(summary_ws, "Images sheet skipped: Pillow is not available for image embedding.")
        return

    plans = [
        (result, label, _embeddable_images(result))
        for result, label in zip(results, row_labels, strict=True)
    ]
    if not any(image_list for _, _, image_list in plans):
        _append_note(summary_ws, "Images sheet skipped: no analyzed images were available.")
        return

    from openpyxl.drawing.image import Image as XLImage

    ws = wb.create_sheet("Images")
    row = 1
    for result, label, image_list in plans:
        ws.cell(row=row, column=1, value=_xlsx_cell(_row_label(result, label)))
        row += 1
        if not image_list:
            ws.cell(row=row, column=1, value="(no analyzed image for this run)")
            row += 2
            continue
        for module_label, path in image_list:
            if module_label:
                ws.cell(row=row, column=1, value=_xlsx_cell(module_label))
                row += 1
            try:
                image = XLImage(path)
                image.width = 480
                image.height = 480
                ws.add_image(image, f"A{row}")
                row += 34
            except Exception:
                ws.cell(row=row, column=1, value="(image could not be embedded)")
                row += 2


def build_qa_workbook(
    results: list[QAResult],
    labels: list[str] | None = None,
    *,
    app_version: str = "",  # pyright: ignore[reportUnusedParameter]
) -> Workbook:
    """
    Build an in-memory openpyxl Workbook for one or more QA runs.

    The caller saves the returned workbook (``workbook.save(path)``); this
    builder stays pure and Qt-free so it is testable via an in-memory
    round-trip. Single-run export passes a 1-element ``results`` list; batch
    export passes N.

    Args:
        results: QA results, one per run/series.
        labels: Optional display labels, parallel to ``results``. When None,
            each row falls back to its ``series_uid``.
        app_version: Reserved for future provenance metadata (unused today;
            accepted for signature parity with the JSON/CSV builders).

    Sheets:
        Summary -- one row per run: Series/Run ID, object ROI mean,
            background mean/std, CNR, status, warnings, then modality-aware
            key columns pulled from the canonical flatten (PIU, PSG, LC score,
            MTF@50% row/col, slice thickness/shift). Each extra column is
            best-effort: it stays blank when its flatten key is absent for the
            run, so CT and MRI rows share one header with blanks where a metric
            does not apply (CT slice thickness, CT SNR, and ``mri_snr`` are
            excluded by design).
        Detail -- full flatten per run (``build_metric_rows``; path denylist).
        Images -- per-module embedded PNGs from ``analyzed_module_images``
            (stable key sort), each preceded by its module label, stacked
            vertically per run; falls back to the legacy composite
            ``analyzed_image_path`` when a run has no module images. When the
            sheet exists, a run with nothing embeddable still gets a
            placeholder row. Skipped (with a Summary note) when Pillow is
            unavailable or no run yields an embeddable image. Series/Run and
            module labels on every sheet are formula-injection neutralized.
    """
    if labels is not None and len(labels) != len(results):
        raise ValueError("labels must be parallel to results (same length)")
    row_labels: list[str | None] = list(labels) if labels is not None else [None] * len(results)

    wb = Workbook()
    # A fresh Workbook always has an active worksheet; cast narrows the
    # Optional/chartsheet union that openpyxl's stubs declare for .active.
    summary_ws = cast(Worksheet, wb.active)
    summary_ws.title = "Summary"
    _build_summary_sheet(summary_ws, results, row_labels)

    detail_ws = wb.create_sheet("Detail")
    _build_detail_sheet(detail_ws, results, row_labels)

    _build_images_sheet(wb, summary_ws, results, row_labels)

    return wb
