"""
XLSX worksheet writers for ROI / crosshair / measurement export.

Extracted from ``roi_export_service.write_xlsx`` to clear Sonar ``python:S3776``
(cognitive complexity) while preserving series → slice → annotation layout,
area unit thresholds, and multichannel rows.

Inputs:
    - openpyxl worksheet and bold Font
    - Collected series/slice annotation tuples
    - DICOM datasets and optional ROI manager for statistics

Outputs:
    - Rows appended to the worksheet; caller saves the workbook

Requirements:
    - openpyxl (imported by caller)
    - ``core.roi_export_service`` helpers for stats / crosshair / measurements
"""

from __future__ import annotations

from typing import Any

from pydicom.dataset import Dataset

from core.dicom_color import multichannel_axis_labels
from core.spreadsheet_safety import (
    neutralize_spreadsheet_value as _safe_spreadsheet_value,
)

# Late-bound helpers from roi_export_service (avoid circular import at load of
# that module's body; these are defined before write_xlsx is re-exported).


def _svc():
    """Return roi_export_service without a static import cycle for basedpyright."""
    import importlib

    return importlib.import_module("core.roi_export_service")


def resolve_export_roi_manager(
    subwindow_managers: dict[int, dict[str, Any]],
) -> Any | None:
    """Pick the first available ROI manager from subwindow manager dicts."""
    roi_manager = None
    for idx in sorted(subwindow_managers.keys()):
        if subwindow_managers[idx].get("roi_manager"):
            roi_manager = subwindow_managers[idx]["roi_manager"]
            break
    if roi_manager is None and subwindow_managers:
        roi_manager = next(iter(subwindow_managers.values())).get("roi_manager")
    return roi_manager


def xlsx_area_label_value_unit(
    area_mm2_f: float | None,
    area_px_f: float,
) -> tuple[str, float, str]:
    """
    Choose Area row value and unit (cm² / mm² / pixels).

    Threshold matches prior write_xlsx: ``area_mm2_f >= 100.0`` → cm².
    """
    if area_mm2_f is not None:
        if area_mm2_f >= 100.0:
            return ("Area", round(area_mm2_f / 100.0, 2), "cm²")
        return ("Area", round(area_mm2_f, 2), "mm²")
    return ("Area", area_px_f, "pixels")


def write_xlsx_labeled_value_row(
    ws: Any,
    row: int,
    label: str,
    value: Any,
    unit: str = "",
) -> int:
    """Write columns 1–3 (label, value, optional unit) and return next row."""
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value=value)
    ws.cell(row=row, column=3, value=_safe_spreadsheet_value(unit) if unit else "")
    return row + 1


def write_xlsx_roi_statistics_rows(
    ws: Any,
    row: int,
    *,
    roi_item: Any,
    dataset: Dataset,
    use_rescale: bool,
    roi_manager: Any,
    dicom_processor: type,
) -> int:
    """Append mean/std/min/max/pixels/area and optional per-channel rows."""
    svc = _svc()
    stats, rescale_unit = svc.compute_roi_statistics(
        roi_item, dataset, use_rescale, roi_manager, dicom_processor
    )
    unit_str = rescale_unit or ""
    mean_v = stats.get("mean")
    std_v = stats.get("std")
    min_v = stats.get("min")
    max_v = stats.get("max")
    count_v = stats.get("count")
    area_px_v = stats.get("area_pixels")
    area_mm2_v = stats.get("area_mm2")

    mean_f = float(mean_v) if mean_v is not None else 0.0
    std_f = float(std_v) if std_v is not None else 0.0
    min_f = float(min_v) if min_v is not None else 0.0
    max_f = float(max_v) if max_v is not None else 0.0
    count_i = int(count_v) if count_v is not None else 0
    area_px_f = float(area_px_v) if area_px_v is not None else 0.0
    area_mm2_f = float(area_mm2_v) if area_mm2_v is not None else None

    for label, value in (
        ("Mean", round(mean_f, 2)),
        ("Std Dev", round(std_f, 2)),
        ("Min", round(min_f, 2)),
        ("Max", round(max_f, 2)),
    ):
        row = write_xlsx_labeled_value_row(ws, row, label, value, unit_str)

    row = write_xlsx_labeled_value_row(ws, row, "Pixels", count_i, "")

    area_label, area_value, area_unit = xlsx_area_label_value_unit(area_mm2_f, area_px_f)
    row = write_xlsx_labeled_value_row(ws, row, area_label, area_value, area_unit)

    channel_count, channel_values = svc._extract_channel_stats(stats)
    ch_labels = multichannel_axis_labels(dataset, channel_count)
    for c in range(channel_count):
        lab = ch_labels[c]
        for metric, key in (
            ("Mean", f"mean_ch{c}"),
            ("Std Dev", f"std_ch{c}"),
            ("Min", f"min_ch{c}"),
            ("Max", f"max_ch{c}"),
        ):
            row = write_xlsx_labeled_value_row(
                ws,
                row,
                f"{lab} {metric}",
                _safe_spreadsheet_value(channel_values.get(key, "")),
                unit_str,
            )
    return row


def write_xlsx_crosshair_rows(ws: Any, row: int, cross_item: Any, dataset: Dataset) -> int:
    """Append one crosshair heading's data rows; returns next row index."""
    svc = _svc()
    data = svc.get_crosshair_export_data(cross_item, dataset)
    crosshair_rows = [
        ("Pixel X", data["pixel_x"]),
        ("Pixel Y", data["pixel_y"]),
        ("Slice Index (0-based)", data["slice_index"]),
        ("Pixel Value", _safe_spreadsheet_value(data["pixel_value_str"])),
        ("Patient X (mm)", svc._format_float(data["patient_x"])),
        ("Patient Y (mm)", svc._format_float(data["patient_y"])),
        ("Patient Z (mm)", svc._format_float(data["patient_z"])),
    ]
    for label, value in crosshair_rows:
        row = write_xlsx_labeled_value_row(ws, row, label, value, "")
    return row


def write_xlsx_measurement_rows(ws: Any, row: int, m_item: Any) -> int:
    """Append key/value rows for one distance or angle measurement."""
    svc = _svc()
    for label, value in svc._measurement_xlsx_label_value_pairs(m_item):
        row = write_xlsx_labeled_value_row(
            ws, row, label, _safe_spreadsheet_value(value), ""
        )
    return row


def write_xlsx_slice_annotations(
    ws: Any,
    row: int,
    bold_font: Any,
    *,
    z: int,
    rois: list[Any],
    crosshairs: list[Any],
    measurements: list[Any],
    series_dict: list[Dataset],
    roi_manager: Any,
    use_rescale: bool,
    dicom_processor: type,
) -> int:
    """Write one slice header plus all ROI / crosshair / measurement blocks."""
    dataset = series_dict[z] if z < len(series_dict) else None
    ws.cell(row=row, column=1, value=f"  Slice Index (0-based): {z}")
    row += 1

    for roi_idx, roi_item in enumerate(rois, start=1):
        shape = getattr(roi_item, "shape_type", "ellipse").capitalize()
        ws.cell(row=row, column=1, value=f"  {shape} ROI {roi_idx}")
        ws.cell(row=row, column=1).font = bold_font
        row += 1
        if dataset and roi_manager:
            row = write_xlsx_roi_statistics_rows(
                ws,
                row,
                roi_item=roi_item,
                dataset=dataset,
                use_rescale=use_rescale,
                roi_manager=roi_manager,
                dicom_processor=dicom_processor,
            )

    for cross_idx, cross_item in enumerate(crosshairs, start=1):
        ws.cell(row=row, column=1, value=f"  Crosshair {cross_idx}")
        ws.cell(row=row, column=1).font = bold_font
        row += 1
        if dataset:
            row = write_xlsx_crosshair_rows(ws, row, cross_item, dataset)

    for meas_idx, m_item in enumerate(measurements, start=1):
        ws.cell(row=row, column=1, value=f"  Measurement {meas_idx}")
        ws.cell(row=row, column=1).font = bold_font
        row += 1
        row = write_xlsx_measurement_rows(ws, row, m_item)

    return row


def write_xlsx_series_block(
    ws: Any,
    row: int,
    bold_font: Any,
    *,
    study_uid: str,
    series_uid: str,
    slice_list: list[Any],
    current_studies: dict[str, dict[str, list[Dataset]]],
    roi_manager: Any,
    use_rescale: bool,
    dicom_processor: type,
) -> int | None:
    """
    Write one series header and its slices.

    Returns the next row index, or ``None`` when the series is not loaded
    (caller should skip without advancing).
    """
    svc = _svc()
    series_dict = current_studies.get(study_uid, {}).get(series_uid, [])
    if not series_dict:
        return None

    first_ds = series_dict[0]
    series_num = getattr(first_ds, "SeriesNumber", "")
    series_desc = getattr(
        first_ds, svc._TAG_SERIES_DESCRIPTION, svc._DEFAULT_UNKNOWN_SERIES
    )
    study_desc = getattr(first_ds, "StudyDescription", "Study")[:31]
    study_desc = svc._sanitize_filename(study_desc) or "Study"
    _ = study_desc  # historical; workbook sheet title remains "ROI Statistics"

    ws.cell(
        row=row,
        column=1,
        value=_safe_spreadsheet_value(f"Series {series_num}: {series_desc}"),
    )
    ws.cell(row=row, column=1).font = bold_font
    row += 1

    if not slice_list:
        ws.cell(row=row, column=1, value="No annotations")
        return row + 2

    for z, rois, crosshairs, measurements in slice_list:
        row = write_xlsx_slice_annotations(
            ws,
            row,
            bold_font,
            z=z,
            rois=rois,
            crosshairs=crosshairs,
            measurements=measurements,
            series_dict=series_dict,
            roi_manager=roi_manager,
            use_rescale=use_rescale,
            dicom_processor=dicom_processor,
        )
    return row + 1


def write_xlsx_workbook(
    file_path: str,
    collected: list[Any],
    current_studies: dict[str, dict[str, list[Dataset]]],
    subwindow_managers: dict[int, dict[str, Any]],
    use_rescale: bool,
    dicom_processor: type,
) -> None:
    """
    Build and save the ROI Statistics workbook.

    Raises ImportError when openpyxl is missing (same message as before).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as e:
        raise ImportError(
            "openpyxl is required for XLSX export. Install with: pip install openpyxl"
        ) from e

    roi_manager = resolve_export_roi_manager(subwindow_managers)

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook has no active sheet")
    ws.title = "ROI Statistics"[:31]

    bold_font = Font(bold=True)
    row = 1

    for (study_uid, series_uid), slice_list in collected:
        next_row = write_xlsx_series_block(
            ws,
            row,
            bold_font,
            study_uid=study_uid,
            series_uid=series_uid,
            slice_list=slice_list,
            current_studies=current_studies,
            roi_manager=roi_manager,
            use_rescale=use_rescale,
            dicom_processor=dicom_processor,
        )
        if next_row is None:
            continue
        row = next_row

    wb.save(file_path)
