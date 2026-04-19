"""
ROI Export Service

Aggregates ROI, crosshair, and distance/angle measurement data from all subwindows,
recomputes ROI statistics per slice, and writes TXT, CSV, or XLSX export files.

Inputs:
    - current_studies: {study_uid: {series_uid: [Dataset]}}
    - subwindow_managers: {idx: {'roi_manager', 'crosshair_manager', 'measurement_tool', ...}}
    - selected_series: list of (study_uid, series_uid)
    - use_rescale: whether to use rescale slope/intercept (e.g. HU)
    - file path and format (TXT, CSV, XLSX)

Outputs:
    - Written files (TXT, CSV, or XLSX) with series → slice → ROI/crosshair/measurement hierarchy

Requirements:
    - core.dicom_processor.DICOMProcessor (get_pixel_array, get_rescale_parameters)
    - utils.dicom_utils (get_pixel_spacing, pixel_to_patient_coordinates)
    - tools.roi_manager.ROIManager.calculate_statistics (stateless per call)
    - tools.measurement_items.MeasurementItem, tools.angle_measurement_items.AngleMeasurementItem
    - openpyxl for XLSX, csv for CSV
"""

from __future__ import annotations

import csv
import math
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from pydicom.dataset import Dataset

from core.dicom_processor import DICOMProcessor
from tools.angle_measurement_items import AngleMeasurementItem
from tools.measurement_items import MeasurementItem
from utils.dicom_utils import get_pixel_spacing, pixel_to_patient_coordinates


# Type aliases for keys and data structures
SeriesKey = Tuple[str, str]  # (study_uid, series_uid)
SliceData = Tuple[int, List[Any], List[Any], List[Any]]  # (z, rois, crosshairs, measurements)
CollectedSeries = List[Tuple[SeriesKey, List[SliceData]]]

# Trailing CSV columns for distance/angle rows (ROI and crosshair rows leave these blank).
MEASUREMENT_CSV_HEADERS: List[str] = [
    "Measurement Type",
    "Measurement Index",
    "Distance (mm)",
    "Distance (pixels)",
    "Angle (degrees)",
    "Measurement display",
    "P1 scene X",
    "P1 scene Y",
    "P2 scene X",
    "P2 scene Y",
    "P3 scene X",
    "P3 scene Y",
]


def _sanitize_filename(s: str) -> str:
    """Replace characters invalid in filenames with underscore."""
    return re.sub(r'[/\\:*?"<>|]', '_', s).strip()


def collect_roi_data(
    selected_series: List[SeriesKey],
    current_studies: Dict[str, Dict[str, List[Dataset]]],
    subwindow_managers: Dict[int, Dict[str, Any]],
) -> CollectedSeries:
    """
    Collect ROI, crosshair, and measurement items for selected series from all subwindow managers.

    For each (study_uid, series_uid), for each slice index z in the series,
    aggregates all ROIItem, CrosshairItem, and measurement graphics items from every
    subwindow's roi_manager, crosshair_manager, and measurement_tool using key
    (study_uid, series_uid, z). No deduplication: each subwindow holds separate instances.

    Args:
        selected_series: List of (study_uid, series_uid) to export.
        current_studies: App studies dict {study_uid: {series_uid: [datasets]}}.
        subwindow_managers: Dict of {idx: {'roi_manager', 'crosshair_manager', 'measurement_tool', ...}}.

    Returns:
        List of ((study_uid, series_uid), [(z, rois, crosshairs, measurements), ...]) where each
        slice tuple is kept only if at least one of rois, crosshairs, or measurements is non-empty.
    """
    result: CollectedSeries = []

    for study_uid, series_uid in selected_series:
        series_dict = current_studies.get(study_uid, {}).get(series_uid)
        if not series_dict:
            continue
        num_slices = len(series_dict)
        slice_data_list: List[SliceData] = []

        for z in range(num_slices):
            key = (study_uid, series_uid, z)
            rois: List[Any] = []
            crosshairs: List[Any] = []
            measurements: List[Any] = []

            for idx in sorted(subwindow_managers.keys()):
                managers = subwindow_managers[idx]
                roi_mgr = managers.get("roi_manager")
                crosshair_mgr = managers.get("crosshair_manager")
                meas_tool = managers.get("measurement_tool")
                if roi_mgr and hasattr(roi_mgr, "rois"):
                    rois.extend(roi_mgr.rois.get(key, []))
                if crosshair_mgr and hasattr(crosshair_mgr, "crosshairs"):
                    crosshairs.extend(crosshair_mgr.crosshairs.get(key, []))
                if meas_tool is not None and hasattr(meas_tool, "get_measurements_for_slice"):
                    measurements.extend(
                        meas_tool.get_measurements_for_slice(study_uid, series_uid, z)
                    )

            if rois or crosshairs or measurements:
                slice_data_list.append((z, rois, crosshairs, measurements))

        # Include every selected series (even with no annotations) per plan E14
        result.append(((study_uid, series_uid), slice_data_list))

    return result


def compute_roi_statistics(
    roi_item: Any,
    dataset: Dataset,
    use_rescale: bool,
    roi_manager: Any,
    dicom_processor: type,
) -> Tuple[Dict[str, float | int | None], Optional[str]]:
    """
    Recompute ROI statistics for an ROI item (never use roi_item.statistics cache).

    calculate_statistics does not use manager-level state; any ROIManager instance is fine.

    Args:
        roi_item: ROIItem with shape_type, get_bounds, get_mask.
        dataset: DICOM dataset for the slice.
        use_rescale: If True, use rescale slope/intercept for values (e.g. HU).
        roi_manager: Any ROIManager instance to call calculate_statistics.
        dicom_processor: DICOMProcessor class for get_pixel_array, get_rescale_parameters.

    Returns:
        (stats_dict, rescale_unit). rescale_unit is e.g. "HU" or None.
    """
    pixel_array = dicom_processor.get_pixel_array(dataset)
    if pixel_array is None:
        return (
            {
                "mean": 0.0,
                "std": 0.0,
                "min": 0.0,
                "max": 0.0,
                "count": 0,
                "area_pixels": 0.0,
                "area_mm2": None,
            },
            None,
        )

    pixel_spacing = get_pixel_spacing(dataset)
    rescale_slope: Optional[float] = None
    rescale_intercept: Optional[float] = None
    rescale_type: Optional[str] = None

    if use_rescale:
        rescale_slope, rescale_intercept, rescale_type = dicom_processor.get_rescale_parameters(dataset)

    stats = roi_manager.calculate_statistics(
        roi_item,
        pixel_array,
        rescale_slope=rescale_slope,
        rescale_intercept=rescale_intercept,
        pixel_spacing=pixel_spacing,
    )
    return (stats, rescale_type)


def get_crosshair_export_data(
    crosshair_item: Any,
    dataset: Dataset,
) -> Dict[str, Any]:
    """
    Get crosshair data for export: pixel coords, pixel value string, patient coords.

    Uses item.x_coord, item.y_coord, item.z_coord (raw). Does not parse
    item.pixel_value_str for coordinates; patient coords are computed fresh via
    pixel_to_patient_coordinates. If patient coords are unavailable, use None → "N/A" in output.

    Args:
        crosshair_item: CrosshairItem with x_coord, y_coord, z_coord, pixel_value_str.
        dataset: DICOM dataset for the slice.

    Returns:
        Dict with pixel_x, pixel_y, slice_index, pixel_value_str, patient_x, patient_y, patient_z
        (patient_* are float or None).
    """
    patient = pixel_to_patient_coordinates(
        dataset,
        crosshair_item.x_coord,
        crosshair_item.y_coord,
        crosshair_item.z_coord,
    )
    return {
        "pixel_x": crosshair_item.x_coord,
        "pixel_y": crosshair_item.y_coord,
        "slice_index": crosshair_item.z_coord,
        "pixel_value_str": getattr(crosshair_item, "pixel_value_str", ""),
        "patient_x": patient[0] if patient else None,
        "patient_y": patient[1] if patient else None,
        "patient_z": patient[2] if patient else None,
    }


def _empty_measurement_csv_row() -> List[str]:
    """Blank measurement columns for ROI and crosshair CSV rows."""
    return [""] * len(MEASUREMENT_CSV_HEADERS)


def serialize_measurement_for_export(item: Any, index: int) -> List[str]:
    """
    Build trailing measurement columns for CSV (same order as MEASUREMENT_CSV_HEADERS).

    Scene coordinates are QGraphics scene units (image-plane) at export time.
    """
    empty = _empty_measurement_csv_row()
    idx_str = str(index)

    if isinstance(item, AngleMeasurementItem):
        p1, p2, p3 = item.p1, item.p2, item.p3
        display = getattr(item, "angle_formatted", "") or ""
        deg = float(getattr(item, "angle_degrees", 0.0))
        return [
            "angle",
            idx_str,
            "",
            "",
            f"{deg:.4f}",
            display,
            f"{p1.x():.4f}",
            f"{p1.y():.4f}",
            f"{p2.x():.4f}",
            f"{p2.y():.4f}",
            f"{p3.x():.4f}",
            f"{p3.y():.4f}",
        ]

    if isinstance(item, MeasurementItem):
        dx = item.end_point.x() - item.start_point.x()
        dy = item.end_point.y() - item.start_point.y()
        dist_px = math.sqrt(dx * dx + dy * dy)
        spacing = getattr(item, "pixel_spacing", None)
        dist_mm_str = ""
        if spacing is not None and len(spacing) >= 2:
            dx_scaled = dx * float(spacing[1])
            dy_scaled = dy * float(spacing[0])
            dist_mm = math.sqrt(dx_scaled * dx_scaled + dy_scaled * dy_scaled)
            dist_mm_str = f"{dist_mm:.6f}"
        p1, p2 = item.start_point, item.end_point
        display = getattr(item, "distance_formatted", "") or ""
        return [
            "distance",
            idx_str,
            dist_mm_str,
            f"{dist_px:.6f}",
            "",
            display,
            f"{p1.x():.4f}",
            f"{p1.y():.4f}",
            f"{p2.x():.4f}",
            f"{p2.y():.4f}",
            "",
            "",
        ]

    return empty


def measurement_txt_block_lines(item: Any, index: int) -> List[str]:
    """Human-readable lines for one measurement in TXT export."""
    lines: List[str] = []
    lines.append(f"  Measurement {index}")
    if isinstance(item, AngleMeasurementItem):
        p1, p2, p3 = item.p1, item.p2, item.p3
        lines.append("    Type            Angle")
        lines.append(f"    Display         {getattr(item, 'angle_formatted', '')}")
        lines.append(f"    Angle (deg)     {float(getattr(item, 'angle_degrees', 0.0)):.4f}")
        lines.append(
            f"    P1 scene        ({p1.x():.2f}, {p1.y():.2f})"
        )
        lines.append(
            f"    P2 (vertex)     ({p2.x():.2f}, {p2.y():.2f})"
        )
        lines.append(
            f"    P3 scene        ({p3.x():.2f}, {p3.y():.2f})"
        )
    elif isinstance(item, MeasurementItem):
        dx = item.end_point.x() - item.start_point.x()
        dy = item.end_point.y() - item.start_point.y()
        dist_px = math.sqrt(dx * dx + dy * dy)
        spacing = getattr(item, "pixel_spacing", None)
        lines.append("    Type            Distance")
        lines.append(f"    Display         {getattr(item, 'distance_formatted', '')}")
        if spacing is not None and len(spacing) >= 2:
            dx_scaled = dx * float(spacing[1])
            dy_scaled = dy * float(spacing[0])
            dist_mm = math.sqrt(dx_scaled * dx_scaled + dy_scaled * dy_scaled)
            lines.append(f"    Distance (mm)   {dist_mm:.4f}")
        lines.append(f"    Distance (px)   {dist_px:.4f}")
        lines.append(
            f"    Start scene     ({item.start_point.x():.2f}, {item.start_point.y():.2f})"
        )
        lines.append(
            f"    End scene       ({item.end_point.x():.2f}, {item.end_point.y():.2f})"
        )
    else:
        lines.append("    Type            (unknown)")
    lines.append("")
    return lines


def _measurement_xlsx_label_value_pairs(item: Any) -> List[Tuple[str, Any]]:
    """Key/value rows for one measurement in XLSX export (columns 1–2)."""
    pairs: List[Tuple[str, Any]] = []
    if isinstance(item, AngleMeasurementItem):
        p1, p2, p3 = item.p1, item.p2, item.p3
        pairs.append(("Type", "Angle"))
        pairs.append(("Display", getattr(item, "angle_formatted", "")))
        pairs.append(("Angle (deg)", round(float(getattr(item, "angle_degrees", 0.0)), 4)))
        pairs.append(("P1 scene X", round(p1.x(), 4)))
        pairs.append(("P1 scene Y", round(p1.y(), 4)))
        pairs.append(("P2 vertex X", round(p2.x(), 4)))
        pairs.append(("P2 vertex Y", round(p2.y(), 4)))
        pairs.append(("P3 scene X", round(p3.x(), 4)))
        pairs.append(("P3 scene Y", round(p3.y(), 4)))
    elif isinstance(item, MeasurementItem):
        dx = item.end_point.x() - item.start_point.x()
        dy = item.end_point.y() - item.start_point.y()
        dist_px = math.sqrt(dx * dx + dy * dy)
        spacing = getattr(item, "pixel_spacing", None)
        pairs.append(("Type", "Distance"))
        pairs.append(("Display", getattr(item, "distance_formatted", "")))
        if spacing is not None and len(spacing) >= 2:
            dx_scaled = dx * float(spacing[1])
            dy_scaled = dy * float(spacing[0])
            dist_mm = math.sqrt(dx_scaled * dx_scaled + dy_scaled * dy_scaled)
            pairs.append(("Distance (mm)", round(dist_mm, 4)))
        pairs.append(("Distance (px)", round(dist_px, 4)))
        pairs.append(("Start scene X", round(item.start_point.x(), 4)))
        pairs.append(("Start scene Y", round(item.start_point.y(), 4)))
        pairs.append(("End scene X", round(item.end_point.x(), 4)))
        pairs.append(("End scene Y", round(item.end_point.y(), 4)))
    else:
        pairs.append(("Type", "unknown"))
    return pairs


def _format_float(v: Optional[float]) -> str:
    if v is None:
        return "N/A"
    return f"{v:.4f}"


def _extract_channel_stats(
    stats: Dict[str, float | int | None],
) -> Tuple[int, Dict[str, str]]:
    """
    Extract per-channel statistics from ROI stats payload.

    Returns:
        (channel_count, values) where values maps keys like ``mean_ch0`` to
        formatted strings for export.
    """
    channel_count = int(stats.get("multichannel_count") or 0)
    values: Dict[str, str] = {}
    for c in range(channel_count):
        for metric in ("mean", "std", "min", "max"):
            k = f"{metric}_ch{c}"
            v = stats.get(k)
            values[k] = f"{float(v):.4f}" if v is not None else ""
    return channel_count, values


def _safe_spreadsheet_value(v: Any) -> Any:
    """Neutralize formula-like cell values for CSV/XLSX exports."""
    if isinstance(v, str) and v[:1] in ("=", "+", "-", "@"):
        return "'" + v
    return v


def write_txt(
    file_path: str,
    collected: CollectedSeries,
    current_studies: Dict[str, Dict[str, List[Dataset]]],
    subwindow_managers: Dict[int, Dict[str, Any]],
    use_rescale: bool,
    dicom_processor: type,
) -> None:
    """
    Write export in TXT format: series header, slice header, ROI/crosshair sections with key-value lines.
    Skips slices with no annotations. Labels z as "Slice Index (0-based)".
    """
    roi_manager = None
    for idx in sorted(subwindow_managers.keys()):
        if subwindow_managers[idx].get("roi_manager"):
            roi_manager = subwindow_managers[idx]["roi_manager"]
            break
    if roi_manager is None:
        roi_manager = list(subwindow_managers.values())[0].get("roi_manager") if subwindow_managers else None

    lines: List[str] = []
    sep = "=" * 60
    subsep = "-" * 40

    for (study_uid, series_uid), slice_list in collected:
        series_dict = current_studies.get(study_uid, {}).get(series_uid, [])
        if not series_dict:
            continue
        first_ds = series_dict[0]
        series_num = getattr(first_ds, "SeriesNumber", "")
        series_desc = getattr(first_ds, "SeriesDescription", "Unknown Series")
        lines.append(sep)
        lines.append(f"Series {series_num}: {series_desc}")
        lines.append(sep)
        if not slice_list:
            lines.append("  No annotations")
            lines.append("")
            continue
        for z, rois, crosshairs, measurements in slice_list:
            dataset = series_dict[z] if z < len(series_dict) else None
            lines.append(f"  Slice Index (0-based): {z}")
            lines.append(subsep)
            roi_idx = 0
            for roi_item in rois:
                roi_idx += 1
                shape = getattr(roi_item, "shape_type", "ellipse").capitalize()
                lines.append(f"  {shape} ROI {roi_idx}")
                if dataset and roi_manager:
                    stats, rescale_unit = compute_roi_statistics(
                        roi_item, dataset, use_rescale, roi_manager, dicom_processor
                    )
                    unit_str = rescale_unit or ""
                    mean_v = stats.get("mean")
                    std_v = stats.get("std")
                    min_v = stats.get("min")
                    max_v = stats.get("max")
                    count_v = stats.get("count")
                    area_px_v = stats.get("area_pixels")

                    mean_f = float(mean_v) if mean_v is not None else 0.0
                    std_f = float(std_v) if std_v is not None else 0.0
                    min_f = float(min_v) if min_v is not None else 0.0
                    max_f = float(max_v) if max_v is not None else 0.0
                    count_i = int(count_v) if count_v is not None else 0
                    area_px_f = float(area_px_v) if area_px_v is not None else 0.0

                    lines.append(f"    Mean       {mean_f:.2f}    {unit_str}")
                    lines.append(f"    Std Dev    {std_f:.2f}    {unit_str}")
                    lines.append(f"    Min        {min_f:.2f}    {unit_str}")
                    lines.append(f"    Max        {max_f:.2f}    {unit_str}")
                    lines.append(f"    Pixels     {count_i}    ")

                    area_mm2 = stats.get("area_mm2")
                    if area_mm2 is not None:
                        area_mm2_f = float(area_mm2)
                        if area_mm2_f >= 100.0:
                            area_cm2 = area_mm2_f / 100.0
                            lines.append(f"    Area       {area_cm2:.2f}    cm²")
                        else:
                            lines.append(f"    Area       {area_mm2_f:.2f}    mm²")
                    else:
                        lines.append(f"    Area       {area_px_f:.1f}    pixels")
                    channel_count, channel_values = _extract_channel_stats(stats)
                    for c in range(channel_count):
                        ch = c + 1
                        lines.append(
                            f"    Ch{ch} Mean   {channel_values.get(f'mean_ch{c}', '')}    {unit_str}"
                        )
                        lines.append(
                            f"    Ch{ch} Std    {channel_values.get(f'std_ch{c}', '')}    {unit_str}"
                        )
                        lines.append(
                            f"    Ch{ch} Min    {channel_values.get(f'min_ch{c}', '')}    {unit_str}"
                        )
                        lines.append(
                            f"    Ch{ch} Max    {channel_values.get(f'max_ch{c}', '')}    {unit_str}"
                        )
                lines.append("")
            cross_idx = 0
            for cross_item in crosshairs:
                cross_idx += 1
                lines.append(f"  Crosshair {cross_idx}")
                if dataset:
                    data = get_crosshair_export_data(cross_item, dataset)
                    lines.append(f"    Pixel X        {data['pixel_x']}    ")
                    lines.append(f"    Pixel Y        {data['pixel_y']}    ")
                    lines.append(f"    Slice Index    {data['slice_index']}    ")
                    lines.append(f"    Pixel Value    {data['pixel_value_str']}    ")
                    lines.append(f"    Patient X (mm) {_format_float(data['patient_x'])}    ")
                    lines.append(f"    Patient Y (mm) {_format_float(data['patient_y'])}    ")
                    lines.append(f"    Patient Z (mm) {_format_float(data['patient_z'])}    ")
                lines.append("")
            for meas_idx, m_item in enumerate(measurements, start=1):
                lines.extend(measurement_txt_block_lines(m_item, meas_idx))
        lines.append("")

    Path(file_path).write_text("\n".join(lines), encoding="utf-8")


def write_csv(
    file_path: str,
    collected: CollectedSeries,
    current_studies: Dict[str, Dict[str, List[Dataset]]],
    subwindow_managers: Dict[int, Dict[str, Any]],
    use_rescale: bool,
    dicom_processor: type,
) -> None:
    """
    Write one data row per ROI, crosshair, or measurement. Base columns per plan E12.
    ROI rows fill stat columns; crosshair and measurement rows leave ROI stats blank.
    Trailing columns hold per-channel ROI stats (when present) then distance/angle fields.
    """
    roi_manager = None
    for idx in sorted(subwindow_managers.keys()):
        if subwindow_managers[idx].get("roi_manager"):
            roi_manager = subwindow_managers[idx]["roi_manager"]
            break
    if roi_manager is None and subwindow_managers:
        roi_manager = list(subwindow_managers.values())[0].get("roi_manager")

    headers = [
        "Study UID",
        "Series Number",
        "Series Description",
        "Slice Index (0-based)",
        "ROI Type",
        "ROI Index",
        "Mean",
        "Std Dev",
        "Min",
        "Max",
        "Pixels",
        "Area (pixels)",
        "Area (mm²)",
        "Rescale Unit",
        "Pixel X",
        "Pixel Y",
        "Pixel Z",
        "Pixel Value",
        "Patient X (mm)",
        "Patient Y (mm)",
        "Patient Z (mm)",
    ]
    row_payloads: List[Tuple[List[str], Dict[str, str], List[str]]] = [
        (headers, {}, _empty_measurement_csv_row())
    ]
    max_channel_count = 0

    for (study_uid, series_uid), slice_list in collected:
        series_dict = current_studies.get(study_uid, {}).get(series_uid, [])
        if not series_dict:
            continue
        first_ds = series_dict[0]
        study_uid_short = (study_uid[:36] + "..") if len(study_uid) > 38 else study_uid
        series_num = getattr(first_ds, "SeriesNumber", "")
        series_desc = getattr(first_ds, "SeriesDescription", "Unknown Series")

        for z, rois, crosshairs, measurements in slice_list:
            dataset = series_dict[z] if z < len(series_dict) else None
            for roi_idx, roi_item in enumerate(rois, start=1):
                shape = getattr(roi_item, "shape_type", "ellipse")
                row = [
                    study_uid_short,
                    str(series_num),
                    series_desc,
                    str(z),
                    shape,
                    str(roi_idx),
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
                channel_values: Dict[str, str] = {}
                if dataset and roi_manager:
                    stats, rescale_unit = compute_roi_statistics(
                        roi_item, dataset, use_rescale, roi_manager, dicom_processor
                    )
                    mean_v = stats.get("mean")
                    std_v = stats.get("std")
                    min_v = stats.get("min")
                    max_v = stats.get("max")
                    count_v = stats.get("count")
                    area_px_v = stats.get("area_pixels")
                    area_mm2_v = stats.get("area_mm2")

                    mean_f = float(mean_v) if mean_v is not None else 0.0
                    std_f = float(std_v) if std_v is not None else 0.0
                    min_f = float(min_v) if min_v is not None else 0.0
                    max_f = float(max_v) if max_v is not None else 0.0
                    count_i = int(count_v) if count_v is not None else 0
                    area_px_f = float(area_px_v) if area_px_v is not None else 0.0

                    row[6] = f"{mean_f:.4f}"
                    row[7] = f"{std_f:.4f}"
                    row[8] = f"{min_f:.4f}"
                    row[9] = f"{max_f:.4f}"
                    row[10] = str(count_i)
                    row[11] = f"{area_px_f:.1f}"
                    row[12] = f"{float(area_mm2_v):.4f}" if area_mm2_v is not None else ""
                    row[13] = rescale_unit or ""
                    channel_count, channel_values = _extract_channel_stats(stats)
                    max_channel_count = max(max_channel_count, channel_count)
                row_payloads.append((row, channel_values, _empty_measurement_csv_row()))

            for cross_item in crosshairs:
                data = get_crosshair_export_data(cross_item, dataset) if dataset else {}
                row = [
                    study_uid_short,
                    str(series_num),
                    series_desc,
                    str(z),
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    str(data.get("pixel_x", "")),
                    str(data.get("pixel_y", "")),
                    str(data.get("slice_index", "")),
                    data.get("pixel_value_str", ""),
                    _format_float(data.get("patient_x")),
                    _format_float(data.get("patient_y")),
                    _format_float(data.get("patient_z")),
                ]
                row_payloads.append((row, {}, _empty_measurement_csv_row()))

            for meas_idx, m_item in enumerate(measurements, start=1):
                # Same 21 base columns as ROI/crosshair rows (crosshair uses 10 blanks then 7 coords).
                row = [
                    study_uid_short,
                    str(series_num),
                    series_desc,
                    str(z),
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
                row_payloads.append(
                    (row, {}, serialize_measurement_for_export(m_item, meas_idx))
                )

    dynamic_headers: List[str] = []
    for c in range(max_channel_count):
        ch = c + 1
        dynamic_headers.extend(
            [
                f"Mean Ch{ch}",
                f"Std Dev Ch{ch}",
                f"Min Ch{ch}",
                f"Max Ch{ch}",
            ]
        )

    rows: List[List[str]] = []
    for base_row, channel_values, meas_part in row_payloads:
        out_row = list(base_row)
        if base_row is headers:
            out_row.extend(dynamic_headers)
            out_row.extend(MEASUREMENT_CSV_HEADERS)
            rows.append(out_row)
            continue
        for c in range(max_channel_count):
            out_row.extend(
                [
                    channel_values.get(f"mean_ch{c}", ""),
                    channel_values.get(f"std_ch{c}", ""),
                    channel_values.get(f"min_ch{c}", ""),
                    channel_values.get(f"max_ch{c}", ""),
                ]
            )
        out_row.extend(meas_part)
        rows.append([_safe_spreadsheet_value(cell) for cell in out_row])

    with open(file_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(rows)


def write_xlsx(
    file_path: str,
    collected: CollectedSeries,
    current_studies: Dict[str, Dict[str, List[Dataset]]],
    subwindow_managers: Dict[int, Dict[str, Any]],
    use_rescale: bool,
    dicom_processor: type,
) -> None:
    """
    Write one sheet per study (name from StudyDescription, max 31 chars, sanitized).
    Bold merged series headers, indented slice headers, bold ROI/crosshair headings, data rows.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as e:
        raise ImportError(
            "openpyxl is required for XLSX export. Install with: pip install openpyxl"
        ) from e

    roi_manager = None
    for idx in sorted(subwindow_managers.keys()):
        if subwindow_managers[idx].get("roi_manager"):
            roi_manager = subwindow_managers[idx]["roi_manager"]
            break
    if roi_manager is None and subwindow_managers:
        roi_manager = list(subwindow_managers.values())[0].get("roi_manager")

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook has no active sheet")
    ws.title = "ROI Statistics"[:31]

    bold_font = Font(bold=True)
    row = 1

    for (study_uid, series_uid), slice_list in collected:
        series_dict = current_studies.get(study_uid, {}).get(series_uid, [])
        if not series_dict:
            continue
        first_ds = series_dict[0]
        series_num = getattr(first_ds, "SeriesNumber", "")
        series_desc = getattr(first_ds, "SeriesDescription", "Unknown Series")
        study_desc = getattr(first_ds, "StudyDescription", "Study")[:31]
        study_desc = _sanitize_filename(study_desc) or "Study"

        ws.cell(
            row=row,
            column=1,
            value=_safe_spreadsheet_value(f"Series {series_num}: {series_desc}"),
        )
        ws.cell(row=row, column=1).font = bold_font
        row += 1

        if not slice_list:
            ws.cell(row=row, column=1, value="No annotations")
            row += 2
            continue

        for z, rois, crosshairs, measurements in slice_list:
            dataset = series_dict[z] if z < len(series_dict) else None
            ws.cell(row=row, column=1, value=f"  Slice Index (0-based): {z}")
            row += 1
            for roi_idx, roi_item in enumerate(rois, start=1):
                shape = getattr(roi_item, "shape_type", "ellipse").capitalize()
                ws.cell(row=row, column=1, value=f"  {shape} ROI {roi_idx}")
                ws.cell(row=row, column=1).font = bold_font
                row += 1
                if dataset and roi_manager:
                    stats, rescale_unit = compute_roi_statistics(
                        roi_item, dataset, use_rescale, roi_manager, dicom_processor
                    )
                    unit_str = rescale_unit or ""
                    mean_v = stats.get("mean")
                    std_v = stats.get("std")
                    min_v = stats.get("min")
                    max_v = stats.get("max")
                    count_v = stats.get("count")
                    area_px_v = stats.get("area_pixels")
                    area_mm2_v = stats.get("area_mm2")

                    mean_f = float(mean_v) if mean_v is not None else 0.0
                    std_f = float(std_v) if std_v is not None else 0.0
                    min_f = float(min_v) if min_v is not None else 0.0
                    max_f = float(max_v) if max_v is not None else 0.0
                    count_i = int(count_v) if count_v is not None else 0
                    area_px_f = float(area_px_v) if area_px_v is not None else 0.0
                    area_mm2_f = float(area_mm2_v) if area_mm2_v is not None else None

                    ws.cell(row=row, column=1, value="Mean")
                    ws.cell(row=row, column=2, value=round(mean_f, 2))
                    ws.cell(row=row, column=3, value=_safe_spreadsheet_value(unit_str))
                    row += 1
                    ws.cell(row=row, column=1, value="Std Dev")
                    ws.cell(row=row, column=2, value=round(std_f, 2))
                    ws.cell(row=row, column=3, value=_safe_spreadsheet_value(unit_str))
                    row += 1
                    ws.cell(row=row, column=1, value="Min")
                    ws.cell(row=row, column=2, value=round(min_f, 2))
                    ws.cell(row=row, column=3, value=_safe_spreadsheet_value(unit_str))
                    row += 1
                    ws.cell(row=row, column=1, value="Max")
                    ws.cell(row=row, column=2, value=round(max_f, 2))
                    ws.cell(row=row, column=3, value=_safe_spreadsheet_value(unit_str))
                    row += 1
                    ws.cell(row=row, column=1, value="Pixels")
                    ws.cell(row=row, column=2, value=count_i)
                    ws.cell(row=row, column=3, value="")
                    row += 1
                    if area_mm2_f is not None:
                        if area_mm2_f >= 100.0:
                            ws.cell(row=row, column=1, value="Area")
                            ws.cell(row=row, column=2, value=round(area_mm2_f / 100.0, 2))
                            ws.cell(row=row, column=3, value="cm²")
                        else:
                            ws.cell(row=row, column=1, value="Area")
                            ws.cell(row=row, column=2, value=round(area_mm2_f, 2))
                            ws.cell(row=row, column=3, value="mm²")
                    else:
                        ws.cell(row=row, column=1, value="Area")
                        ws.cell(row=row, column=2, value=area_px_f)
                        ws.cell(row=row, column=3, value="pixels")
                    row += 1
                    channel_count, channel_values = _extract_channel_stats(stats)
                    for c in range(channel_count):
                        ch = c + 1
                        ws.cell(row=row, column=1, value=f"Ch{ch} Mean")
                        ws.cell(
                            row=row,
                            column=2,
                            value=_safe_spreadsheet_value(channel_values.get(f"mean_ch{c}", "")),
                        )
                        ws.cell(row=row, column=3, value=_safe_spreadsheet_value(unit_str))
                        row += 1
                        ws.cell(row=row, column=1, value=f"Ch{ch} Std Dev")
                        ws.cell(
                            row=row,
                            column=2,
                            value=_safe_spreadsheet_value(channel_values.get(f"std_ch{c}", "")),
                        )
                        ws.cell(row=row, column=3, value=_safe_spreadsheet_value(unit_str))
                        row += 1
                        ws.cell(row=row, column=1, value=f"Ch{ch} Min")
                        ws.cell(
                            row=row,
                            column=2,
                            value=_safe_spreadsheet_value(channel_values.get(f"min_ch{c}", "")),
                        )
                        ws.cell(row=row, column=3, value=_safe_spreadsheet_value(unit_str))
                        row += 1
                        ws.cell(row=row, column=1, value=f"Ch{ch} Max")
                        ws.cell(
                            row=row,
                            column=2,
                            value=_safe_spreadsheet_value(channel_values.get(f"max_ch{c}", "")),
                        )
                        ws.cell(row=row, column=3, value=_safe_spreadsheet_value(unit_str))
                        row += 1
            for cross_idx, cross_item in enumerate(crosshairs, start=1):
                ws.cell(row=row, column=1, value=f"  Crosshair {cross_idx}")
                ws.cell(row=row, column=1).font = bold_font
                row += 1
                if dataset:
                    data = get_crosshair_export_data(cross_item, dataset)
                    ws.cell(row=row, column=1, value="Pixel X"); ws.cell(row=row, column=2, value=data["pixel_x"]); ws.cell(row=row, column=3, value="")
                    row += 1
                    ws.cell(row=row, column=1, value="Pixel Y"); ws.cell(row=row, column=2, value=data["pixel_y"]); ws.cell(row=row, column=3, value="")
                    row += 1
                    ws.cell(row=row, column=1, value="Slice Index (0-based)"); ws.cell(row=row, column=2, value=data["slice_index"]); ws.cell(row=row, column=3, value="")
                    row += 1
                    ws.cell(row=row, column=1, value="Pixel Value"); ws.cell(row=row, column=2, value=_safe_spreadsheet_value(data["pixel_value_str"])); ws.cell(row=row, column=3, value="")
                    row += 1
                    ws.cell(row=row, column=1, value="Patient X (mm)"); ws.cell(row=row, column=2, value=_format_float(data["patient_x"])); ws.cell(row=row, column=3, value="")
                    row += 1
                    ws.cell(row=row, column=1, value="Patient Y (mm)"); ws.cell(row=row, column=2, value=_format_float(data["patient_y"])); ws.cell(row=row, column=3, value="")
                    row += 1
                    ws.cell(row=row, column=1, value="Patient Z (mm)"); ws.cell(row=row, column=2, value=_format_float(data["patient_z"])); ws.cell(row=row, column=3, value="")
                    row += 1
            for meas_idx, m_item in enumerate(measurements, start=1):
                ws.cell(row=row, column=1, value=f"  Measurement {meas_idx}")
                ws.cell(row=row, column=1).font = bold_font
                row += 1
                for label, value in _measurement_xlsx_label_value_pairs(m_item):
                    ws.cell(row=row, column=1, value=label)
                    ws.cell(row=row, column=2, value=_safe_spreadsheet_value(value))
                    ws.cell(row=row, column=3, value="")
                    row += 1
        row += 1

    wb.save(file_path)


def run_export(
    file_path: str,
    format_key: str,
    selected_series: List[SeriesKey],
    current_studies: Dict[str, Dict[str, List[Dataset]]],
    subwindow_managers: Dict[int, Dict[str, Any]],
    use_rescale: bool,
) -> None:
    """
    Run the export: collect data and call the appropriate writer.

    format_key: "TXT", "CSV", or "XLSX".
    Raises on write error (caller should show dialog and keep dialog open).
    """
    collected = collect_roi_data(selected_series, current_studies, subwindow_managers)
    if format_key.upper() == "TXT":
        write_txt(
            file_path,
            collected,
            current_studies,
            subwindow_managers,
            use_rescale,
            DICOMProcessor,
        )
    elif format_key.upper() == "CSV":
        write_csv(
            file_path,
            collected,
            current_studies,
            subwindow_managers,
            use_rescale,
            DICOMProcessor,
        )
    elif format_key.upper() == "XLSX":
        write_xlsx(
            file_path,
            collected,
            current_studies,
            subwindow_managers,
            use_rescale,
            DICOMProcessor,
        )
    else:
        raise ValueError(f"Unsupported format: {format_key}")
