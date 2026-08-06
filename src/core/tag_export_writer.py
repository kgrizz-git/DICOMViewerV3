"""
DICOM Tag Export Writer

Provides filename generation and file writing (Excel, CSV, UTF-8 text) for DICOM tag
exports. CSV uses comma-separated fields; text exports use the same columns as tab-
separated values (``.txt``) for easy viewing and paste into spreadsheets.
This module is pure logic with no Qt dependency.
"""

import csv
from pathlib import Path
from typing import Any

from pydicom.dataset import Dataset

from core.dicom_parser import DICOMParser
from core.spreadsheet_safety import SafeCsvWriter, neutralize_spreadsheet_value
from core.tag_export_catalog import missing_tag_export_display_fields

_TAG_EXPORT_HEADERS = ['Instance', 'Tag Number', 'Name', 'Value']


def _tag_number_for_export(tag_str: str, tag_data: dict[str, Any]) -> str:
    """The identifier written to the "Tag Number" column.

    ``tag_data['tag']`` is the *canonical* tag, which for a nested sequence leaf
    discards the path — two leaves under different sequences would both export as
    ``(0008, 0104)``, re-creating in the spreadsheet exactly the ambiguity path keys
    exist to remove. So nested rows (``depth > 0``) export their full path key instead.

    Root-level rows export the canonical tag, which is identical to their key.
    """
    if tag_data.get('depth', 0):
        return tag_str
    return tag_data.get('tag', tag_str)


def _tag_value_for_export(value: Any) -> str:
    """Return a parsed tag value in the single-cell form used by every exporter."""
    if isinstance(value, list):
        return ', '.join(str(item) for item in value)
    return str(value)


def _present_tag_export_row(
    instance_id: str, tag_str: str, tag_data: dict[str, Any]
) -> list[str]:
    """Build a row for a selected tag present in a parsed dataset."""
    return [
        instance_id,
        _tag_number_for_export(tag_str, tag_data),
        tag_data.get('name', ''),
        _tag_value_for_export(tag_data.get('value', '')),
    ]


def _missing_tag_export_row(instance_id: str, tag_str: str) -> list[str]:
    """Build the optional empty row for a selected tag absent from a dataset."""
    tag_num, tag_name = missing_tag_export_display_fields(tag_str)
    return [instance_id, tag_num or tag_str, tag_name, '']


def _selected_tag_export_rows(
    instance_id: str,
    selected_tags: list[str],
    all_tags: dict[str, Any],
    include_missing_selected_tags: bool,
) -> list[list[str]]:
    """Build the rows for selected tags from one parsed dataset."""
    rows: list[list[str]] = []
    for tag_str in selected_tags:
        if tag_str in all_tags:
            rows.append(_present_tag_export_row(instance_id, tag_str, all_tags[tag_str]))
        elif include_missing_selected_tags:
            rows.append(_missing_tag_export_row(instance_id, tag_str))
    return rows


def _parsed_tags_for_export(
    dataset: Dataset, include_private: bool, include_sequences: bool
) -> dict[str, Any]:
    """Read a dataset through the parser with the export's selected options."""
    return DICOMParser(dataset).get_all_tags(
        include_private=include_private, include_sequences=include_sequences
    )


def _series_tag_partitions(
    variation_analysis: dict[str, dict[str, list[str]]],
    series_uid: str,
    selected_tags: list[str],
) -> tuple[list[str], list[str]]:
    """Return a series' varying and constant selected-tag partitions."""
    analysis = variation_analysis.get(
        series_uid, {'varying_tags': [], 'constant_tags': selected_tags}
    )
    return analysis['varying_tags'], analysis['constant_tags']


def _constant_tag_export_rows(
    datasets: list[Dataset],
    instance_indices: list[int],
    constant_tags: list[str],
    include_private: bool,
    include_missing_selected_tags: bool,
    include_sequences: bool,
) -> list[list[str]]:
    """Build one ``All`` row for each constant tag in a selected series."""
    if not constant_tags or not instance_indices:
        return []

    first_instance_idx = instance_indices[0]
    if first_instance_idx >= len(datasets):
        return []

    all_tags = _parsed_tags_for_export(
        datasets[first_instance_idx], include_private, include_sequences
    )
    return _selected_tag_export_rows(
        'All', constant_tags, all_tags, include_missing_selected_tags
    )


def _varying_tag_export_rows(
    datasets: list[Dataset],
    instance_indices: list[int],
    varying_tags: list[str],
    include_private: bool,
    include_missing_selected_tags: bool,
    include_sequences: bool,
) -> list[list[str]]:
    """Build per-selected-instance rows for a series' varying tags."""
    rows: list[list[str]] = []
    for instance_idx in instance_indices:
        if instance_idx >= len(datasets):
            continue

        dataset = datasets[instance_idx]
        instance_num = getattr(dataset, 'InstanceNumber', None)
        instance_id = (
            f"Instance {instance_num}"
            if instance_num is not None
            else f"Instance {instance_idx + 1}"
        )
        all_tags = _parsed_tags_for_export(dataset, include_private, include_sequences)
        rows.extend(
            _selected_tag_export_rows(
                instance_id, varying_tags, all_tags, include_missing_selected_tags
            )
        )
    return rows


def _series_tag_export_rows(
    series_uid: str,
    instance_indices: list[int],
    studies: dict[str, dict[str, list[Dataset]]],
    study_uid: str,
    variation_analysis: dict[str, dict[str, list[str]]],
    selected_tags: list[str],
    include_private: bool,
    include_missing_selected_tags: bool,
    include_sequences: bool,
) -> list[list[str]]:
    """Build the heading, tag data, and trailing blank row for one series."""
    datasets = studies[study_uid][series_uid]
    if not datasets:
        return []

    first_ds = datasets[0]
    series_num = getattr(first_ds, 'SeriesNumber', '')
    series_desc = getattr(first_ds, 'SeriesDescription', 'Unknown')
    rows = [[f"Series {series_num}: {series_desc}", '', '', '']]
    varying_tags, constant_tags = _series_tag_partitions(
        variation_analysis, series_uid, selected_tags
    )
    rows.extend(
        _constant_tag_export_rows(
            datasets,
            instance_indices,
            constant_tags,
            include_private,
            include_missing_selected_tags,
            include_sequences,
        )
    )
    rows.extend(
        _varying_tag_export_rows(
            datasets,
            instance_indices,
            varying_tags,
            include_private,
            include_missing_selected_tags,
            include_sequences,
        )
    )
    rows.append([])
    return rows


def _tag_export_study_rows(
    study_uid: str,
    series_dict: dict[str, list[int]],
    studies: dict[str, dict[str, list[Dataset]]],
    variation_analysis: dict[str, dict[str, list[str]]],
    selected_tags: list[str],
    include_private: bool,
    include_missing_selected_tags: bool,
    include_sequences: bool,
) -> list[list[list[str]]]:
    """Build the export rows for each non-empty selected series in a study."""
    return [
        rows
        for series_uid, instance_indices in series_dict.items()
        if (
            rows := _series_tag_export_rows(
                series_uid,
                instance_indices,
                studies,
                study_uid,
                variation_analysis,
                selected_tags,
                include_private,
                include_missing_selected_tags,
                include_sequences,
            )
        )
    ]


def _configure_excel_worksheet(ws: Any, font_factory: Any) -> None:
    """Write the fixed header and column widths for an export worksheet."""
    for column, value in enumerate(_TAG_EXPORT_HEADERS, start=1):
        cell = ws.cell(row=1, column=column, value=value)
        cell.font = font_factory(bold=True)

    for column, width in zip(('A', 'B', 'C', 'D'), (15, 15, 40, 60), strict=True):
        ws.column_dimensions[column].width = width


def _write_excel_series_heading(ws: Any, row: int, heading: str, font_factory: Any) -> int:
    """Write and style an XLSX series heading, returning the next row number."""
    ws[f'A{row}'] = neutralize_spreadsheet_value(heading)
    ws[f'A{row}'].font = font_factory(bold=True, italic=True)
    ws.merge_cells(f'A{row}:D{row}')
    return row + 1


def _write_excel_tag_rows(ws: Any, row: int, rows: list[list[str]]) -> int:
    """Write data or blank tag rows to an XLSX worksheet and return the next row."""
    for values in rows:
        for column, value in enumerate(values, start=1):
            ws.cell(row=row, column=column, value=neutralize_spreadsheet_value(value))
        row += 1
    return row


def _write_excel_study_rows(
    ws: Any,
    study_uid: str,
    series_dict: dict[str, list[int]],
    studies: dict[str, dict[str, list[Dataset]]],
    variation_analysis: dict[str, dict[str, list[str]]],
    selected_tags: list[str],
    include_private: bool,
    include_missing_selected_tags: bool,
    include_sequences: bool,
    font_factory: Any,
) -> None:
    """Render one study's shared tag-export rows with XLSX-specific styling."""
    row = 2
    for series_rows in _tag_export_study_rows(
        study_uid,
        series_dict,
        studies,
        variation_analysis,
        selected_tags,
        include_private,
        include_missing_selected_tags,
        include_sequences,
    ):
        row = _write_excel_series_heading(ws, row, series_rows[0][0], font_factory)
        row = _write_excel_tag_rows(ws, row, series_rows[1:])


def generate_default_filename(
    studies: dict[str, dict[str, list[Dataset]]],
    selected_series: dict[str, dict[str, list[int]]],
) -> str:
    """Generate a default filename for tag export (Excel format)."""
    first_study_uid = next(iter(selected_series.keys()))
    first_series_uid = next(iter(selected_series[first_study_uid].keys()))
    first_instance_idx = selected_series[first_study_uid][first_series_uid][0]
    first_dataset = studies[first_study_uid][first_series_uid][first_instance_idx]

    modality = getattr(first_dataset, 'Modality', 'Unknown')
    accession = getattr(first_dataset, 'AccessionNumber', 'Unknown')
    return f"{modality} DICOM Tag Export {accession}.xlsx"


def write_excel_file(
    file_path: str,
    variation_analysis: dict[str, dict[str, list[str]]],
    studies: dict[str, dict[str, list[Dataset]]],
    selected_series: dict[str, dict[str, list[int]]],
    selected_tags: list[str],
    include_private: bool,
    include_missing_selected_tags: bool = True,
    include_sequences: bool = False,
) -> None:
    """Write selected tags to an Excel file with per-instance export for varying tags.

    ``include_sequences`` (default False) mirrors
    :meth:`~core.dicom_parser.DICOMParser.get_all_tags`'s flag: off reproduces the
    historical scalar-only export byte-for-byte; on lets a selected SQ tag resolve
    to its single-cell summary value (item count, or the compact code summary for
    code sequences) instead of always falling through to the missing-tag row.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        raise ImportError(  # noqa: B904 - the ModuleNotFoundError adds nothing here
            "openpyxl library is required for Excel export. "
            "Install it with: pip install openpyxl"
        )

    wb = Workbook()
    default_ws = wb.active
    if default_ws is not None:
        wb.remove(default_ws)  # Remove default sheet

    # Create one sheet per study.
    for study_uid, series_dict in selected_series.items():
        first_series_uid = next(iter(series_dict.keys()))
        first_instance_idx = series_dict[first_series_uid][0]
        first_dataset = studies[study_uid][first_series_uid][first_instance_idx]
        study_desc = getattr(first_dataset, 'StudyDescription', 'Study')
        sheet_name = study_desc[:31].replace('/', '-').replace('\\', '-').replace(':', '-')
        ws = wb.create_sheet(title=sheet_name)
        _configure_excel_worksheet(ws, Font)
        _write_excel_study_rows(
            ws,
            study_uid,
            series_dict,
            studies,
            variation_analysis,
            selected_tags,
            include_private,
            include_missing_selected_tags,
            include_sequences,
            Font,
        )

    wb.save(file_path)


def _write_tag_export_sheet_rows(
    writer: Any,
    study_uid: str,
    series_dict: dict[str, list[int]],
    studies: dict[str, dict[str, list[Dataset]]],
    variation_analysis: dict[str, dict[str, list[str]]],
    selected_tags: list[str],
    include_private: bool,
    include_missing_selected_tags: bool,
    include_sequences: bool = False,
) -> None:
    """Write one study's tag rows using *writer* (``csv.writer`` with any dialect).

    ``include_sequences`` (default False) is forwarded to
    :meth:`~core.dicom_parser.DICOMParser.get_all_tags` — see :func:`write_excel_file`.
    """
    writer.writerow(_TAG_EXPORT_HEADERS)
    for series_rows in _tag_export_study_rows(
        study_uid,
        series_dict,
        studies,
        variation_analysis,
        selected_tags,
        include_private,
        include_missing_selected_tags,
        include_sequences,
    ):
        writer.writerows(series_rows)


def write_csv_files(
    base_file_path: str,
    variation_analysis: dict[str, dict[str, list[str]]],
    studies: dict[str, dict[str, list[Dataset]]],
    selected_series: dict[str, dict[str, list[int]]],
    selected_tags: list[str],
    include_private: bool,
    include_missing_selected_tags: bool = True,
    include_sequences: bool = False,
) -> list[Path]:
    """
    Write selected tags to CSV files (one per study) with per-instance export for varying tags.

    ``include_sequences`` (default False) is forwarded to
    :meth:`~core.dicom_parser.DICOMParser.get_all_tags` — see :func:`write_excel_file`.

    Returns:
        List of created file paths.
    """
    base_path = Path(base_file_path)
    base_name = base_path.stem
    base_dir = base_path.parent
    exported_files: list[Path] = []

    for study_uid, series_dict in selected_series.items():
        first_series_uid = next(iter(series_dict.keys()))
        first_instance_idx = series_dict[first_series_uid][0]
        first_dataset = studies[study_uid][first_series_uid][first_instance_idx]
        study_desc = getattr(first_dataset, 'StudyDescription', 'Study')
        safe_study_desc = study_desc.replace('/', '-').replace('\\', '-').replace(':', '-')[:50]

        if len(selected_series) > 1:
            csv_filename = f"{base_name}_{safe_study_desc}.csv"
        else:
            csv_filename = f"{base_name}.csv"

        csv_path = base_dir / csv_filename

        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = SafeCsvWriter(csv.writer(csvfile))
            _write_tag_export_sheet_rows(
                writer,
                study_uid,
                series_dict,
                studies,
                variation_analysis,
                selected_tags,
                include_private,
                include_missing_selected_tags,
                include_sequences,
            )

        exported_files.append(csv_path)

    return exported_files


def write_txt_files(
    base_file_path: str,
    variation_analysis: dict[str, dict[str, list[str]]],
    studies: dict[str, dict[str, list[Dataset]]],
    selected_series: dict[str, dict[str, list[int]]],
    selected_tags: list[str],
    include_private: bool,
    include_missing_selected_tags: bool = True,
    include_sequences: bool = False,
) -> list[Path]:
    """
    Write selected tags to UTF-8 text files (one per study), tab-separated columns,
    same row layout as :func:`write_csv_files`.

    ``include_sequences`` (default False) is forwarded to
    :meth:`~core.dicom_parser.DICOMParser.get_all_tags` — see :func:`write_excel_file`.

    Returns:
        List of created file paths (``.txt``).
    """
    base_path = Path(base_file_path)
    base_name = base_path.stem
    base_dir = base_path.parent
    exported_files: list[Path] = []

    for study_uid, series_dict in selected_series.items():
        first_series_uid = next(iter(series_dict.keys()))
        first_instance_idx = series_dict[first_series_uid][0]
        first_dataset = studies[study_uid][first_series_uid][first_instance_idx]
        study_desc = getattr(first_dataset, 'StudyDescription', 'Study')
        safe_study_desc = study_desc.replace('/', '-').replace('\\', '-').replace(':', '-')[:50]

        if len(selected_series) > 1:
            txt_filename = f"{base_name}_{safe_study_desc}.txt"
        else:
            txt_filename = f"{base_name}.txt"

        txt_path = base_dir / txt_filename

        with open(txt_path, 'w', newline='', encoding='utf-8') as txtfile:
            writer = SafeCsvWriter(csv.writer(txtfile, delimiter='\t'))
            _write_tag_export_sheet_rows(
                writer,
                study_uid,
                series_dict,
                studies,
                variation_analysis,
                selected_tags,
                include_private,
                include_missing_selected_tags,
                include_sequences,
            )

        exported_files.append(txt_path)

    return exported_files
